#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
УНИВЕРСАЛЬНОЕ ПРИЛОЖЕНИЕ ДЛЯ РАБОТЫ С ФГИС АРШИН
Версия 1.2 - С атрибутивным поиском и динамическим rate limiting

Изменения v1.2:
- Атрибутивный поиск (mi_number=, mit_number=) вместо search=
- Динамический rate limiting (0.1-0.5 сек) с адаптацией к HTTP 429
- Исправлено кодирование: пробелы → ?, кириллица не кодируется
- Парсинг тела ошибок API (status, message, time, requestId)
- sort=mi_number+desc по умолчанию для VRI
- Сохранение ВСЕХ колонок из исходного Excel
- Одиночный поиск через mi_number с пагинацией
"""

import os
import sys
import json
import csv
import sqlite3
import asyncio
import aiohttp
import requests
import threading
import time
import logging
import re
import webbrowser
from datetime import datetime
from typing import Dict, List, Set, Optional, Tuple, Any
from collections import defaultdict
from dataclasses import dataclass, asdict, field
from urllib.parse import quote, urlencode
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tkinter.font import Font

# Попытка импорта дополнительных библиотек
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logging.warning("pandas/openpyxl не установлены - экспорт в Excel недоступен")

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

# ============== РАСШИРЕННОЕ ЛОГИРОВАНИЕ ==============
def setup_logging(debug_mode: bool = False):
    """Настройка расширенного логирования"""
    log_dir = "logs"
    os.makedirs(log_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(log_dir, f"arshin_app_{timestamp}.log")

    log_format = logging.Formatter(
        '%(asctime)s.%(msecs)03d | %(levelname)-8s | %(name)s | %(funcName)s:%(lineno)d | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setFormatter(log_format)
    file_handler.setLevel(logging.DEBUG)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    console_handler.setLevel(logging.DEBUG if debug_mode else logging.INFO)

    root_logger = logging.getLogger()
    root_logger.setLevel(logging.DEBUG)
    root_logger.addHandler(file_handler)
    root_logger.addHandler(console_handler)

    logger = logging.getLogger(__name__)
    logger.info(f"Логирование инициализировано: {log_file}")
    logger.debug(f"Режим отладки: {'ВКЛЮЧЕН' if debug_mode else 'ВЫКЛЮЧЕН'}")

    return logger, log_file

logger, LOG_FILE = setup_logging(debug_mode=True)

# ============== КОНФИГУРАЦИЯ ==============
class Config:
    """Настройки приложения с загрузкой из файлов"""

    # API ФГИС АРШИН (согласно документации v.2.2)
    API_BASE_URL = "https://fgis.gost.ru/fundmetrology/eapi"
    API_VRI = f"{API_BASE_URL}/vri"
    API_MIT = f"{API_BASE_URL}/mit"

    # Параметры запросов (согласно спецификации API v.2.2)
    MAX_ROWS_PER_REQUEST = 100
    DEFAULT_ROWS = 50
    REQUEST_TIMEOUT = 30
    RETRY_ATTEMPTS = 3
    RETRY_DELAY = 2

    # Динамический rate limiting
    RATE_LIMIT_MIN = 0.1       # Минимальная пауза (сек)
    RATE_LIMIT_MAX = 0.5       # Максимальная пауза (сек)
    RATE_LIMIT_START = 0.1     # Стартовая пауза
    RATE_LIMIT_STEP = 0.1      # Шаг изменения
    RATE_LIMIT_429_COOLDOWN = 10   # Секунды без ошибок для ускорения
    RATE_LIMIT_ERROR_WINDOW = 60   # Секунды без ошибок для замедления

    # База данных
    DB_PATH = "arshin_data.db"

    # Пути к файлам
    EXPORT_DIR = "exports"
    CONFIG_DIR = "config"

    # Файлы конфигурации
    EXACT_QUERIES_FILE = os.path.join(CONFIG_DIR, "exact_queries.csv")
    MANUFACTURERS_FILE = os.path.join(CONFIG_DIR, "manufacturers.csv")

    # Годы для фильтрации
    YEARS = list(range(2010, 2027))

    # Размеры окна (оптимизировано для 1200x700)
    WINDOW_WIDTH = 1180
    WINDOW_HEIGHT = 680

    @classmethod
    def load_exact_queries(cls) -> List[str]:
        """Загрузка точных запросов из файла"""
        queries = []
        if os.path.exists(cls.EXACT_QUERIES_FILE):
            try:
                with open(cls.EXACT_QUERIES_FILE, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if 'query' in row and row['query'].strip():
                            queries.append(row['query'].strip())
                logger.info(f"Загружено {len(queries)} точных запросов из {cls.EXACT_QUERIES_FILE}")
            except Exception as e:
                logger.error(f"Ошибка загрузки точных запросов: {e}")
        else:
            logger.warning(f"Файл точных запросов не найден: {cls.EXACT_QUERIES_FILE}")
        return queries

    @classmethod
    def load_manufacturers(cls) -> Dict[str, str]:
        """Загрузка словаря производителей из файла"""
        manufacturers = {}
        if os.path.exists(cls.MANUFACTURERS_FILE):
            try:
                with open(cls.MANUFACTURERS_FILE, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if 'keyword' in row and 'manufacturer' in row:
                            manufacturers[row['keyword'].lower().strip()] = row['manufacturer'].strip()
                logger.info(f"Загружено {len(manufacturers)} правил для производителей")
            except Exception as e:
                logger.error(f"Ошибка загрузки производителей: {e}")
        else:
            logger.warning(f"Файл производителей не найден: {cls.MANUFACTURERS_FILE}")
        return manufacturers

EXACT_QUERIES = Config.load_exact_queries()
MANUFACTURERS_RULES = Config.load_manufacturers()

# ============== МОДЕЛИ ДАННЫХ ==============
@dataclass
class VerificationRecord:
    """Запись о поверке СИ с полной служебной информацией"""

    # Основные данные из API (список VRI)
    vri_id: str = ""
    mit_number: str = ""
    mit_title: str = ""
    mit_notation: str = ""
    mi_modification: str = ""
    mi_number: str = ""
    verification_date: str = ""
    valid_date: str = ""
    applicability: bool = True
    org_title: str = ""
    result_docnum: str = ""
    sticker_num: str = ""

    # Служебные поля
    manufacturer: str = ""
    collected_at: str = ""

    # Связи с Excel
    search_query: str = ""
    row_index: int = 0
    id_pu: str = ""

    # Все дополнительные колонки из Excel (динамические)
    extra_fields: Dict[str, str] = field(default_factory=dict)

    @staticmethod
    def generate_url(vri_id: str) -> str:
        """Генерация URL записи в системе ФГИС АРШИН"""
        return f"https://fgis.gost.ru/fundmetrology/cm/erts/?id={vri_id}"

    @property
    def record_url(self) -> str:
        return self.generate_url(self.vri_id)

    @classmethod
    def from_api_response(cls, data: dict, search_query: str = "",
                          row_index: int = 0, id_pu: str = "",
                          original_data: dict = None,
                          extra_fields: dict = None) -> 'VerificationRecord':
        """Создание записи из ответа API"""
        original = original_data or {}

        record = cls(
            vri_id=data.get('vri_id', ''),
            mit_number=data.get('mit_number', ''),
            mit_title=data.get('mit_title', ''),
            mit_notation=data.get('mit_notation', ''),
            mi_modification=data.get('mi_modification', ''),
            mi_number=data.get('mi_number', ''),
            verification_date=data.get('verification_date', ''),
            valid_date=data.get('valid_date', ''),
            applicability=data.get('applicability', True),
            org_title=data.get('org_title', ''),
            result_docnum=data.get('result_docnum', ''),
            sticker_num=data.get('sticker_num', ''),
            search_query=search_query,
            row_index=row_index,
            id_pu=id_pu,
            extra_fields=extra_fields or {},
            collected_at=datetime.now().isoformat()
        )

        record.manufacturer = cls._detect_manufacturer(record.mit_title)
        logger.debug(f"Создана запись: vri_id={record.vri_id}, query={search_query}")
        return record

    @staticmethod
    def _detect_manufacturer(title: str) -> str:
        """Определение производителя по названию"""
        if not title:
            return 'Другие'

        title_lower = title.lower()

        for keyword, manufacturer in MANUFACTURERS_RULES.items():
            if keyword in title_lower:
                return manufacturer

        fallback_rules = {
            'меркурий': 'Меркурий',
            'нева': 'Нева',
            'энергомера': 'Энергомера',
        }

        for key, value in fallback_rules.items():
            if key in title_lower:
                return value

        return 'Другие'

    @staticmethod
    def is_electric_meter(title: str) -> bool:
        """
        Проверка: относится ли запись к приборам учёта электрической энергии
        """
        if not title:
            return False

        title_lower = title.lower()

        # Исключения (не электрическая энергия)
        exclude_keywords = [
            'воды', 'водомер', 'водосчетчик', 'холодной воды', 'горячей воды',
            'счетчик воды', 'крыльчатые',
            'газа', 'газ', 'газосчетчик', 'газовые', 'бытовые газовые', 'диафрагменные',
            'счетчик газа',
            'тепл', 'теплопотребление', 'теплопотребления', 'тепловычислитель',
            'вычислитель количества теплоты', 'распределения теплопотребления',
            'термометры медицинские', 'термометр медицинский', 'медицинские максимальные',
            'ртутные стеклянные медицинские', 'максимальные стеклянные ртутные',
            'весы', 'взвешивания', 'для новорожденных', 'почтовые электронные',
            'рычажные настольные',
            'манометры', 'вакуумметры', 'мановакуумметры', 'напоромеры', 'тягомеры',
            'тягонапоромеры', 'обм',
            'термометры ртутные', 'термометры стеклянные', 'термометры лабораторные',
            'термопреобразователи', 'термометры сопротивления',
            'сигнализаторы', 'метана', 'с внешним сенсором',
            'штангенглубиномеры', 'нутромеры', 'ареометры', 'гигрометры',
            'психрометрические', 'мегаомметры', 'приемники-ловушки',
            'счетчики времени наработки', 'бутирометры', 'поверка си массы',
            'расхода электромагнитные', 'преобразователи расхода',
            'устройства для распределения', 'для статического взвешивания',
            'транcформаторы тока', 'трансформаторы тока до 0,4 кв',
        ]

        for keyword in exclude_keywords:
            if keyword in title_lower:
                return False

        # Включения (электрическая энергия)
        include_keywords = [
            # Счетчики электрической энергии — расширенный список
            'электрической энергии', 'электроэнергии', 'электросчетчик',
            'счетчик электрической', 'счетчик электроэнергии',
            'активной электроэнергии', 'реактивной электроэнергии',
            'активной энергии', 'реактивной энергии',  # без "электро" — тоже электросчётчики
            'электрические активной', 'электрические реактивной',
            'счетчики электрические',
            'однофазный', 'трехфазный', 'статический', 'индукционный',
            'меркурий', 'нева', 'энергомера', 'матрица', 'альфа', 'милур',
            'псч', 'сэт', 'цэ', 'се 101', 'се 102', 'се 201', 'се 301',
            'меркурий 20', 'меркурий 23', 'нева 10', 'нева 30',
            'трансформатор тока измерительный', 'трансформатор тока 0,4 кв',
        ]

        for keyword in include_keywords:
            if keyword in title_lower:
                return True

        return False

    def to_dict(self) -> dict:
        """Преобразование в словарь с полной служебной информацией"""
        result = asdict(self)
        result['record_url'] = self.record_url
        # Распаковка extra_fields в верхний уровень
        result.update(self.extra_fields)
        return result


@dataclass
class SearchQuery:
    """Поисковый запрос с метаданными"""
    search_term: str
    query_type: str  # 'mi_number', 'mit_number', 'search'
    year: Optional[int] = None
    source: str = ""
    row_index: int = 0


# ============== ДИНАМИЧЕСКИЙ RATE LIMITER ==============
class DynamicRateLimiter:
    """
    Адаптивный контроллер частоты запросов к API.

    Логика:
    - Старт с RATE_LIMIT_START (0.1 сек)
    - При получении 429: увеличивать интервал на STEP до MAX
    - Если 429 получены 2+ подряд: сразу jump до MAX
    - Если RATE_LIMIT_429_COOLDOWN (10 сек) без ошибок: уменьшать на STEP до MIN
    - Если RATE_LIMIT_ERROR_WINDOW (60 сек) без ошибок: уменьшать на STEP до MIN
    """

    def __init__(self):
        self.current_delay = Config.RATE_LIMIT_START
        self.consecutive_429 = 0
        self.last_429_time = 0.0
        self.last_request_time = 0.0
        self.lock = threading.Lock()
        logger.info(f"DynamicRateLimiter: старт={Config.RATE_LIMIT_START}, "
                     f"min={Config.RATE_LIMIT_MIN}, max={Config.RATE_LIMIT_MAX}")

    def wait(self):
        """Ожидание перед запросом"""
        with self.lock:
            now = time.time()
            elapsed = now - self.last_request_time
            wait_time = max(0, self.current_delay - elapsed)

        if wait_time > 0:
            logger.debug(f"Rate limiter: ожидание {wait_time:.2f}с (текущий delay={self.current_delay:.2f})")
            time.sleep(wait_time)

        with self.lock:
            self.last_request_time = time.time()

    def on_429(self):
        """Обработка получения HTTP 429"""
        with self.lock:
            self.consecutive_429 += 1
            self.last_429_time = time.time()

            # Если 2+ подряд 429 — сразу максимум
            if self.consecutive_429 >= 2:
                self.current_delay = Config.RATE_LIMIT_MAX
                logger.warning(f"429 x{self.consecutive_429} подряд — интервал увеличен до MAX ({self.current_delay:.2f}с)")
            else:
                self.current_delay = min(
                    Config.RATE_LIMIT_MAX,
                    self.current_delay + Config.RATE_LIMIT_STEP
                )
                logger.warning(f"Получен 429 — интервал увеличен до {self.current_delay:.2f}с")

    def on_success(self):
        """Обработка успешного запроса"""
        with self.lock:
            now = time.time()

            # Сброс счётчика подрядных 429
            self.consecutive_429 = 0

            # Проверка: прошло ли достаточно времени без ошибок
            time_since_429 = now - self.last_429_time
            if time_since_429 >= Config.RATE_LIMIT_ERROR_WINDOW and self.current_delay > Config.RATE_LIMIT_MIN:
                self.current_delay = max(
                    Config.RATE_LIMIT_MIN,
                    self.current_delay - Config.RATE_LIMIT_STEP
                )
                logger.debug(f"60 сек без 429 — интервал уменьшен до {self.current_delay:.2f}с")
            elif time_since_429 >= Config.RATE_LIMIT_429_COOLDOWN and self.current_delay > Config.RATE_LIMIT_MIN:
                # Более мягкое условие: 10 сек без ошибок тоже уменьшаем
                self.current_delay = max(
                    Config.RATE_LIMIT_MIN,
                    self.current_delay - Config.RATE_LIMIT_STEP
                )
                logger.debug(f"10 сек без 429 — интервал уменьшен до {self.current_delay:.2f}с")

    @property
    def delay(self) -> float:
        with self.lock:
            return self.current_delay


# ============== БАЗА ДАННЫХ ==============
class Database:
    """Работа с SQLite базой данных"""

    def __init__(self, db_path: str = Config.DB_PATH):
        self.db_path = db_path
        logger.debug(f"Инициализация БД: {db_path}")
        self._init_db()

    def _init_db(self):
        """Инициализация структуры БД"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute('''
                CREATE TABLE IF NOT EXISTS verification_records (
                    vri_id TEXT PRIMARY KEY,
                    mit_number TEXT,
                    mit_title TEXT,
                    mit_notation TEXT,
                    mi_modification TEXT,
                    mi_number TEXT,
                    verification_date TEXT,
                    valid_date TEXT,
                    applicability INTEGER,
                    org_title TEXT,
                    result_docnum TEXT,
                    sticker_num TEXT,
                    manufacturer TEXT,
                    search_query TEXT,
                    row_index INTEGER,
                    id_pu TEXT,
                    extra_fields TEXT,
                    collected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            conn.execute('''
                CREATE TABLE IF NOT EXISTS measurement_types (
                    mit_id TEXT PRIMARY KEY,
                    number TEXT,
                    title TEXT,
                    notation TEXT,
                    status TEXT,
                    manufacturer TEXT,
                    country TEXT,
                    collected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            conn.execute('''
                CREATE TABLE IF NOT EXISTS search_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    query TEXT,
                    query_type TEXT,
                    year INTEGER,
                    results_count INTEGER,
                    search_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    duration_ms INTEGER,
                    source TEXT
                )
            ''')

            # Индексы
            conn.execute('CREATE INDEX IF NOT EXISTS idx_applicability ON verification_records(applicability)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_manufacturer ON verification_records(manufacturer)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_verification_date ON verification_records(verification_date)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_search_query ON verification_records(search_query)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_mi_number ON verification_records(mi_number)')

            logger.info("БД инициализирована успешно")

    def save_records(self, records: List[VerificationRecord]) -> int:
        """Пакетное сохранение записей с проверкой на полные дубли"""
        if not records:
            return 0

        saved = 0
        duplicates = 0
        errors = 0

        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()

            for record in records:
                try:
                    extra_json = json.dumps(record.extra_fields, ensure_ascii=False) if record.extra_fields else ''

                    cursor.execute('''
                        SELECT COUNT(*) FROM verification_records
                        WHERE vri_id = ?
                          AND mit_number = ?
                          AND mi_number = ?
                          AND verification_date = ?
                          AND valid_date = ?
                          AND applicability = ?
                          AND search_query = ?
                          AND row_index = ?
                          AND id_pu = ?
                    ''', (
                        record.vri_id, record.mit_number, record.mi_number,
                        record.verification_date, record.valid_date,
                        1 if record.applicability else 0,
                        record.search_query, record.row_index, record.id_pu
                    ))

                    if cursor.fetchone()[0] > 0:
                        duplicates += 1
                        logger.debug(f"Дубль: vri_id={record.vri_id}")
                    else:
                        cursor.execute('''
                            INSERT OR IGNORE INTO verification_records
                            (vri_id, mit_number, mit_title, mit_notation, mi_modification,
                             mi_number, verification_date, valid_date, applicability,
                             org_title, result_docnum, sticker_num, manufacturer,
                             search_query, row_index, id_pu, extra_fields)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            record.vri_id, record.mit_number, record.mit_title,
                            record.mit_notation, record.mi_modification, record.mi_number,
                            record.verification_date, record.valid_date,
                            1 if record.applicability else 0,
                            record.org_title, record.result_docnum, record.sticker_num,
                            record.manufacturer,
                            record.search_query, record.row_index, record.id_pu,
                            extra_json
                        ))

                        if cursor.rowcount > 0:
                            saved += 1

                except Exception as e:
                    errors += 1
                    logger.error(f"Ошибка сохранения {record.vri_id}: {e}")

            conn.commit()

        logger.info(f"Сохранено: {saved} новых, {duplicates} дубликатов, {errors} ошибок")
        return saved

    def get_existing_ids(self) -> Set[str]:
        """Получить все сохранённые ID"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute('SELECT vri_id FROM verification_records')
            ids = {row[0] for row in cursor.fetchall()}
            logger.debug(f"В БД найдено {len(ids)} существующих записей")
            return ids

    def get_stats(self) -> Dict:
        """Получить детальную статистику по базе"""
        with sqlite3.connect(self.db_path) as conn:
            total = conn.execute('SELECT COUNT(*) FROM verification_records').fetchone()[0]

            if total == 0:
                return {
                    'total': 0, 'by_year': {}, 'by_manufacturer': {},
                    'applicable': 0, 'inapplicable': 0
                }

            by_year = dict(conn.execute(
                "SELECT strftime('%Y', verification_date) as year, COUNT(*) "
                "FROM verification_records WHERE verification_date != '' "
                "GROUP BY year ORDER BY year"
            ).fetchall())

            by_manufacturer = dict(conn.execute(
                'SELECT manufacturer, COUNT(*) FROM verification_records GROUP BY manufacturer ORDER BY COUNT(*) DESC'
            ).fetchall())

            applicable = conn.execute('SELECT COUNT(*) FROM verification_records WHERE applicability = 1').fetchone()[0]

            return {
                'total': total,
                'by_year': by_year,
                'by_manufacturer': by_manufacturer,
                'applicable': applicable,
                'inapplicable': total - applicable
            }

    def search(self, query: str = "", query_type: str = 'substring',
               year: Optional[int] = None, manufacturer: str = "",
               limit: int = 500, offset: int = 0) -> List[Dict]:
        """Расширенный поиск в локальной БД"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            params = []

            sql = 'SELECT * FROM verification_records WHERE 1=1'

            if query:
                if query_type == 'exact':
                    sql += ' AND (mit_number = ? OR mi_number = ? OR vri_id = ?)'
                    params.extend([query, query, query])
                else:
                    sql += ' AND (mit_number LIKE ? OR mi_number LIKE ? OR mit_title LIKE ? OR org_title LIKE ?)'
                    pattern = f'%{query}%'
                    params.extend([pattern, pattern, pattern, pattern])

            if year:
                sql += " AND strftime('%Y', verification_date) = ?"
                params.append(str(year))

            if manufacturer:
                sql += ' AND manufacturer = ?'
                params.append(manufacturer)

            sql += ' ORDER BY verification_date DESC LIMIT ? OFFSET ?'
            params.extend([limit, offset])

            cursor = conn.execute(sql, params)
            results = [dict(row) for row in cursor.fetchall()]
            logger.debug(f"Найдено {len(results)} записей")
            return results

    def log_search(self, query: str, query_type: str, year: Optional[int],
                   results_count: int, duration_ms: int, source: str = "manual"):
        """Логирование поиска в историю"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                'INSERT INTO search_history (query, query_type, year, results_count, duration_ms, source) VALUES (?, ?, ?, ?, ?, ?)',
                (query, query_type, year, results_count, duration_ms, source)
            )
            conn.commit()

    def get_search_history(self, limit: int = 50) -> List[Dict]:
        """Получение истории поиска"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(
                'SELECT * FROM search_history ORDER BY search_date DESC LIMIT ?',
                (limit,)
            )
            return [dict(row) for row in cursor.fetchall()]

    def clean_duplicates(self) -> int:
        """Удаление дубликатов"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute('''
                DELETE FROM verification_records
                WHERE rowid NOT IN (
                    SELECT MIN(rowid) FROM verification_records GROUP BY vri_id
                )
            ''')
            conn.commit()
            deleted = cursor.rowcount
        logger.info(f"Удалено дубликатов: {deleted}")
        return deleted


# ============== КЛИЕНТ API (v.2.2 compliant) ==============
class APIClient:
    """
    Клиент для работы с API ФГИС АРШИН
    Реализация согласно спецификации внешнего публичного интерфейса v.2.2
    """

    def __init__(self, rate_limiter: DynamicRateLimiter = None):
        self.base_url = Config.API_BASE_URL
        self.session = requests.Session()
        self.rate_limiter = rate_limiter or DynamicRateLimiter()

        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Cache-Control': 'no-cache',
            'Pragma': 'no-cache'
        })

        logger.info(f"APIClient инициализирован: {self.base_url}")

    @staticmethod
    def encode_search_value(value: str) -> str:
        """
        Кодирование поискового значения согласно документации v.2.2:
        - Пробелы заменяются на '?' (любой символ)
        - Кириллица НЕ кодируется
        - '*' остаётся как есть (подстановочный символ)
        """
        # Заменяем пробелы на '?'
        result = value.replace(' ', '?')
        # Кодируем только спецсимволы URL, но не кириллицу и не '*'
        result = quote(result, safe='*?абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ-_.')
        return result

    def _build_vri_params(self, mi_number: str = None, mit_number: str = None,
                          search_term: str = None, year: int = None,
                          start: int = 0, rows: int = Config.DEFAULT_ROWS) -> dict:
        """
        Формирование параметров запроса к /vri согласно документации v.2.2

        Приоритет: атрибутивный поиск > search
        - Если указан mi_number → используем mi_number=...
        - Если указан mit_number → используем mit_number=...
        - Иначе → search=...
        """
        rows = min(rows, Config.MAX_ROWS_PER_REQUEST)
        rows = max(1, rows)
        start = max(0, start)

        params = {
            'start': start,
            'rows': rows,
            # requests сам закодирует пробел как + в URL
            'sort': 'mi_number desc'
        }

        if year:
            params['year'] = year

        # Атрибутивный поиск (приоритет)
        if mi_number:
            params['mi_number'] = mi_number
            logger.debug(f"Атрибутивный поиск: mi_number={mi_number}")
        elif mit_number:
            params['mit_number'] = mit_number
            logger.debug(f"Атрибутивный поиск: mit_number={mit_number}")
        elif search_term:
            # Общий поиск с правильным кодированием
            encoded = self.encode_search_value(search_term)
            params['search'] = encoded
            logger.debug(f"Общий поиск: search={encoded} (исходное: {search_term})")

        return params

    def search_vri(self, mi_number: str = None, mit_number: str = None,
                   search_term: str = None, year: int = None,
                   start: int = 0, rows: int = Config.DEFAULT_ROWS) -> Dict:
        """
        Поиск в реестре поверок (VRI)

        Согласно документации API v.2.2:
        - GET /vri
        - Атрибутивный поиск: mi_number=..., mit_number=...
        - Общий поиск: search=... (пробелы → ?)
        - sort=mi_number+desc по умолчанию
        """
        params = self._build_vri_params(
            mi_number=mi_number, mit_number=mit_number,
            search_term=search_term, year=year, start=start, rows=rows
        )

        query_string = urlencode(params)
        full_url = f"{Config.API_VRI}?{query_string}"
        logger.debug(f"API запрос: GET {full_url}")

        # Rate limiting
        self.rate_limiter.wait()

        for attempt in range(Config.RETRY_ATTEMPTS):
            try:
                start_time = time.time()
                response = self.session.get(
                    Config.API_VRI,
                    params=params,
                    timeout=Config.REQUEST_TIMEOUT
                )
                elapsed_ms = int((time.time() - start_time) * 1000)
                logger.debug(f"Ответ API: статус={response.status_code}, время={elapsed_ms}мс")

                if response.status_code == 200:
                    self.rate_limiter.on_success()
                    data = response.json()
                    items = data.get('result', {}).get('items', [])
                    total_count = data.get('result', {}).get('count', 0)
                    logger.debug(f"Получено {len(items)} записей из {total_count}")
                    return data

                elif response.status_code == 429:
                    self.rate_limiter.on_429()
                    logger.warning(f"API 429: Too Many Requests (попытка {attempt+1}/{Config.RETRY_ATTEMPTS})")
                    # Дополнительная пауза после 429
                    time.sleep(self.rate_limiter.delay * 2)
                    continue

                elif response.status_code == 409:
                    logger.warning(f"API 409: Превышен лимит страниц (start={start})")
                    return {"result": {"items": [], "count": 0}}

                elif response.status_code == 408:
                    logger.error(f"API 408: Request Timeout (попытка {attempt+1})")
                    time.sleep(Config.RETRY_DELAY * (2 ** attempt))
                    continue

                elif response.status_code == 400:
                    # Парсинг тела ошибки
                    error_body = self._parse_error_body(response)
                    logger.error(f"API 400: Bad Request — {error_body}")
                    return {"result": {"items": [], "count": 0, "error": "bad_request", "error_detail": error_body}}

                else:
                    error_body = self._parse_error_body(response)
                    logger.warning(f"API ошибка: статус={response.status_code}, тело={error_body}")
                    if response.status_code >= 500:
                        logger.error(f"5XX ошибка — рекомендуется обратиться в fgis2@rst.gov.ru с requestId={error_body.get('requestId', 'N/A')}")

            except requests.Timeout:
                logger.error(f"Таймаут запроса (попытка {attempt+1}/{Config.RETRY_ATTEMPTS})")
            except requests.ConnectionError as e:
                logger.error(f"Ошибка соединения (попытка {attempt+1}): {e}")
            except Exception as e:
                logger.error(f"Неожиданная ошибка (попытка {attempt+1}): {e}")

            if attempt < Config.RETRY_ATTEMPTS - 1:
                delay = Config.RETRY_DELAY * (2 ** attempt)
                logger.debug(f"Пауза перед повтором: {delay}с")
                time.sleep(delay)

        return {"result": {"items": [], "count": 0, "error": "max_retries"}}

    def _parse_error_body(self, response) -> dict:
        """Парсинг тела ошибки API согласно схеме v.2.2"""
        try:
            data = response.json()
            return {
                'status': data.get('status', ''),
                'message': data.get('message', ''),
                'time': data.get('time', ''),
                'requestId': data.get('requestId', '')
            }
        except Exception:
            return {'raw': response.text[:500]}

    def close(self):
        """Закрытие сессии"""
        self.session.close()
        logger.debug("API сессия закрыта")


# ============== АСИНХРОННЫЙ СБОРЩИК ==============
class AsyncCollector:
    """Асинхронный сборщик данных с динамическим rate limiting"""

    def __init__(self, db: Database = None, rate_limiter: DynamicRateLimiter = None):
        self.db = db or Database()
        self.rate_limiter = rate_limiter or DynamicRateLimiter()
        self.base_url = Config.API_VRI
        self.session = None
        self.is_running = False
        self.stats = {'requests': 0, 'errors': 0, 'found': 0}
        logger.info("AsyncCollector инициализирован")

    async def __aenter__(self):
        self.session = aiohttp.ClientSession(headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json',
            'Accept-Language': 'ru-RU,ru;q=0.9',
        })
        return self

    async def __aexit__(self, *args):
        if self.session:
            await self.session.close()

    async def fetch_page(self, year: int, page: int, mi_number: str = None,
                         search_term: str = None) -> List[Dict]:
        """Асинхронный запрос одной страницы с атрибутивным поиском"""
        params = {
            'year': year,
            'rows': Config.MAX_ROWS_PER_REQUEST,
            'start': page * Config.MAX_ROWS_PER_REQUEST,
            'sort': 'mi_number desc'
        }

        # Атрибутивный поиск
        if mi_number:
            params['mi_number'] = mi_number
        elif search_term:
            # Кодирование согласно v.2.2
            encoded = APIClient.encode_search_value(search_term)
            params['search'] = encoded

        self.stats['requests'] += 1

        # Rate limiting
        self.rate_limiter.wait()

        try:
            async with self.session.get(self.base_url, params=params, timeout=15) as resp:
                if resp.status == 200:
                    self.rate_limiter.on_success()
                    data = await resp.json()
                    items = data.get('result', {}).get('items', [])
                    self.stats['found'] += len(items)
                    logger.debug(f"Год {year}, стр.{page}: получено {len(items)} записей")
                    return items
                elif resp.status == 429:
                    self.rate_limiter.on_429()
                    logger.warning(f"Год {year}, стр.{page}: 429 Too Many Requests")
                    await asyncio.sleep(self.rate_limiter.delay * 2)
                    return []
                elif resp.status == 409:
                    logger.debug(f"Год {year}: превышен лимит страниц")
                    return []
                else:
                    self.stats['errors'] += 1
                    logger.debug(f"Год {year}, стр.{page}: статус {resp.status}")
                    return []
        except asyncio.TimeoutError:
            self.stats['errors'] += 1
            logger.debug(f"Год {year}, стр.{page}: таймаут")
            return []
        except Exception as e:
            self.stats['errors'] += 1
            logger.debug(f"Год {year}, стр.{page}: ошибка {e}")
            return []

    async def collect_year(self, year: int, progress_callback=None,
                           mi_number: str = None, search_term: str = None) -> int:
        """Сбор данных за указанный год"""
        logger.info(f"Начало сбора за {year} год (mi_number={mi_number}, search={search_term})")

        existing_ids = self.db.get_existing_ids()
        new_count = 0
        page = 0
        max_pages = 50

        pbar = None
        if TQDM_AVAILABLE and progress_callback is None:
            pbar = tqdm(total=max_pages, desc=f"  {year}", unit="стр", leave=False)

        while page < max_pages and self.is_running:
            items = await self.fetch_page(year, page, mi_number=mi_number, search_term=search_term)

            if not items:
                logger.debug(f"Год {year}: данных больше нет на странице {page}")
                break

            records = []
            for item in items:
                vri_id = item.get('vri_id')
                if vri_id and vri_id not in existing_ids and VerificationRecord.is_electric_meter(item.get('mit_title', '')):
                    record = VerificationRecord.from_api_response(item)
                    records.append(record)
                    existing_ids.add(vri_id)

            if records:
                saved = self.db.save_records(records)
                new_count += saved
                logger.info(f"Год {year}, стр.{page}: сохранено {saved} новых записей")

                if progress_callback:
                    progress_callback(year, page, saved)

            if pbar:
                pbar.update(1)
                pbar.set_postfix({"найдено": new_count})

            await asyncio.sleep(0.1)

            if len(items) < Config.MAX_ROWS_PER_REQUEST:
                break

            page += 1

        if pbar:
            pbar.close()

        logger.info(f"Год {year} завершён: собрано {new_count} новых записей")
        return new_count

    def stop(self):
        self.is_running = False
        logger.info("Получена команда остановки сбора")


# ============== ЭКСПОРТЕР ==============
class Exporter:
    """Экспорт данных с полной служебной информацией"""

    @staticmethod
    def ensure_export_dir():
        os.makedirs(Config.EXPORT_DIR, exist_ok=True)

    @staticmethod
    def to_csv(data: List[Dict], filename: str = None,
               include_service_fields: bool = True,
               extra_columns: List[str] = None) -> str:
        """
        Экспорт в CSV с русскими заголовками

        Args:
            data: Список записей
            filename: Имя файла
            include_service_fields: Включать служебные поля
            extra_columns: Дополнительные колонки из Excel
        """
        Exporter.ensure_export_dir()

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{Config.EXPORT_DIR}/export_{timestamp}.csv"

        if not data:
            logger.warning("Нет данных для экспорта")
            with open(filename, 'w', encoding='utf-8-sig') as f:
                f.write("Нет данных")
            return filename

        field_mapping = {
            'vri_id': 'ID',
            'mit_number': '№ реестра',
            'mit_title': 'Наименование',
            'mit_notation': 'Обозначение типа',
            'mi_modification': 'Модификация',
            'mi_number': 'Зав. номер',
            'verification_date': 'Дата поверки',
            'valid_date': 'Действ. до',
            'applicability': 'Пригоден',
            'org_title': 'Поверитель',
            'result_docnum': '№ документа',
            'sticker_num': '№ наклейки',
            'manufacturer': 'Производитель',
            'search_query': 'Поисковый запрос',
            'row_index': '№ строки',
            'id_pu': 'Id_ПУ',
            'collected_at': 'Время сохранения',
        }

        base_fields = [
            'vri_id', 'mit_number', 'mit_title', 'mit_notation',
            'mi_modification', 'mi_number', 'verification_date',
            'valid_date', 'applicability', 'org_title', 'result_docnum',
            'sticker_num', 'manufacturer'
        ]

        if include_service_fields:
            base_fields.extend([
                'id_pu', 'search_query', 'row_index'
            ])

        base_fields.append('collected_at')

        # Добавляем дополнительные колонки из Excel
        if extra_columns:
            base_fields.extend(extra_columns)

        export_fields = [f for f in base_fields if f in data[0]]
        russian_headers = [field_mapping.get(f, f) for f in export_fields]

        with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=export_fields, delimiter=';', extrasaction='ignore')
            writer.writerow(dict(zip(export_fields, russian_headers)))
            for row in data:
                writer.writerow(row)

        logger.info(f"Экспортировано {len(data)} записей в {filename}")
        return filename

    @staticmethod
    def to_excel(data: List[Dict], filename: str = None,
                 include_service_fields: bool = True,
                 extra_columns: List[str] = None) -> str:
        """Экспорт в Excel с русскими заголовками и форматированием"""
        if not PANDAS_AVAILABLE:
            logger.warning("pandas не установлен, используется CSV")
            return Exporter.to_csv(data, filename, include_service_fields, extra_columns)

        Exporter.ensure_export_dir()

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{Config.EXPORT_DIR}/export_{timestamp}.xlsx"

        field_mapping = {
            'vri_id': 'ID',
            'mit_number': '№ реестра',
            'mit_title': 'Наименование',
            'mit_notation': 'Обозначение типа',
            'mi_modification': 'Модификация',
            'mi_number': 'Зав. номер',
            'verification_date': 'Дата поверки',
            'valid_date': 'Действ. до',
            'applicability': 'Пригоден',
            'org_title': 'Поверитель',
            'result_docnum': '№ документа',
            'sticker_num': '№ наклейки',
            'manufacturer': 'Производитель',
            'search_query': 'Поисковый запрос',
            'row_index': '№ строки',
            'id_pu': 'Id_ПУ',
            'collected_at': 'Время сохранения',
        }

        base_fields = [
            'vri_id', 'mit_number', 'mit_title', 'mit_notation',
            'mi_modification', 'mi_number', 'verification_date',
            'valid_date', 'applicability', 'org_title', 'result_docnum',
            'sticker_num', 'manufacturer'
        ]

        if include_service_fields:
            base_fields.extend(['id_pu', 'search_query', 'row_index'])

        base_fields.append('collected_at')

        if extra_columns:
            base_fields.extend(extra_columns)

        if data:
            export_fields = [f for f in base_fields if f in data[0]]
        else:
            export_fields = base_fields

        filtered_data = []
        for row in data:
            filtered_row = {field_mapping.get(k, k): v for k, v in row.items() if k in export_fields}
            filtered_data.append(filtered_row)

        df = pd.DataFrame(filtered_data)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Результаты')

            workbook = writer.book
            worksheet = writer.sheets['Результаты']

            header_font = ExcelFont(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')

            for col_num, value in enumerate(df.columns.values):
                cell = worksheet.cell(row=1, column=col_num+1)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Экспортировано {len(data)} записей в {filename}")
        return filename

    @staticmethod
    def export_stats(stats: Dict) -> str:
        """Экспорт статистики в JSON"""
        Exporter.ensure_export_dir()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{Config.EXPORT_DIR}/stats_{timestamp}.json"

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(stats, f, ensure_ascii=False, indent=2)

        logger.info(f"Статистика экспортирована в {filename}")
        return filename


# ============== ОБРАБОТЧИК EXCEL ==============
class ExcelHandler:
    """Работа с Excel файлами для пакетного поиска с сохранением ВСЕХ связей"""

    @staticmethod
    def detect_columns(df: pd.DataFrame) -> Dict[str, Tuple[str, int]]:
        """
        Автоматическое определение колонок по заголовкам
        Структура таблицы определяется автоматически
        """
        column_mapping = {}

        possible_columns = {
            'mi_number': [
                'серийный номер', 'заводской номер', 'номер пу', 'серийный номер пу',
                'зав. номер', 'зав №', 'serial number', 'device_number',
                'серийный', 'заводской', 'serial', 'pu_number', 'id_пу',
                'номер пу', 'прибор', 'счетчик', 'измеритель', 'id'
            ],
            'mit_number': [
                'номер в реестре', 'реестровый номер', 'рег. номер', 'регистрационный',
                'тип си номер', 'номер типа', 'mit_number', 'vri_id',
                'реестр', 'тип си', 'тип'
            ],
            'mit_title': [
                'модель', 'model', 'наименование типа', 'тип прибора', 'тип пу',
                'наименование', 'название', 'name', 'тип си', 'устройство',
                'описание', 'прибор'
            ],
            'org_title': [
                'организация поверитель', 'поверитель', 'организация', 'org_title',
                'org', 'кто поверял', 'поверка', 'company', 'предприятие',
                'эксплуатационная ответственность'
            ],
            'vri_id': [
                'vri_id', 'vri', 'идентификатор', 'уникальный', 'uuid',
                'ключ', 'key', 'id записи'
            ],
            'search_term': [
                'поисковый запрос', 'ключевое слово', 'поиск', 'query', 'запрос',
                'search', 'термин', 'ключ'
            ],
            'year': [
                'год поверки', 'поверка год', 'год поверки си', 'verification year',
                'год', 'year', 'дата поверки', 'mpi', 'мпИ'
            ],
            'manufacture_year': [
                'год выпуска', 'год производства', 'год изготовления', 'дата выпуска',
                'выпуск', 'manufacture', 'production year'
            ]
        }

        df_lower = [col.lower().strip() for col in df.columns]

        for field, keywords in possible_columns.items():
            best_match = None
            best_score = 0

            for idx, col_lower in enumerate(df_lower):
                score = 0
                original_col = df.columns[idx]

                if col_lower == field:
                    score = 100
                else:
                    for priority, keyword in enumerate(keywords):
                        if keyword == col_lower:
                            score = max(score, 95 - priority)
                        elif col_lower.startswith(keyword):
                            score = max(score, 80 - priority)
                        elif keyword in col_lower:
                            score = max(score, 60 - priority)

                if score > best_score:
                    best_score = score
                    best_match = (original_col, idx)

            if best_match and best_score > 50:
                col_name, col_index = best_match
                used_cols = [v[0] for v in column_mapping.values()]
                if col_name not in used_cols:
                    column_mapping[field] = (col_name, col_index)

        if 'search_term' not in column_mapping and 'mi_number' in column_mapping:
            logger.debug("search_term не найден, будет использоваться mi_number для поиска")

        logger.info(f"Автоматически определено колонок: {len(column_mapping)}")
        return column_mapping

    @staticmethod
    def read_queries(filename: str) -> Tuple[List[Dict], List[str]]:
        """
        Чтение запросов из Excel с сохранением ВСЕХ колонок

        Returns:
            (queries, all_extra_columns) — список запросов и список всех дополнительных колонок
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas не установлен")

        logger.info(f"Чтение Excel файла: {filename}")
        df = pd.read_excel(filename)
        logger.debug(f"Прочитано {len(df)} строк, колонки: {list(df.columns)}")

        column_mapping = ExcelHandler.detect_columns(df)

        if not column_mapping:
            logger.warning("Колонки не определены, используется первая колонка")
            column_mapping = {'mi_number': (df.columns[0], 0)}

        # Определяем все дополнительные колонки (кроме определённых как поисковые)
        mapped_cols = {col_name for col_name, _ in column_mapping.values()}
        extra_columns = [col for col in df.columns if col not in mapped_cols]

        queries = []
        for idx, row in df.iterrows():
            query = {
                'row_index': idx + 1,
                'original_data': {},
                'extra_fields': {}
            }

            # Сохранение ВСЕХ исходных данных
            for col in df.columns:
                value = row[col]
                if pd.notna(value):
                    query['original_data'][col] = str(value)

            # Маппинг полей
            for field_name, (col_name, _) in column_mapping.items():
                value = row[col_name]
                if pd.notna(value):
                    query[field_name] = str(value)

            # Дополнительные поля из Excel
            for col in extra_columns:
                value = row[col]
                if pd.notna(value):
                    query['extra_fields'][col] = str(value)

            # Для поиска используем mi_number (атрибутивный поиск)
            if 'mi_number' in query or 'search_term' in query:
                queries.append(query)

        logger.info(f"Загружено {len(queries)} запросов, дополнительных колонок: {len(extra_columns)}")
        return queries, extra_columns

    @staticmethod
    def create_template(filename: str = None) -> str:
        """Создание шаблона Excel файла"""
        if not PANDAS_AVAILABLE:
            return ""

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{Config.EXPORT_DIR}/batch_search_template_{timestamp}.xlsx"

        Exporter.ensure_export_dir()

        data = {
            'Заводской номер': ['12345678', '87654321', ''],
            'Наименование': ['Счетчик электрической энергии', '', ''],
            'Номер в реестре': ['77310-20', '', ''],
            'Год поверки': [2023, 2022, ''],
            'Примечание': ['Первый запрос', 'Второй запрос', '']
        }

        df = pd.DataFrame(data)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Шаблон')

            worksheet = writer.sheets['Шаблон']
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 12
            worksheet.column_dimensions['E'].width = 25

        logger.info(f"Шаблон создан: {filename}")
        return filename


# ============== ГРАФИЧЕСКИЙ ИНТЕРФЕЙС ==============
class ArshinApp:
    """Главное окно приложения v1.2"""

    def __init__(self, root):
        self.root = root
        self.root.title("ФГИС Аршин - Поиск счётчиков электроэнергии v1.2")
        self.root.geometry(f"{Config.WINDOW_WIDTH}x{Config.WINDOW_HEIGHT}")
        self.root.minsize(1024, 600)

        # Компоненты
        self.rate_limiter = DynamicRateLimiter()
        self.db = Database()
        self.api_client = APIClient(rate_limiter=self.rate_limiter)
        self.collector = AsyncCollector(self.db, rate_limiter=self.rate_limiter)

        # Пакетный поиск
        self.batch_queries = []
        self.batch_extra_columns = []  # Дополнительные колонки из Excel
        self.batch_results = []
        self.batch_all_columns = []
        self.current_batch_index = 0
        self.batch_running = False
        self.batch_stop_requested = False

        # Одиночный поиск
        self.single_search_results = []
        self.single_search_results_all = []

        # Асинхронный сбор
        self.async_running = False
        self.async_thread = None

        logger.info(f"Приложение запущено v1.2, окно: {Config.WINDOW_WIDTH}x{Config.WINDOW_HEIGHT}")
        logger.info(f"Загружено {len(EXACT_QUERIES)} точных запросов, {len(MANUFACTURERS_RULES)} производителей")
        logger.info(f"Rate limiter: start={Config.RATE_LIMIT_START}, min={Config.RATE_LIMIT_MIN}, max={Config.RATE_LIMIT_MAX}")

        self.setup_ui()
        self.update_stats()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_ui(self):
        """Создание интерфейса"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Экспорт в CSV", command=self.export_csv, accelerator="Ctrl+S")
        file_menu.add_command(label="Экспорт в Excel", command=self.export_excel)
        file_menu.add_command(label="Экспорт статистики", command=self.export_stats_json)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.on_close, accelerator="Ctrl+Q")

        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Инструменты", menu=tools_menu)
        tools_menu.add_command(label="Обновить статистику", command=self.update_stats)
        tools_menu.add_command(label="Очистить дубликаты", command=self.clean_duplicates)
        tools_menu.add_command(label="История поиска", command=self.show_search_history)
        tools_menu.add_separator()
        tools_menu.add_command(label="Создать шаблон Excel", command=lambda: ExcelHandler.create_template())

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Помощь", menu=help_menu)
        help_menu.add_command(label="О программе", command=self.show_about)
        help_menu.add_command(label="Лог файл", command=self.show_log_file)

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)

        self.setup_single_search_tab()
        self.setup_batch_operations_tab()
        self.setup_results_tab()
        self.setup_filters_tab()
        self.setup_stats_tab()

        self.root.bind('<Control-s>', lambda e: self.export_csv())
        self.root.bind('<Control-q>', lambda e: self.on_close())

    def setup_single_search_tab(self):
        """Вкладка одиночного поиска"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔍 Поиск")

        params_frame = ttk.LabelFrame(tab, text="Параметры поиска (атрибутивный поиск по mi_number)", padding=3)
        params_frame.pack(fill=tk.X, padx=5, pady=2)

        search_frame = ttk.Frame(params_frame)
        search_frame.pack(fill=tk.X, pady=1)

        ttk.Label(search_frame, text="Зав. номер (mi_number):", width=22).pack(side=tk.LEFT)
        self.search_entry = ttk.Entry(search_frame, width=30)
        self.search_entry.pack(side=tk.LEFT, padx=3)
        self.search_entry.bind('<Button-3>', lambda e: None)

        self.search_start_btn = ttk.Button(search_frame, text="▶ Найти", command=self.single_search, width=12)
        self.search_start_btn.pack(side=tk.LEFT, padx=3)

        self.search_stop_btn = ttk.Button(search_frame, text="⏹ Стоп", command=self.stop_single_search, width=10, state=tk.DISABLED)
        self.search_stop_btn.pack(side=tk.LEFT, padx=3)

        clear_btn = ttk.Button(search_frame, text="✖ Очистить", command=self.clear_single_search, width=12)
        clear_btn.pack(side=tk.LEFT, padx=3)

        info_frame = ttk.Frame(params_frame)
        info_frame.pack(fill=tk.X, pady=1)

        self.search_info_var = tk.StringVar(value="Готов к поиску (атрибутивный: mi_number=...)")
        ttk.Label(info_frame, textvariable=self.search_info_var, foreground='blue').pack(side=tk.LEFT, padx=5)

        # Rate limiter status
        self.rate_limit_var = tk.StringVar(value=f"Rate delay: {self.rate_limiter.delay:.2f}с")
        ttk.Label(info_frame, textvariable=self.rate_limit_var, foreground='gray').pack(side=tk.RIGHT, padx=5)

        years_label_frame = ttk.Frame(params_frame)
        years_label_frame.pack(fill=tk.X, pady=1)
        ttk.Label(years_label_frame, text="Годы:").pack(side=tk.LEFT, padx=3)

        years_frame = ttk.Frame(params_frame)
        years_frame.pack(fill=tk.X, pady=1)

        self.search_type = tk.StringVar(value="mi_number")
        self.search_running = False
        self.search_stop_requested = False

        self.search_year_vars = {}
        for year in range(2010, 2027):
            var = tk.BooleanVar(value=True)
            self.search_year_vars[year] = var
            ttk.Checkbutton(years_frame, text=str(year), variable=var, width=5).pack(side=tk.LEFT, padx=0)

        # Результаты
        results_frame = ttk.LabelFrame(tab, text="Результаты поиска", padding=0)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        tree_container = ttk.Frame(results_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)

        columns = (
            'vri_id', 'mit_number', 'mit_title', 'mit_notation', 'mi_modification',
            'mi_number', 'verification_date', 'valid_date', 'org_title', 'result_docnum', 'sticker_num', 'applicability'
        )
        self.single_results_tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=12, selectmode='extended')

        self.single_results_tree.heading('vri_id', text='ID записи')
        self.single_results_tree.heading('mit_number', text='№ в реестре')
        self.single_results_tree.heading('mit_title', text='Наименование СИ')
        self.single_results_tree.heading('mit_notation', text='Обозначение типа')
        self.single_results_tree.heading('mi_modification', text='Модификация')
        self.single_results_tree.heading('mi_number', text='Зав. номер')
        self.single_results_tree.heading('verification_date', text='Дата поверки')
        self.single_results_tree.heading('valid_date', text='Действ. до')
        self.single_results_tree.heading('org_title', text='Поверитель')
        self.single_results_tree.heading('result_docnum', text='№ документа')
        self.single_results_tree.heading('sticker_num', text='№ наклейки')
        self.single_results_tree.heading('applicability', text='Пригоден')

        self.single_results_tree.column('vri_id', width=90)
        self.single_results_tree.column('mit_number', width=80)
        self.single_results_tree.column('mit_title', width=250)
        self.single_results_tree.column('mit_notation', width=100)
        self.single_results_tree.column('mi_modification', width=150)
        self.single_results_tree.column('mi_number', width=100)
        self.single_results_tree.column('verification_date', width=90)
        self.single_results_tree.column('valid_date', width=90)
        self.single_results_tree.column('org_title', width=180)
        self.single_results_tree.column('result_docnum', width=130)
        self.single_results_tree.column('sticker_num', width=80)
        self.single_results_tree.column('applicability', width=70)

        v_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.single_results_tree.yview)
        h_scroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.single_results_tree.xview)
        self.single_results_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.single_results_tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        self.setup_tree_context_menu(self.single_results_tree)

        self.single_status_var = tk.StringVar(value="Готов")
        status_bar = ttk.Label(tab, textvariable=self.single_status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, padx=3, pady=1)

    def setup_batch_operations_tab(self):
        """Вкладка пакетных операций"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📦 Пакетная")

        file_frame = ttk.Frame(tab)
        file_frame.pack(fill=tk.X, padx=5, pady=2)

        ttk.Label(file_frame, text="Excel:").pack(side=tk.LEFT)
        self.batch_file_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.batch_file_var, width=45).pack(side=tk.LEFT, padx=3, fill=tk.X, expand=True)
        ttk.Button(file_frame, text="📂", command=self.select_batch_file, width=3).pack(side=tk.LEFT, padx=2)
        ttk.Button(file_frame, text="📄", command=lambda: ExcelHandler.create_template(), width=3).pack(side=tk.LEFT, padx=2)

        params_frame = ttk.LabelFrame(tab, text="Параметры", padding=2)
        params_frame.pack(fill=tk.X, padx=5, pady=2)

        top_row = ttk.Frame(params_frame)
        top_row.pack(fill=tk.X, pady=1)

        years_group = ttk.LabelFrame(top_row, text="Годы", padding=2)
        years_group.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.year_vars = {}

        years_row1 = ttk.Frame(years_group)
        years_row1.pack(fill=tk.X, pady=0)
        for year in range(2010, 2017):
            var = tk.BooleanVar(value=True)
            self.year_vars[year] = var
            ttk.Checkbutton(years_row1, text=str(year), variable=var, width=5).pack(side=tk.LEFT, padx=0)

        years_row2 = ttk.Frame(years_group)
        years_row2.pack(fill=tk.X, pady=0)
        for year in range(2020, 2027):
            var = tk.BooleanVar(value=True)
            self.year_vars[year] = var
            ttk.Checkbutton(years_row2, text=str(year), variable=var, width=5).pack(side=tk.LEFT, padx=0)

        ttk.Button(years_row2, text="Все...", command=self.select_all_years_dialog, width=6).pack(side=tk.LEFT, padx=2)

        buttons_frame = ttk.Frame(top_row)
        buttons_frame.pack(side=tk.RIGHT, padx=5)

        self.batch_start_btn = ttk.Button(buttons_frame, text="▶ Запуск", command=self.start_batch_search, width=12)
        self.batch_start_btn.pack(side=tk.LEFT, padx=2)

        self.batch_stop_btn = ttk.Button(buttons_frame, text="⏹ Стоп", command=self.stop_batch_search, width=10, state=tk.DISABLED)
        self.batch_stop_btn.pack(side=tk.LEFT, padx=2)

        progress_frame = ttk.Frame(params_frame)
        progress_frame.pack(fill=tk.X, pady=1)

        self.batch_progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.batch_progress.pack(fill=tk.X, pady=1)

        stats_row = ttk.Frame(progress_frame)
        stats_row.pack(fill=tk.X, pady=0)

        self.batch_processed_var = tk.StringVar(value="Обработано: 0 / 0")
        ttk.Label(stats_row, textvariable=self.batch_processed_var, width=25).pack(side=tk.LEFT, padx=5)

        self.batch_found_var = tk.StringVar(value="Найдено: 0")
        ttk.Label(stats_row, textvariable=self.batch_found_var, width=15).pack(side=tk.LEFT, padx=5)

        self.batch_filtered_var = tk.StringVar(value="Отфильтровано: 0")
        ttk.Label(stats_row, textvariable=self.batch_filtered_var, width=20).pack(side=tk.LEFT, padx=5)

        # Rate limiter display
        self.batch_rate_var = tk.StringVar(value=f"Rate: {self.rate_limiter.delay:.2f}с")
        ttk.Label(stats_row, textvariable=self.batch_rate_var, foreground='gray').pack(side=tk.RIGHT, padx=5)

        preview_frame = ttk.LabelFrame(tab, text="Запросы из файла (предпросмотр)", padding=0)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

        tree_container = ttk.Frame(preview_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)

        columns = ('row_num', 'id_pu', 'mi_number', 'mit_title', 'mit_number', 'year',
                   'contract_number', 'edo_code', 'balance_owner', 'operation_responsibility', 'mpi')
        self.batch_preview_tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=14)

        self.batch_preview_tree.heading('row_num', text='№')
        self.batch_preview_tree.heading('id_pu', text='Id_ПУ')
        self.batch_preview_tree.heading('mi_number', text='Серийный номер')
        self.batch_preview_tree.heading('mit_title', text='Модель')
        self.batch_preview_tree.heading('mit_number', text='Тип ПУ')
        self.batch_preview_tree.heading('year', text='Год')
        self.batch_preview_tree.heading('contract_number', text='Номер договора')
        self.batch_preview_tree.heading('edo_code', text='Код ЭДО')
        self.batch_preview_tree.heading('balance_owner', text='Балансовая принадлежность')
        self.batch_preview_tree.heading('operation_responsibility', text='Эксплуатационная ответственность')
        self.batch_preview_tree.heading('mpi', text='МПИ')

        self.batch_preview_tree.column('row_num', width=35)
        self.batch_preview_tree.column('id_pu', width=90)
        self.batch_preview_tree.column('mi_number', width=110)
        self.batch_preview_tree.column('mit_title', width=180)
        self.batch_preview_tree.column('mit_number', width=90)
        self.batch_preview_tree.column('year', width=50)
        self.batch_preview_tree.column('contract_number', width=100)
        self.batch_preview_tree.column('edo_code', width=75)
        self.batch_preview_tree.column('balance_owner', width=150)
        self.batch_preview_tree.column('operation_responsibility', width=180)
        self.batch_preview_tree.column('mpi', width=50)

        v_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.batch_preview_tree.yview)
        h_scroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.batch_preview_tree.xview)
        self.batch_preview_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.batch_preview_tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

    def setup_filters_tab(self):
        """Вкладка управления фильтрами"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔧 Фильтры")

        exact_frame = ttk.LabelFrame(tab, text="Точные запросы", padding=10)
        exact_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.exact_queries_list = scrolledtext.ScrolledText(exact_frame, height=10, font=('Consolas', 9))
        self.exact_queries_list.pack(fill=tk.BOTH, expand=True, pady=5)

        exact_btn_frame = ttk.Frame(exact_frame)
        exact_btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(exact_btn_frame, text="📂 Загрузить", command=self.load_exact_queries_file, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(exact_btn_frame, text="💾 Сохранить", command=self.save_exact_queries_file, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(exact_btn_frame, text="🗑 Очистить", command=self.clear_exact_queries, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(exact_btn_frame, text="➕ Добавить", command=self.add_exact_query, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(exact_btn_frame, text="🔄 Из файла", command=self.load_exact_queries_display, width=12).pack(side=tk.LEFT, padx=2)

        mfg_frame = ttk.LabelFrame(tab, text="Производители", padding=10)
        mfg_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.manufacturers_list = scrolledtext.ScrolledText(mfg_frame, height=10, font=('Consolas', 9))
        self.manufacturers_list.pack(fill=tk.BOTH, expand=True, pady=5)

        mfg_btn_frame = ttk.Frame(mfg_frame)
        mfg_btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(mfg_btn_frame, text="📂 Загрузить", command=self.load_manufacturers_file, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(mfg_btn_frame, text="💾 Сохранить", command=self.save_manufacturers_file, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(mfg_btn_frame, text="🗑 Очистить", command=self.clear_manufacturers, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(mfg_btn_frame, text="➕ Добавить", command=self.add_manufacturer, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(mfg_btn_frame, text="🔄 Из файла", command=self.load_manufacturers_display, width=12).pack(side=tk.LEFT, padx=2)

        self.filters_status_var = tk.StringVar(value="Готов")
        status_bar = ttk.Label(tab, textvariable=self.filters_status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, padx=3, pady=1)

        self.load_exact_queries_display()
        self.load_manufacturers_display()

    def setup_stats_tab(self):
        """Вкладка статистики"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📊 Статистика")

        stats_frame = ttk.LabelFrame(tab, text="База данных", padding=5)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        self.stats_text = scrolledtext.ScrolledText(stats_frame, height=20, font=('Consolas', 10))
        self.stats_text.pack(fill=tk.BOTH, expand=True)

        btn_frame = ttk.Frame(tab)
        btn_frame.pack(pady=3)

        ttk.Button(btn_frame, text="Обновить", command=self.update_stats, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="Экспорт JSON", command=self.export_stats_json, width=12).pack(side=tk.LEFT, padx=3)

    def setup_results_tab(self):
        """Вкладка просмотра результатов"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📋 Результаты")

        filter_frame = ttk.Frame(tab)
        filter_frame.pack(fill=tk.X, padx=5, pady=2)

        row0 = ttk.Frame(filter_frame)
        row0.pack(fill=tk.X, pady=1)

        ttk.Label(row0, text="Поиск:").pack(side=tk.LEFT)
        self.db_search_var = tk.StringVar()
        ttk.Entry(row0, textvariable=self.db_search_var, width=25).pack(side=tk.LEFT, padx=2)
        ttk.Button(row0, text="🔍", command=self.load_db_data, width=3).pack(side=tk.LEFT)
        ttk.Button(row0, text="✖", command=self.reset_db_filters, width=3).pack(side=tk.LEFT, padx=5)

        btn_row = ttk.Frame(row0)
        btn_row.pack(side=tk.RIGHT)
        ttk.Button(btn_row, text="📄 CSV", command=self.export_results_csv, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="📊 Excel", command=self.export_results_excel, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="🗑 БД", command=self.clear_database, width=7).pack(side=tk.LEFT, padx=2)

        row1 = ttk.Frame(filter_frame)
        row1.pack(fill=tk.X, pady=1)

        ttk.Label(row1, text="Год:").pack(side=tk.LEFT)
        self.db_year_var = tk.StringVar(value="")
        year_combo = ttk.Combobox(row1, textvariable=self.db_year_var, width=5)
        year_combo['values'] = [''] + Config.YEARS
        year_combo.pack(side=tk.LEFT, padx=2)

        ttk.Label(row1, text="| Производитель:").pack(side=tk.LEFT, padx=(10, 0))
        self.db_manufacturer_var = tk.StringVar(value="")
        self.db_manufacturer_combo = ttk.Combobox(row1, textvariable=self.db_manufacturer_var, width=12)
        self.db_manufacturer_combo.pack(side=tk.LEFT, padx=2)

        table_frame = ttk.LabelFrame(tab, text="Результаты", padding=0)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

        tree_container = ttk.Frame(table_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)

        columns = (
            'vri_id', 'id_pu', 'mit_number', 'mit_title', 'mit_notation', 'mi_modification',
            'mi_number', 'verification_date', 'valid_date', 'applicability',
            'org_title', 'result_docnum', 'sticker_num', 'manufacturer',
            'search_query', 'row_index'
        )
        self.db_tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=12)

        self.db_tree.heading('vri_id', text='ID')
        self.db_tree.heading('id_pu', text='Id_ПУ')
        self.db_tree.heading('mit_number', text='№ реестра')
        self.db_tree.heading('mit_title', text='Наименование')
        self.db_tree.heading('mit_notation', text='Обозначение')
        self.db_tree.heading('mi_modification', text='Модиф.')
        self.db_tree.heading('mi_number', text='Зав. №')
        self.db_tree.heading('verification_date', text='Дата поверки')
        self.db_tree.heading('valid_date', text='Действ. до')
        self.db_tree.heading('applicability', text='Пригоден')
        self.db_tree.heading('org_title', text='Поверитель')
        self.db_tree.heading('result_docnum', text='№ док.')
        self.db_tree.heading('sticker_num', text='№ наклейки')
        self.db_tree.heading('manufacturer', text='Производитель')
        self.db_tree.heading('search_query', text='Запрос')
        self.db_tree.heading('row_index', text='№ строки')

        self.db_tree.column('vri_id', width=75)
        self.db_tree.column('id_pu', width=85)
        self.db_tree.column('mit_number', width=70)
        self.db_tree.column('mit_title', width=180)
        self.db_tree.column('mit_notation', width=85)
        self.db_tree.column('mi_modification', width=85)
        self.db_tree.column('mi_number', width=85)
        self.db_tree.column('verification_date', width=85)
        self.db_tree.column('valid_date', width=85)
        self.db_tree.column('applicability', width=60)
        self.db_tree.column('org_title', width=120)
        self.db_tree.column('result_docnum', width=95)
        self.db_tree.column('sticker_num', width=75)
        self.db_tree.column('manufacturer', width=95)
        self.db_tree.column('search_query', width=110)
        self.db_tree.column('row_index', width=50)

        v_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.db_tree.yview)
        h_scroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.db_tree.xview)
        self.db_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.db_tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        self.setup_tree_context_menu(self.db_tree)

    def setup_tree_context_menu(self, tree):
        """Контекстное меню"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Копировать строку", command=lambda: self.copy_selected(tree))
        menu.add_command(label="Копировать все", command=lambda: self.copy_all(tree))
        menu.add_separator()
        menu.add_command(label="Открыть в браузере", command=lambda: self.open_in_browser(tree))
        menu.add_separator()
        menu.add_command(label="Экспорт выбранного", command=lambda: self.export_selected(tree))

        def show_menu(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                menu.post(event.x_root, event.y_root)

        tree.bind("<Button-3>", show_menu)
        tree.bind("<Double-1>", lambda e: self.open_in_browser(tree))

    def copy_selected(self, tree):
        item = tree.selection()[0]
        values = tree.item(item, 'values')
        self.root.clipboard_clear()
        self.root.clipboard_append('\t'.join(map(str, values)))

    def copy_all(self, tree):
        lines = []
        for item in tree.get_children():
            lines.append('\t'.join(map(str, tree.item(item, 'values'))))
        self.root.clipboard_clear()
        self.root.clipboard_append('\n'.join(lines))

    def open_in_browser(self, tree):
        item = tree.selection()[0]
        values = tree.item(item, 'values')
        if values:
            vri_id = values[0]
            url = f"https://fgis.gost.ru/fundmetrology/cm/erts/?id={vri_id}"
            logger.info(f"Открытие в браузере: {url}")
            webbrowser.open(url)

    def export_selected(self, tree):
        items = tree.selection()
        if not items:
            messagebox.showwarning("Предупреждение", "Выберите записи для экспорта")
            return

        data = []
        for item in items:
            values = tree.item(item, 'values')
            vri_id = values[0]
            results = self.db.search(query=vri_id, query_type='exact', limit=1)
            if results:
                data.extend(results)

        if data:
            filename = filedialog.asksaveasfilename(
                title="Сохранить выбранные записи",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx")]
            )
            if filename:
                if filename.endswith('.xlsx') and PANDAS_AVAILABLE:
                    Exporter.to_excel(data, filename)
                else:
                    Exporter.to_csv(data, filename)
                messagebox.showinfo("Экспорт", f"Экспортировано {len(data)} записей")

    def export_results_csv(self):
        data = self.get_db_data()
        if not data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return

        filename = filedialog.asksaveasfilename(title="Экспорт в CSV", defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if filename:
            Exporter.to_csv(data, filename, extra_columns=self.batch_extra_columns)
            messagebox.showinfo("Экспорт", f"Экспортировано {len(data)} записей в {filename}")

    def export_results_excel(self):
        if not PANDAS_AVAILABLE:
            messagebox.showwarning("Предупреждение", "pandas не установлен")
            return

        data = self.get_db_data()
        if not data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return

        filename = filedialog.asksaveasfilename(title="Экспорт в Excel", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            Exporter.to_excel(data, filename, extra_columns=self.batch_extra_columns)
            messagebox.showinfo("Экспорт", f"Экспортировано {len(data)} записей в {filename}")

    def clear_database(self):
        if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите очистить всю базу данных?"):
            return

        try:
            with sqlite3.connect(self.db.db_path) as conn:
                conn.execute('DELETE FROM verification_records')
                conn.commit()
            messagebox.showinfo("Очистка", "База данных очищена")
            self.load_db_data()
            self.update_stats()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось очистить БД:\n{e}")

    def get_db_data(self) -> List[Dict]:
        """Получение данных из таблицы Результаты"""
        data = []
        for item in self.db_tree.get_children():
            values = self.db_tree.item(item, 'values')
            if values and len(values) >= 16:
                row = {
                    'vri_id': values[0],
                    'id_pu': values[1],
                    'mit_number': values[2],
                    'mit_title': values[3],
                    'mit_notation': values[4],
                    'mi_modification': values[5],
                    'mi_number': values[6],
                    'verification_date': values[7],
                    'valid_date': values[8],
                    'applicability': values[9],
                    'org_title': values[10],
                    'result_docnum': values[11],
                    'sticker_num': values[12],
                    'manufacturer': values[13],
                    'search_query': values[14],
                    'row_index': values[15],
                }
                data.append(row)
        return data

    # ========== Пакетный поиск ==========
    def select_batch_file(self):
        """Выбор Excel файла"""
        filename = filedialog.askopenfilename(
            title="Выберите файл Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.batch_file_var.set(filename)
            try:
                if PANDAS_AVAILABLE:
                    df = pd.read_excel(filename)
                    logger.info(f"Прочитано {len(df)} строк, колонки: {list(df.columns)}")

                    self.batch_all_columns = list(df.columns)

                    # Чтение с сохранением ВСЕХ колонок
                    queries, extra_columns = ExcelHandler.read_queries(filename)
                    self.batch_queries = queries
                    self.batch_extra_columns = extra_columns

                    self.batch_processed_var.set(f"Загружено: {len(queries)} запросов")
                    self.batch_found_var.set(f"Строк в файле: {len(df)}")
                    self.batch_filtered_var.set("Готов к поиску (атрибутивный: mi_number)")
                    logger.info(f"Загружено {len(queries)} запросов, доп. колонки: {extra_columns}")

                    for item in self.batch_preview_tree.get_children():
                        self.batch_preview_tree.delete(item)

                    for i, query in enumerate(queries[:100]):
                        row_index = query.get('row_index', i + 1)
                        original = query.get('original_data', {})

                        values = [row_index]
                        id_pu = original.get('Id_ПУ', '') or original.get('id_пу', '') or '-'
                        values.append(id_pu[:20] if id_pu else '-')

                        mi_number = query.get('mi_number', '') or query.get('search_term', '')
                        values.append(mi_number[:20] if mi_number else '-')

                        mit_title = original.get('Модель', '') or original.get('Тип ПУ', '') or ''
                        values.append(mit_title[:40] if mit_title else '-')

                        mit_number = original.get('Тип ПУ', '') or ''
                        values.append(mit_number[:20] if mit_number else '-')

                        year = original.get('Год выпуска', '') or original.get('Год', '') or '-'
                        values.append(year)

                        values.append(original.get('Номер договора', '-'))
                        values.append(original.get('Код ЭДО', '-'))
                        values.append(original.get('Балансовая принадлежность', '-'))
                        values.append(original.get('Эксплуатационная ответственность', '-'))
                        values.append(original.get('МПИ', '-'))

                        self.batch_preview_tree.insert('', tk.END, values=values)

                    if len(queries) > 100:
                        placeholder = [f'... ещё {len(queries) - 100}'] + ['-'] * 10
                        self.batch_preview_tree.insert('', tk.END, values=placeholder)

                    if not queries:
                        messagebox.showwarning("Предупреждение", "Файл загружен, но не найдено поисковых запросов.")

                else:
                    messagebox.showwarning("Предупреждение", "pandas не установлен")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось прочитать файл:\n{e}")
                logger.error(f"Ошибка чтения файла: {e}", exc_info=True)

    def select_all_years_dialog(self):
        """Диалог выбора всех лет"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Выбор лет")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Выберите годы:").pack(pady=5)

        years_frame = ttk.Frame(dialog)
        years_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        for i, year in enumerate(Config.YEARS):
            if year not in self.year_vars:
                self.year_vars[year] = tk.BooleanVar(value=True)
            cb = ttk.Checkbutton(years_frame, text=str(year), variable=self.year_vars[year])
            cb.grid(row=i//5, column=i%5, sticky=tk.W, padx=5, pady=2)

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)

        def select_all():
            for var in self.year_vars.values():
                var.set(True)
            dialog.destroy()

        def deselect_all():
            for var in self.year_vars.values():
                var.set(False)
            dialog.destroy()

        ttk.Button(btn_frame, text="Выбрать все", command=select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Снять все", command=deselect_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Отмена", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def start_batch_search(self):
        """Запуск пакетного поиска"""
        if not self.batch_queries:
            messagebox.showwarning("Предупреждение", "Загрузите файл с запросами")
            return

        selected_years = [year for year, var in self.year_vars.items() if var.get()]
        if not selected_years:
            messagebox.showwarning("Предупреждение", "Выберите хотя бы один год")
            return

        self.batch_running = True
        self.batch_stop_requested = False
        self.batch_start_btn.config(state=tk.DISABLED)
        self.batch_stop_btn.config(state=tk.NORMAL)

        self.batch_results = []

        thread = threading.Thread(target=self._batch_search_thread, args=(selected_years,))
        thread.daemon = True
        thread.start()

    def _batch_search_thread(self, years):
        """Поток пакетного поиска с атрибутивным поиском mi_number"""
        try:
            total_queries = len(self.batch_queries)
            logger.info(f"Запуск пакетного поиска: {total_queries} запросов (атрибутивный mi_number), годы={years}")
            found_count = 0
            filtered_count = 0

            for i, query_data in enumerate(self.batch_queries):
                if self.batch_stop_requested:
                    break

                mi_number = query_data.get('mi_number', '') or query_data.get('search_term', '')
                row_index = query_data.get('row_index', i + 1)
                extra_fields = query_data.get('extra_fields', {})
                original_data = query_data.get('original_data', {})

                if not mi_number:
                    continue

                logger.debug(f"Запрос {i+1}/{total_queries}: mi_number='{mi_number}' (строка {row_index})")

                for year in years:
                    if self.batch_stop_requested:
                        break

                    # АТРИБУТИВНЫЙ ПОИСК по mi_number
                    response = self.api_client.search_vri(mi_number=mi_number, year=year)
                    items = response.get('result', {}).get('items', [])

                    for item in items:
                        mit_title = item.get('mit_title', '')
                        if not VerificationRecord.is_electric_meter(mit_title):
                            filtered_count += 1
                            logger.debug(f"Отфильтровано (не электросчётчик): {mit_title}")
                            continue

                        id_pu = original_data.get('Id_ПУ', '') or original_data.get('id_пу', '')

                        record = VerificationRecord.from_api_response(
                            item,
                            search_query=mi_number,
                            row_index=row_index,
                            id_pu=id_pu,
                            original_data=original_data,
                            extra_fields=extra_fields
                        )
                        self.batch_results.append(record)
                        found_count += 1

                        self.root.after(0, lambda r=record: self._add_result_to_db_tree(r))

                    progress = int((i + 1) / total_queries * 100)
                    self.root.after(0, lambda p=progress: self.batch_progress.configure(value=p))
                    self.root.after(0, lambda c=i+1, t=total_queries: self.batch_processed_var.set(
                        f"Обработано: {c} / {t}"
                    ))
                    self.root.after(0, lambda f=found_count: self.batch_found_var.set(
                        f"Найдено: {f}"
                    ))
                    self.root.after(0, lambda fl=filtered_count: self.batch_filtered_var.set(
                        f"Отфильтровано: {fl}"
                    ))
                    self.root.after(0, lambda d=self.rate_limiter.delay: self.batch_rate_var.set(
                        f"Rate: {d:.2f}с"
                    ))

                    time.sleep(0.1)

            if self.batch_results:
                saved = self.db.save_records(self.batch_results)
                self.root.after(0, lambda: self.batch_processed_var.set("✅ Завершено"))
                self.root.after(0, lambda: self.batch_found_var.set(f"Найдено: {len(self.batch_results)}"))
                self.root.after(0, lambda: self.batch_filtered_var.set(f"Сохранено: {saved}"))
                logger.info(f"✅ Завершено. Найдено {len(self.batch_results)}, сохранено {saved} новых, отфильтровано {filtered_count}")
                self.root.after(0, lambda: self.notebook.select(2))
            else:
                self.root.after(0, lambda: self.batch_processed_var.set("⚠ Завершено"))
                self.root.after(0, lambda: self.batch_found_var.set("Результатов нет"))

        except Exception as e:
            error_msg = f"Ошибка: {e}"
            logger.error(f"Ошибка в пакетном поиске: {e}", exc_info=True)

        finally:
            self.root.after(0, self._finish_batch_search)

    def _add_result_to_db_tree(self, record: VerificationRecord):
        """Добавление результата в таблицу Результаты"""
        title = record.mit_title[:60] + '...' if len(record.mit_title) > 60 else record.mit_title
        ver_date = record.verification_date[:10] if record.verification_date and len(record.verification_date) >= 10 else ''
        valid_date = record.valid_date[:10] if record.valid_date and len(record.valid_date) >= 10 else ''

        self.db_tree.insert('', tk.END, values=(
            record.vri_id,
            record.id_pu,
            record.mit_number,
            title,
            record.mit_notation,
            record.mi_modification,
            record.mi_number,
            ver_date,
            valid_date,
            'Да' if record.applicability else 'Нет',
            record.org_title,
            record.result_docnum,
            record.sticker_num,
            record.manufacturer,
            record.search_query,
            record.row_index
        ))

    def _finish_batch_search(self):
        self.batch_running = False
        self.batch_start_btn.config(state=tk.NORMAL)
        self.batch_stop_btn.config(state=tk.DISABLED)
        self.update_stats()

    def stop_batch_search(self):
        self.batch_stop_requested = True
        logger.info("Остановка пакетного поиска")

    # ========== Одиночный поиск ==========
    def single_search(self):
        """Одиночный поиск с атрибутивным поиском mi_number"""
        search_term = self.search_entry.get().strip()
        if not search_term:
            messagebox.showwarning("Предупреждение", "Введите заводской номер (mi_number)")
            return

        selected_years = [year for year, var in self.search_year_vars.items() if var.get()]
        if not selected_years:
            messagebox.showwarning("Предупреждение", "Выберите хотя бы один год")
            return

        self.search_running = True
        self.search_stop_requested = False
        self.search_start_btn.config(state=tk.DISABLED)
        self.search_stop_btn.config(state=tk.NORMAL)

        start_time = time.time()

        for item in self.single_results_tree.get_children():
            self.single_results_tree.delete(item)

        logger.info(f"Одиночный поиск (атрибутивный mi_number): '{search_term}', годы={selected_years}")
        all_items = []
        self.single_search_results_all = []

        for year in selected_years:
            if self.search_stop_requested:
                self.search_info_var.set("⏹ Поиск остановлен")
                break

            self.search_info_var.set(f"🔍 Поиск... год={year}, найдено={len(all_items)}")
            self.root.update_idletasks()

            start = 0
            rows = 100

            while True:
                if self.search_stop_requested:
                    break

                # АТРИБУТИВНЫЙ ПОИСК по mi_number
                response = self.api_client.search_vri(
                    mi_number=search_term,
                    year=year,
                    start=start,
                    rows=rows
                )
                items = response.get('result', {}).get('items', [])
                count = response.get('result', {}).get('count', 0)

                if not items:
                    break

                all_items.extend(items)
                self.single_search_results_all.extend(items)
                start += rows

                if start >= count or len(all_items) >= 500:
                    break

        elapsed_ms = int((time.time() - start_time) * 1000)

        self.search_running = False
        self.search_start_btn.config(state=tk.NORMAL)
        self.search_stop_btn.config(state=tk.DISABLED)

        self.single_search_results = all_items

        for item in all_items:
            self.single_results_tree.insert('', tk.END, values=(
                item.get('vri_id', ''),
                item.get('mit_number', ''),
                item.get('mit_title', '')[:100],
                item.get('mit_notation', ''),
                item.get('mi_modification', ''),
                item.get('mi_number', ''),
                item.get('verification_date', '')[:10] if item.get('verification_date') else '',
                item.get('valid_date', '')[:10] if item.get('valid_date') else '',
                item.get('org_title', '')[:50],
                item.get('result_docnum', ''),
                item.get('sticker_num', ''),
                'Да' if item.get('applicability') else 'Нет'
            ))

        self.db.log_search(search_term, 'mi_number', selected_years[0] if selected_years else None, len(all_items), elapsed_ms)

        self.single_status_var.set(f"Найдено {len(all_items)} за {elapsed_ms}мс")

        if self.search_stop_requested:
            self.search_info_var.set(f"⏹ Остановлено. Найдено: {len(all_items)}")
        else:
            self.search_info_var.set(f"✅ Найдено: {len(all_items)} за {elapsed_ms}мс")

        msg = f"Найдено {len(all_items)} записей\nВремя: {elapsed_ms}мс"

        dialog = tk.Toplevel(self.root)
        dialog.title("Поиск завершён")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=msg, justify=tk.CENTER).pack(pady=20)

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)

        def save_csv():
            self.save_single_search_csv(all_items, search_term)
            dialog.destroy()

        def add_to_db():
            self.add_single_search_to_db(all_items, search_term)
            dialog.destroy()

        def close():
            dialog.destroy()

        ttk.Button(btn_frame, text="💾 Сохранить CSV", command=save_csv, width=18).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="➕ В БД", command=add_to_db, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="✖ Закрыть", command=close, width=10).pack(side=tk.LEFT, padx=5)

    def save_single_search_csv(self, items: List[Dict], search_term: str):
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для сохранения")
            return

        filename = filedialog.asksaveasfilename(
            title="Сохранить результаты поиска",
            defaultextension=".csv",
            initialfile=f"search_{search_term}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            filetypes=[("CSV files", "*.csv")]
        )

        if filename:
            try:
                records = [VerificationRecord.from_api_response(item, search_term) for item in items]
                data = [r.to_dict() for r in records]
                Exporter.to_csv(data, filename, include_service_fields=True)
                messagebox.showinfo("Экспорт", f"Экспортировано {len(items)} записей в {filename}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось экспортировать:\n{e}")

    def add_single_search_to_db(self, items: List[Dict], search_term: str):
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для добавления")
            return

        records = []
        for i, item in enumerate(items):
            record = VerificationRecord.from_api_response(
                item,
                search_query=search_term,
                row_index=i + 1,
                id_pu="",
                original_data=None
            )
            records.append(record)

        saved = self.db.save_records(records)
        duplicates = len(records) - saved

        msg = f"Добавлено: {saved} новых записей"
        if duplicates > 0:
            msg += f"\nПропущено дублей: {duplicates}"

        messagebox.showinfo("Добавление в БД", msg)
        self.update_stats()

    def stop_single_search(self):
        self.search_stop_requested = True
        self.search_info_var.set("⏹ Остановка поиска...")

    def clear_single_search(self):
        self.search_entry.delete(0, tk.END)
        for item in self.single_results_tree.get_children():
            self.single_results_tree.delete(item)
        self.single_search_results = []
        self.search_info_var.set("Готов к поиску (атрибутивный: mi_number=...)")
        self.single_status_var.set("Готов")

    # ========== Работа с БД ==========
    def load_db_data(self):
        search = self.db_search_var.get().strip()
        year = self.db_year_var.get()
        year = int(year) if year.isdigit() else None
        manufacturer = self.db_manufacturer_var.get().strip()

        for item in self.db_tree.get_children():
            self.db_tree.delete(item)

        results = self.db.search(search, 'substring' if search else 'exact', year, manufacturer, limit=500)

        for row in results:
            self.db_tree.insert('', tk.END, values=(
                row.get('vri_id', ''),
                row.get('id_pu', ''),
                row.get('mit_number', ''),
                row.get('mit_title', '')[:80],
                row.get('mit_notation', ''),
                row.get('mi_modification', ''),
                row.get('mi_number', ''),
                row.get('verification_date', '')[:10] if row.get('verification_date') else '',
                row.get('valid_date', '')[:10] if row.get('valid_date') else '',
                'Да' if row.get('applicability', 1) else 'Нет',
                row.get('org_title', ''),
                row.get('result_docnum', ''),
                row.get('sticker_num', ''),
                row.get('manufacturer', ''),
                row.get('search_query', ''),
                row.get('row_index', '')
            ))

        with sqlite3.connect(self.db.db_path) as conn:
            cursor = conn.execute('SELECT DISTINCT manufacturer FROM verification_records ORDER BY manufacturer')
            manufacturers = [row[0] for row in cursor.fetchall() if row[0]]
        self.db_manufacturer_combo['values'] = [''] + manufacturers

    def reset_db_filters(self):
        self.db_search_var.set('')
        self.db_year_var.set('')
        self.db_manufacturer_var.set('')
        self.load_db_data()

    # ========== Статистика и экспорт ==========
    def update_stats(self):
        stats = self.db.get_stats()

        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete(1.0, tk.END)

        self.stats_text.insert(tk.END, "📊 СТАТИСТИКА БАЗЫ ДАННЫХ\n")
        self.stats_text.insert(tk.END, "="*50 + "\n")
        self.stats_text.insert(tk.END, f"Всего записей: {stats['total']}\n")
        if stats['total']:
            self.stats_text.insert(tk.END, f"Пригодных: {stats['applicable']} ({stats['applicable']/stats['total']*100:.1f}%)\n")
            self.stats_text.insert(tk.END, f"Непригодных: {stats['inapplicable']}\n\n")
        else:
            self.stats_text.insert(tk.END, "Пригодных: 0\nНепригодных: 0\n\n")

        if stats['by_year']:
            self.stats_text.insert(tk.END, "📅 ПО ГОДАМ:\n")
            for year, count in sorted(stats['by_year'].items()):
                pct = (count/stats['total']*100) if stats['total'] else 0
                self.stats_text.insert(tk.END, f"  {year}: {count} ({pct:.1f}%)\n")
            self.stats_text.insert(tk.END, "\n")

        if stats['by_manufacturer']:
            self.stats_text.insert(tk.END, "🏭 ПО ПРОИЗВОДИТЕЛЯМ (ТОП-15):\n")
            for man, count in list(stats['by_manufacturer'].items())[:15]:
                pct = (count/stats['total']*100) if stats['total'] else 0
                self.stats_text.insert(tk.END, f"  {man}: {count} ({pct:.1f}%)\n")

        self.stats_text.config(state=tk.DISABLED)

    def clean_duplicates(self):
        deleted = self.db.clean_duplicates()
        messagebox.showinfo("Дедупликация", f"Удалено дубликатов: {deleted}")
        self.update_stats()

    def export_csv(self):
        filename = filedialog.asksaveasfilename(title="Сохранить как CSV", defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if filename:
            try:
                with sqlite3.connect(self.db.db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.execute('SELECT * FROM verification_records ORDER BY verification_date DESC')
                    data = [dict(row) for row in cursor.fetchall()]

                Exporter.to_csv(data, filename, include_service_fields=True, extra_columns=self.batch_extra_columns)
                messagebox.showinfo("Экспорт", f"Данные экспортированы в {filename}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось экспортировать: {e}")

    def export_excel(self):
        if not PANDAS_AVAILABLE:
            messagebox.showwarning("Предупреждение", "pandas не установлен")
            return

        filename = filedialog.asksaveasfilename(title="Сохранить как Excel", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            try:
                with sqlite3.connect(self.db.db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.execute('SELECT * FROM verification_records ORDER BY verification_date DESC')
                    data = [dict(row) for row in cursor.fetchall()]

                Exporter.to_excel(data, filename, include_service_fields=True, extra_columns=self.batch_extra_columns)
                messagebox.showinfo("Экспорт", f"Данные экспортированы в {filename}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось экспортировать: {e}")

    def export_stats_json(self):
        stats = self.db.get_stats()
        filename = Exporter.export_stats(stats)
        messagebox.showinfo("Экспорт", f"Статистика экспортирована в {filename}")

    # ========== Вспомогательные ==========
    def show_search_history(self):
        history = self.db.get_search_history(50)

        dialog = tk.Toplevel(self.root)
        dialog.title("История поиска")
        dialog.geometry("700x400")
        dialog.transient(self.root)

        text = scrolledtext.ScrolledText(dialog, height=20, font=('Consolas', 10))
        text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        for record in history:
            line = f"{record['search_date'][:19]} | {record['query'][:30]:30} | {record['results_count']:4} рез. | {record['duration_ms']:4}мс | тип={record['query_type']}\n"
            text.insert(tk.END, line)

        text.config(state=tk.DISABLED)

    def show_log_file(self):
        if os.path.exists(LOG_FILE):
            if sys.platform == 'win32':
                os.startfile(LOG_FILE)
            elif sys.platform == 'darwin':
                os.system(f'open "{LOG_FILE}"')
            else:
                os.system(f'xdg-open "{LOG_FILE}"')
        else:
            messagebox.showinfo("Лог", f"Файл логов не найден: {LOG_FILE}")

    def show_about(self):
        messagebox.showinfo("О программе",
            "ФГИС АРШИН — Поиск счётчиков электроэнергии\n"
            "Версия 1.2\n\n"
            "Изменения:\n"
            "• Атрибутивный поиск (mi_number=...)\n"
            "• Динамический rate limiting (0.1-0.5 сек)\n"
            "• Исправлено кодирование search\n"
            "• Парсинг ошибок API (status, message, requestId)\n"
            "• sort=mi_number+desc по умолчанию\n"
            "• Сохранение ВСЕХ колонок из Excel\n"
            "• Фильтрация только счётчиков электроэнергии"
        )

    # ========== Управление фильтрами ==========
    def load_exact_queries_display(self):
        self.exact_queries_list.delete(1.0, tk.END)
        for q in EXACT_QUERIES:
            self.exact_queries_list.insert(tk.END, q + '\n')
        self.filters_status_var.set(f"Точных запросов: {len(EXACT_QUERIES)}")

    def load_exact_queries_file(self):
        filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if filename:
            global EXACT_QUERIES
            EXACT_QUERIES = Config.load_exact_queries()
            self.load_exact_queries_display()

    def save_exact_queries_file(self):
        filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if filename:
            with open(filename, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['query'])
                for q in EXACT_QUERIES:
                    writer.writerow([q])
            messagebox.showinfo("Сохранение", f"Точные запросы сохранены в {filename}")

    def clear_exact_queries(self):
        global EXACT_QUERIES
        EXACT_QUERIES = []
        self.load_exact_queries_display()

    def add_exact_query(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Добавить запрос")
        dialog.geometry("300x100")
        dialog.transient(self.root)

        ttk.Label(dialog, text="Запрос:").pack(pady=5)
        entry = ttk.Entry(dialog, width=40)
        entry.pack(pady=5)

        def add():
            global EXACT_QUERIES
            q = entry.get().strip()
            if q:
                EXACT_QUERIES.append(q)
                self.load_exact_queries_display()
            dialog.destroy()

        ttk.Button(dialog, text="Добавить", command=add).pack(pady=5)

    def load_manufacturers_display(self):
        self.manufacturers_list.delete(1.0, tk.END)
        for keyword, manufacturer in MANUFACTURERS_RULES.items():
            self.manufacturers_list.insert(tk.END, f"{keyword} → {manufacturer}\n")
        self.filters_status_var.set(f"Производителей: {len(MANUFACTURERS_RULES)}")

    def load_manufacturers_file(self):
        filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if filename:
            global MANUFACTURERS_RULES
            MANUFACTURERS_RULES = Config.load_manufacturers()
            self.load_manufacturers_display()

    def save_manufacturers_file(self):
        filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if filename:
            with open(filename, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['keyword', 'manufacturer'])
                for keyword, manufacturer in MANUFACTURERS_RULES.items():
                    writer.writerow([keyword, manufacturer])
            messagebox.showinfo("Сохранение", f"Производители сохранены в {filename}")

    def clear_manufacturers(self):
        global MANUFACTURERS_RULES
        MANUFACTURERS_RULES = {}
        self.load_manufacturers_display()

    def add_manufacturer(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Добавить производителя")
        dialog.geometry("350x150")
        dialog.transient(self.root)

        ttk.Label(dialog, text="Ключевое слово:").pack(pady=2)
        keyword_entry = ttk.Entry(dialog, width=40)
        keyword_entry.pack(pady=2)

        ttk.Label(dialog, text="Производитель:").pack(pady=2)
        mfg_entry = ttk.Entry(dialog, width=40)
        mfg_entry.pack(pady=2)

        def add():
            global MANUFACTURERS_RULES
            keyword = keyword_entry.get().strip().lower()
            manufacturer = mfg_entry.get().strip()
            if keyword and manufacturer:
                MANUFACTURERS_RULES[keyword] = manufacturer
                self.load_manufacturers_display()
            dialog.destroy()

        ttk.Button(dialog, text="Добавить", command=add).pack(pady=5)

    def on_close(self):
        """Закрытие приложения"""
        logger.info("Закрытие приложения")
        self.api_client.close()
        self.root.destroy()


# ============== ЗАПУСК ==============
if __name__ == "__main__":
    logger.info("="*60)
    logger.info("ЗАПУСК ПРИЛОЖЕНИЯ ФГИС АРШИН v1.2")
    logger.info("="*60)

    root = tk.Tk()
    app = ArshinApp(root)
    root.mainloop()
