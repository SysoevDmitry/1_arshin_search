#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
УНИВЕРСАЛЬНОЕ ПРИЛОЖЕНИЕ ДЛЯ РАБОТЫ С ФГИС АРШИН
Версия 3.0 — объединение v2.1 + v1.2

Возможности:
- Атрибутивный поиск (mi_number=, mit_number=) согласно спецификации v.2.2
- Динамический rate limiting с адаптацией к HTTP 429
- Пакетная обработка из Excel с сохранением связей
- Промышленный сбор данных (асинхронный)
- Расширенное логирование для отладки
- Экспорт с полной служебной информацией
- Сохранение всех колонок из Excel (extra_fields)
"""

import os
import sys
import json
import csv
import sqlite3
import asyncio
import aiohttp
import requests
import threading
import time
import logging
import re
import webbrowser
from datetime import datetime
from typing import Dict, List, Set, Optional, Tuple, Any
from collections import defaultdict
from dataclasses import dataclass, asdict, field
from urllib.parse import quote, urlencode
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tkinter.font import Font

# Попытка импорта дополнительных библиотек
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logging.warning("pandas/openpyxl не установлены - экспорт в Excel недоступен")

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

# ============== РАСШИРЕННОЕ ЛОГИРОВАНИЕ ==============
def setup_logging(debug_mode: bool = False):
    """Настройка расширенного логирования"""
    log_dir = "logs"
    os.makedirs(log_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(log_dir, f"arshin_app_{timestamp}.log")
    
    # Формат с детальной информацией
    log_format = logging.Formatter(
        '%(asctime)s.%(msecs)03d | %(levelname)-8s | %(name)s | %(funcName)s:%(lineno)d | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # Файловый обработчик (всегда включен)
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setFormatter(log_format)
    file_handler.setLevel(logging.DEBUG)
    
    # Консольный обработчик
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    console_handler.setLevel(logging.DEBUG if debug_mode else logging.INFO)
    
    # Настройка корневого логгера
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.DEBUG)
    root_logger.addHandler(file_handler)
    root_logger.addHandler(console_handler)
    
    logger = logging.getLogger(__name__)
    logger.info(f"Логирование инициализировано: {log_file}")
    logger.debug(f"Режим отладки: {'ВКЛЮЧЕН' if debug_mode else 'ВЫКЛЮЧЕН'}")
    
    return logger, log_file

# Инициализация логирования
logger, LOG_FILE = setup_logging(debug_mode=True)

# ============== КОНФИГУРАЦИЯ ==============
class Config:
    """Настройки приложения с загрузкой из файлов"""
    
    # API ФГИС АРШИН (согласно документации)
    API_BASE_URL = "https://fgis.gost.ru/fundmetrology/eapi"
    API_VRI = f"{API_BASE_URL}/vri"  # Реестр поверок
    API_MIT = f"{API_BASE_URL}/mit"  # Реестр типов СИ
    
    # Параметры запросов (согласно спецификации API)
    MAX_ROWS_PER_REQUEST = 100  # Максимум по API
    DEFAULT_ROWS = 50
    REQUEST_TIMEOUT = 30
    RETRY_ATTEMPTS = 3
    RETRY_DELAY = 2
    
    # Динамический rate limiting
    RATE_LIMIT_MIN = 0.1       # Минимальная пауза (сек)
    RATE_LIMIT_MAX = 0.5       # Максимальная пауза (сек)
    RATE_LIMIT_START = 0.1     # Стартовая пауза
    RATE_LIMIT_STEP = 0.1      # Шаг изменения
    RATE_LIMIT_429_COOLDOWN = 10   # Секунды без ошибок для ускорения
    RATE_LIMIT_ERROR_WINDOW = 60   # Секунды без ошибок для замедления
    
    # База данных
    DB_PATH = "arshin_data.db"
    
    # Пути к файлам
    EXPORT_DIR = "exports"
    CONFIG_DIR = "config"
    
    # Файлы конфигурации
    EXACT_QUERIES_FILE = os.path.join(CONFIG_DIR, "exact_queries.csv")
    MANUFACTURERS_FILE = os.path.join(CONFIG_DIR, "manufacturers.csv")
    
    # Годы для фильтрации
    YEARS = list(range(2010, 2027))
    
    # Размеры окна (оптимизировано для 1200x700)
    WINDOW_WIDTH = 1180
    WINDOW_HEIGHT = 680
    
    @classmethod
    def load_exact_queries(cls) -> List[str]:
        """Загрузка точных запросов из файла"""
        queries = []
        if os.path.exists(cls.EXACT_QUERIES_FILE):
            try:
                with open(cls.EXACT_QUERIES_FILE, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if 'query' in row and row['query'].strip():
                            queries.append(row['query'].strip())
                logger.info(f"Загружено {len(queries)} точных запросов из {cls.EXACT_QUERIES_FILE}")
            except Exception as e:
                logger.error(f"Ошибка загрузки точных запросов: {e}")
        else:
            logger.warning(f"Файл точных запросов не найден: {cls.EXACT_QUERIES_FILE}")
        return queries
    
    @classmethod
    def load_manufacturers(cls) -> Dict[str, str]:
        """Загрузка словаря производителей из файла"""
        manufacturers = {}
        if os.path.exists(cls.MANUFACTURERS_FILE):
            try:
                with open(cls.MANUFACTURERS_FILE, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if 'keyword' in row and 'manufacturer' in row:
                            manufacturers[row['keyword'].lower().strip()] = row['manufacturer'].strip()
                logger.info(f"Загружено {len(manufacturers)} правил для производителей")
            except Exception as e:
                logger.error(f"Ошибка загрузки производителей: {e}")
        else:
            logger.warning(f"Файл производителей не найден: {cls.MANUFACTURERS_FILE}")
        return manufacturers

# Глобальные конфигурационные данные
EXACT_QUERIES = Config.load_exact_queries()
MANUFACTURERS_RULES = Config.load_manufacturers()

# ============== ДИНАМИЧЕСКИЙ RATE LIMITER ==============
class DynamicRateLimiter:
    """
    Адаптивный контроллер частоты запросов к API.
    
    Логика:
    - Старт с RATE_LIMIT_START (0.1 сек)
    - При получении 429: увеличивать интервал на STEP до MAX
    - Если 429 получены 2+ подряд: сразу jump до MAX
    - Если 10 сек без ошибок: уменьшать на STEP до MIN
    - Если 60 сек без ошибок: уменьшать на STEP до MIN
    """

    def __init__(self):
        self.current_delay = Config.RATE_LIMIT_START
        self.consecutive_429 = 0
        self.last_429_time = 0.0
        self.last_request_time = 0.0
        self.lock = threading.Lock()

    def wait(self):
        with self.lock:
            now = time.time()
            elapsed = now - self.last_request_time
            wait_time = max(0, self.current_delay - elapsed)
        if wait_time > 0:
            time.sleep(wait_time)
        with self.lock:
            self.last_request_time = time.time()

    def on_429(self):
        with self.lock:
            self.consecutive_429 += 1
            self.last_429_time = time.time()
            if self.consecutive_429 >= 2:
                self.current_delay = Config.RATE_LIMIT_MAX
            else:
                self.current_delay = min(Config.RATE_LIMIT_MAX, self.current_delay + Config.RATE_LIMIT_STEP)
            logger.warning(f"API 429 — задержка увеличена до {self.current_delay:.2f}с")

    def on_success(self):
        with self.lock:
            self.consecutive_429 = 0
            now = time.time()
            time_since_429 = now - self.last_429_time
            if time_since_429 >= Config.RATE_LIMIT_ERROR_WINDOW and self.current_delay > Config.RATE_LIMIT_MIN:
                self.current_delay = max(Config.RATE_LIMIT_MIN, self.current_delay - Config.RATE_LIMIT_STEP)
            elif time_since_429 >= Config.RATE_LIMIT_429_COOLDOWN and self.current_delay > Config.RATE_LIMIT_MIN:
                self.current_delay = max(Config.RATE_LIMIT_MIN, self.current_delay - Config.RATE_LIMIT_STEP)

    @property
    def delay(self) -> float:
        with self.lock:
            return self.current_delay

# ============== МОДЕЛИ ДАННЫХ ==============
@dataclass
class VerificationRecord:
    """Запись о поверке СИ с полной служебной информацией (оптимизированная)"""

    # Основные данные из API
    vri_id: str = ""                    # Идентификатор записи (ключевое поле)
    mit_number: str = ""                # Номер в реестре типов СИ
    mit_title: str = ""                 # Наименование типа СИ
    mit_notation: str = ""              # Обозначение типа СИ
    mi_modification: str = ""           # Модификация СИ
    mi_number: str = ""                 # Заводской/серийный номер
    verification_date: str = ""         # Дата поверки (ISO 8601)
    valid_date: str = ""                # Действительна до (ISO 8601)
    applicability: bool = True          # Пригодность
    org_title: str = ""                 # Организация-поверитель
    result_docnum: str = ""             # Номер документа о поверке
    sticker_num: str = ""               # № наклейки

    # Служебные поля (добавляются приложением)
    manufacturer: str = ""              # Определенный производитель
    collected_at: str = ""              # Время сохранения

    # Связи (для пакетного поиска)
    search_query: str = ""              # Исходный поисковый запрос
    row_index: int = 0                  # Индекс строки в исходном файле
    id_pu: str = ""                     # Id_ПУ из исходного Excel файла
    
    # Данные из исходного Excel файла
    contract_number: str = ""           # Номер договора
    edo_code: str = ""                  # Код ЭДО
    balance_owner: str = ""             # Балансовая принадлежность
    operation_responsibility: str = ""  # Эксплуатационная ответственность
    mpi: str = ""                       # МПИ
    
    # Все дополнительные колонки из Excel (динамические)
    extra_fields: Dict[str, str] = field(default_factory=dict)

    # Статический метод для генерации URL (не хранится в БД)
    @staticmethod
    def generate_url(vri_id: str) -> str:
        """Генерация URL записи в системе ФГИС АРШИН"""
        return f"https://fgis.gost.ru/fundmetrology/cm/erts/?id={vri_id}"

    @property
    def record_url(self) -> str:
        """URL записи (генерируется динамически из vri_id)"""
        return self.generate_url(self.vri_id)

    @classmethod
    def from_api_response(cls, data: dict, search_query: str = "",
                          row_index: int = 0, id_pu: str = "",
                          original_data: dict = None,
                          extra_fields: Dict[str, str] = None) -> 'VerificationRecord':
        """Создание записи из ответа API с сохранением связей"""
        original = original_data or {}
        extras = extra_fields or {}
        
        record = cls(
            vri_id=data.get('vri_id', ''),
            mit_number=data.get('mit_number', ''),
            mit_title=data.get('mit_title', ''),
            mit_notation=data.get('mit_notation', ''),
            mi_modification=data.get('mi_modification', ''),
            mi_number=data.get('mi_number', ''),
            verification_date=data.get('verification_date', ''),
            valid_date=data.get('valid_date', ''),
            applicability=data.get('applicability', True),
            org_title=data.get('org_title', ''),
            result_docnum=data.get('result_docnum', ''),
            sticker_num=data.get('sticker_num', ''),
            search_query=search_query,
            row_index=row_index,
            id_pu=id_pu,
            contract_number=original.get('Номер договора', original.get('contract_number', '')),
            edo_code=original.get('Код ЭДО', original.get('edo_code', '')),
            balance_owner=original.get('Балансовая принадлежность', original.get('balance_owner', '')),
            operation_responsibility=original.get('Эксплуатационная ответственность', original.get('operation_responsibility', '')),
            mpi=original.get('МПИ', original.get('mpi', '')),
            extra_fields=extras,
            collected_at=datetime.now().isoformat()
        )

        # Автоматическое определение производителя
        record.manufacturer = cls._detect_manufacturer(record.mit_title)

        logger.debug(f"Создана запись: vri_id={record.vri_id}, query={search_query}")
        return record

    @staticmethod
    def _detect_manufacturer(title: str) -> str:
        """Определение производителя по названию с использованием загруженных правил"""
        if not title:
            return 'Другие'

        title_lower = title.lower()

        # Используем загруженные правила из файла
        for keyword, manufacturer in MANUFACTURERS_RULES.items():
            if keyword in title_lower:
                return manufacturer

        # Резервные правила (если файл не загружен)
        fallback_rules = {
            'меркурий': 'Меркурий',
            'нева': 'Нева',
            'энергомера': 'Энергомера',
        }

        for key, value in fallback_rules.items():
            if key in title_lower:
                return value

        return 'Другие'

    @staticmethod
    def is_electric_meter(title: str) -> bool:
        """
        Проверка: относится ли запись к приборам учета электрической энергии

        Возвращает True только для счетчиков электроэнергии
        Возвращает False для всех остальных приборов (вода, газ, тепло, медицинские, весы и т.д.)
        """
        if not title:
            return False

        title_lower = title.lower()

        # ===== СПИСОК ИСКЛЮЧЕНИЙ (не электрическая энергия) =====
        exclude_keywords = [
            # Вода
            'воды', 'водомер', 'водосчетчик', 'холодной воды', 'горячей воды',
            'счетчик воды', 'крыльчатые',

            # Газ
            'газа', 'газ', 'газосчетчик', 'газовые', 'бытовые газовые', 'диафрагменные',
            'счетчик газа',

            # Тепло
            'тепл', 'теплопотребление', 'теплопотребления', 'тепловычислитель',
            'вычислитель количества теплоты', 'распределения теплопотребления',

            # Медицинские приборы
            'термометры медицинские', 'термометр медицинский', 'медицинские максимальные',
            'ртутные стеклянные медицинские', 'максимальные стеклянные ртутные',

            # Весы
            'весы', 'взвешивания', 'для новорожденных', 'почтовые электронные',
            'рычажные настольные',

            # Давление
            'манометры', 'вакуумметры', 'мановакуумметры', 'напоромеры', 'тягомеры',
            'тягонапоромеры', 'обм',

            # Температура (не медицинские)
            'термометры ртутные', 'термометры стеклянные', 'термометры лабораторные',
            'термопреобразователи', 'термометры сопротивления',

            # Газовые сигнализаторы
            'сигнализаторы', 'метана', 'с внешним сенсором',

            # Измерительные приборы общего назначения
            'штангенглубиномеры', 'нутромеры', 'ареометры', 'гигрометры',
            'психрометрические', 'мегаомметры', 'приемники-ловушки',
            'счетчики времени наработки', 'бутирометры', 'поверка си массы',

            # Расходомеры (не электрические)
            'расхода электромагнитные', 'преобразователи расхода',

            # Прочие исключения
            'устройства для распределения', 'для статического взвешивания',
            'транcформаторы тока', 'трансформаторы тока до 0,4 кв',
        ]

        # Проверка исключений
        for keyword in exclude_keywords:
            if keyword in title_lower:
                logger.debug(f"Исключено (не электросчетчик): {title}")
                return False

        # ===== СПИСОК ВКЛЮЧЕНИЙ (электрическая энергия) =====
        include_keywords = [
            # Счетчики электрической энергии
            'электрической энергии', 'электроэнергии', 'электросчетчик',
            'счетчик электрической', 'счетчик электроэнергии',
            'активной электроэнергии', 'реактивной электроэнергии',
            'однофазный', 'трехфазный', 'статический', 'индукционный',

            # Специфичные модели
            'меркурий', 'нева', 'энергомера', 'матрица', 'альфа', 'милур',
            'псч', 'сэт', 'цэ', 'се 101', 'се 102', 'се 201', 'се 301',
            'меркурий 20', 'меркурий 23', 'нева 10', 'нева 30',

            # Трансформаторы тока для электроэнергии
            'трансформатор тока измерительный', 'трансформатор тока 0,4 кв',
        ]

        # Проверка включений
        for keyword in include_keywords:
            if keyword in title_lower:
                logger.debug(f"Включено (электросчетчик): {title}")
                return True

        # По умолчанию - не электросчетчик
        logger.debug(f"Не определено как электросчетчик: {title}")
        return False

    def to_dict(self) -> dict:
        """Преобразование в словарь с полной служебной информацией"""
        result = asdict(self)
        result['record_url'] = self.record_url
        # Распаковка дополнительных колонок из Excel (не перезаписывая основные)
        for k, v in self.extra_fields.items():
            if k not in result or not result[k]:
                result[k] = v
        return result


@dataclass
class SearchQuery:
    """Поисковый запрос с метаданными"""
    search_term: str
    query_type: str  # 'exact', 'substring', 'prefix'
    year: Optional[int] = None
    source: str = ""  # Источник: 'manual', 'file', 'config'
    row_index: int = 0


# ============== БАЗА ДАННЫХ ==============
class Database:
    """Работа с SQLite базой данных с расширенным логированием"""

    def __init__(self, db_path: str = Config.DB_PATH):
        self.db_path = db_path
        logger.debug(f"Инициализация БД: {db_path}")
        self._init_db()

    def _init_db(self):
        """Инициализация структуры БД с индексами и миграцией"""
        logger.debug("Инициализация таблиц БД")

        with sqlite3.connect(self.db_path) as conn:
            # Основная таблица для результатов поверки (оптимизированная)
            # Удалены дублирующие поля: year, month, record_url, search_year
            conn.execute('''
                CREATE TABLE IF NOT EXISTS verification_records (
                    vri_id TEXT PRIMARY KEY,
                    mit_number TEXT,
                    mit_title TEXT,
                    mit_notation TEXT,
                    mi_modification TEXT,
                    mi_number TEXT,
                    verification_date TEXT,
                    valid_date TEXT,
                    applicability INTEGER,
                    org_title TEXT,
                    result_docnum TEXT,
                    sticker_num TEXT,
                    manufacturer TEXT,
                    search_query TEXT,
                    row_index INTEGER,
                    id_pu TEXT,
                    contract_number TEXT,
                    edo_code TEXT,
                    balance_owner TEXT,
                    operation_responsibility TEXT,
                    mpi TEXT,
                    extra_fields TEXT,
                    collected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # Таблица для типов СИ
            conn.execute('''
                CREATE TABLE IF NOT EXISTS measurement_types (
                    mit_id TEXT PRIMARY KEY,
                    number TEXT,
                    title TEXT,
                    notation TEXT,
                    status TEXT,
                    manufacturer TEXT,
                    country TEXT,
                    collected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # Таблица для истории поиска с деталями
            conn.execute('''
                CREATE TABLE IF NOT EXISTS search_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    query TEXT,
                    query_type TEXT,
                    year INTEGER,
                    results_count INTEGER,
                    search_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    duration_ms INTEGER,
                    source TEXT
                )
            ''')

            # Миграция: добавление новых полей если их нет
            self._migrate_database(conn)

            # Индексы для ускорения поиска
            conn.execute('CREATE INDEX IF NOT EXISTS idx_applicability ON verification_records(applicability)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_manufacturer ON verification_records(manufacturer)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_verification_date ON verification_records(verification_date)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_search_query ON verification_records(search_query)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_mi_number ON verification_records(mi_number)')

            logger.info("БД инициализирована успешно")

    def _migrate_database(self, conn):
        """Добавление новых полей в существующую БД и удаление устаревших"""
        logger.debug("Проверка миграции БД")

        # Получаем список существующих колонок
        cursor = conn.execute("PRAGMA table_info(verification_records)")
        existing_columns = {row[1] for row in cursor.fetchall()}

        # Новые поля для добавления
        new_columns = [
            ('search_query', 'TEXT'),
            ('row_index', 'INTEGER'),
            ('id_pu', 'TEXT'),
            ('sticker_num', 'TEXT'),
            ('contract_number', 'TEXT'),
            ('edo_code', 'TEXT'),
            ('balance_owner', 'TEXT'),
            ('operation_responsibility', 'TEXT'),
            ('mpi', 'TEXT'),
            ('extra_fields', 'TEXT')
        ]

        for column_name, column_type in new_columns:
            if column_name not in existing_columns:
                logger.info(f"Добавление колонки {column_name} в verification_records")
                conn.execute(f'ALTER TABLE verification_records ADD COLUMN {column_name} {column_type}')
                existing_columns.add(column_name)
            else:
                logger.debug(f"Колонка {column_name} уже существует")

        # Удаляем устаревшие колонки (year, month, record_url, search_year)
        # SQLite не поддерживает DROP COLUMN напрямую, помечаем как устаревшие
        deprecated_columns = ['year', 'month', 'record_url', 'search_year']
        for col in deprecated_columns:
            if col in existing_columns:
                logger.info(f"Колонка {col} помечена как устаревшая (не используется)")

        logger.debug("Миграция БД завершена")

    def save_records(self, records: List[VerificationRecord]) -> int:
        """Пакетное сохранение записей с проверкой на полные дубли"""
        if not records:
            return 0

        saved = 0
        duplicates = 0
        errors = 0

        logger.debug(f"Сохранение {len(records)} записей в БД")

        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()

            for record in records:
                try:
                    # Проверка на полный дубль (все основные поля + служебные)
                    cursor.execute('''
                        SELECT COUNT(*) FROM verification_records
                        WHERE vri_id = ?
                          AND mit_number = ?
                          AND mit_title = ?
                          AND mi_number = ?
                          AND verification_date = ?
                          AND valid_date = ?
                          AND applicability = ?
                          AND org_title = ?
                          AND result_docnum = ?
                          AND sticker_num = ?
                          AND manufacturer = ?
                          AND search_query = ?
                          AND row_index = ?
                          AND id_pu = ?
                    ''', (
                        record.vri_id, record.mit_number, record.mit_title,
                        record.mi_number, record.verification_date, record.valid_date,
                        1 if record.applicability else 0,
                        record.org_title, record.result_docnum,
                        record.sticker_num,
                        record.manufacturer,
                        record.search_query, record.row_index, record.id_pu
                    ))

                    is_duplicate = cursor.fetchone()[0] > 0

                    if is_duplicate:
                        duplicates += 1
                        logger.debug(f"Пропущен дубль: vri_id={record.vri_id}, query={record.search_query}")
                    else:
                        # Вставка новой записи
                        cursor.execute('''
                            INSERT OR IGNORE INTO verification_records
                            (vri_id, mit_number, mit_title, mit_notation, mi_modification,
                             mi_number, verification_date, valid_date, applicability,
                             org_title, result_docnum, sticker_num, manufacturer,
                             search_query, row_index, id_pu,
                             contract_number, edo_code, balance_owner,
                             operation_responsibility, mpi, extra_fields)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            record.vri_id, record.mit_number, record.mit_title,
                            record.mit_notation, record.mi_modification, record.mi_number,
                            record.verification_date, record.valid_date,
                            1 if record.applicability else 0,
                            record.org_title, record.result_docnum,
                            record.sticker_num,
                            record.manufacturer,
                            record.search_query, record.row_index, record.id_pu,
                            record.contract_number, record.edo_code, record.balance_owner,
                            record.operation_responsibility, record.mpi,
                            json.dumps(record.extra_fields, ensure_ascii=False) if record.extra_fields else ''
                        ))

                        if cursor.rowcount > 0:
                            saved += 1

                except Exception as e:
                    errors += 1
                    logger.error(f"Ошибка сохранения записи {record.vri_id}: {e}")

            conn.commit()

        logger.info(f"Сохранено: {saved} новых, {duplicates} дубликатов, {errors} ошибок")
        return saved

    def get_existing_ids(self) -> Set[str]:
        """Получить все сохраненные ID"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute('SELECT vri_id FROM verification_records')
            ids = {row[0] for row in cursor.fetchall()}
            logger.debug(f"В БД найдено {len(ids)} существующих записей")
            return ids

    def get_stats(self) -> Dict:
        """Получить детальную статистику по базе"""
        logger.debug("Запрос статистики БД")

        with sqlite3.connect(self.db_path) as conn:
            total = conn.execute('SELECT COUNT(*) FROM verification_records').fetchone()[0]

            if total == 0:
                return {
                    'total': 0, 'by_year': {}, 'by_manufacturer': {},
                    'applicable': 0, 'inapplicable': 0
                }

            # Год извлекается из verification_date (SQLite strftime)
            by_year = dict(conn.execute(
                "SELECT strftime('%Y', verification_date) as year, COUNT(*) "
                "FROM verification_records WHERE verification_date != '' "
                "GROUP BY year ORDER BY year"
            ).fetchall())

            by_manufacturer = dict(conn.execute(
                'SELECT manufacturer, COUNT(*) FROM verification_records GROUP BY manufacturer ORDER BY COUNT(*) DESC'
            ).fetchall())

            applicable = conn.execute('SELECT COUNT(*) FROM verification_records WHERE applicability = 1').fetchone()[0]

            return {
                'total': total,
                'by_year': by_year,
                'by_manufacturer': by_manufacturer,
                'applicable': applicable,
                'inapplicable': total - applicable
            }

    def search(self, query: str = "", query_type: str = 'substring',
               year: Optional[int] = None, manufacturer: str = "",
               limit: int = 500, offset: int = 0) -> List[Dict]:
        """Расширенный поиск в локальной БД"""
        logger.debug(f"Поиск в БД: query='{query}', type={query_type}, year={year}, mfg={manufacturer}")

        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            params = []

            sql = 'SELECT * FROM verification_records WHERE 1=1'

            if query:
                if query_type == 'exact':
                    sql += ' AND (mit_number = ? OR mi_number = ? OR vri_id = ?)'
                    params.extend([query, query, query])
                else:
                    sql += ' AND (mit_number LIKE ? OR mi_number LIKE ? OR mit_title LIKE ? OR org_title LIKE ?)'
                    pattern = f'%{query}%'
                    params.extend([pattern, pattern, pattern, pattern])

            # Фильтр по году через извлечение из verification_date
            if year:
                sql += " AND strftime('%Y', verification_date) = ?"
                params.append(str(year))

            if manufacturer:
                sql += ' AND manufacturer = ?'
                params.append(manufacturer)

            sql += ' ORDER BY verification_date DESC LIMIT ? OFFSET ?'
            params.extend([limit, offset])

            cursor = conn.execute(sql, params)
            results = [dict(row) for row in cursor.fetchall()]
            logger.debug(f"Найдено {len(results)} записей")
            return results

    def log_search(self, query: str, query_type: str, year: Optional[int], 
                   results_count: int, duration_ms: int, source: str = "manual"):
        """Логирование поиска в историю"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                'INSERT INTO search_history (query, query_type, year, results_count, duration_ms, source) VALUES (?, ?, ?, ?, ?, ?)',
                (query, query_type, year, results_count, duration_ms, source)
            )
            conn.commit()
    
    def get_search_history(self, limit: int = 50) -> List[Dict]:
        """Получение истории поиска"""
        with sqlite3.connect(self.db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.execute(
                'SELECT * FROM search_history ORDER BY search_date DESC LIMIT ?',
                (limit,)
            )
            return [dict(row) for row in cursor.fetchall()]
    
    def clean_duplicates(self) -> int:
        """Удаление дубликатов"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute('''
                DELETE FROM verification_records
                WHERE rowid NOT IN (
                    SELECT MIN(rowid) FROM verification_records GROUP BY vri_id
                )
            ''')
            conn.commit()
            deleted = cursor.rowcount
        logger.info(f"Удалено дубликатов: {deleted}")
        return deleted


# ============== КЛИЕНТ API (согласно документации) ==============
class APIClient:
    """
    Клиент для работы с API ФГИС АРШИН
    v3.0 — атрибутивный поиск, rate limiting, согласно спецификации v.2.2
    """

    def __init__(self, rate_limiter: 'DynamicRateLimiter' = None):
        self.base_url = Config.API_BASE_URL
        self.session = requests.Session()
        self.rate_limiter = rate_limiter or DynamicRateLimiter()
        
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Cache-Control': 'no-cache',
            'Pragma': 'no-cache'
        })
        
        logger.info(f"APIClient инициализирован: {self.base_url}")

    @staticmethod
    def encode_search_value(value: str) -> str:
        """
        Кодирование поискового значения: пробелы → '?', кириллица не кодируется
        """
        result = value.replace(' ', '?')
        result = quote(result, safe='*?абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ-_.')
        return result

    def _build_vri_params(self, mi_number: str = None, mit_number: str = None,
                          search_term: str = None, year: int = None,
                          start: int = 0, rows: int = Config.DEFAULT_ROWS) -> dict:
        """
        Формирование параметров запроса к /vri.
        Приоритет: атрибутивный поиск (mi_number/mit_number) > search
        """
        rows = min(rows, Config.MAX_ROWS_PER_REQUEST)
        rows = max(1, rows)
        start = max(0, start)

        params = {
            'start': start,
            'rows': rows,
            'sort': 'mi_number desc'
        }

        if year:
            params['year'] = year

        if mi_number:
            params['mi_number'] = mi_number
        elif mit_number:
            params['mit_number'] = mit_number
        elif search_term:
            params['search'] = self.encode_search_value(search_term)

        return params

    def search_vri(self, mi_number: str = None, mit_number: str = None,
                   search_term: str = None, year: Optional[int] = None,
                   start: int = 0, rows: int = Config.DEFAULT_ROWS) -> Dict:
        """
        Поиск в реестре поверок (VRI) — атрибутивный поиск, rate limiting, retry
        """
        params = self._build_vri_params(
            mi_number=mi_number, mit_number=mit_number,
            search_term=search_term, year=year, start=start, rows=rows
        )

        query_string = urlencode(params)
        full_url = f"{Config.API_VRI}?{query_string}"
        logger.debug(f"API запрос: GET {full_url}")

        self.rate_limiter.wait()

        for attempt in range(Config.RETRY_ATTEMPTS):
            try:
                start_time = time.time()
                response = self.session.get(
                    Config.API_VRI, params=params, timeout=Config.REQUEST_TIMEOUT
                )
                elapsed_ms = int((time.time() - start_time) * 1000)
                logger.debug(f"Ответ API: статус={response.status_code}, время={elapsed_ms}мс")

                if response.status_code == 200:
                    self.rate_limiter.on_success()
                    data = response.json()
                    items = data.get('result', {}).get('items', [])
                    total_count = data.get('result', {}).get('count', 0)
                    logger.debug(f"Получено {len(items)} записей из {total_count}")
                    return data

                elif response.status_code == 429:
                    self.rate_limiter.on_429()
                    logger.warning(f"API 429: Too Many Requests (попытка {attempt+1}/{Config.RETRY_ATTEMPTS})")
                    time.sleep(self.rate_limiter.delay * 2)
                    continue

                elif response.status_code == 408:
                    logger.error(f"API 408: Request Timeout (попытка {attempt+1})")
                    time.sleep(Config.RETRY_DELAY * (2 ** attempt))
                    continue

                elif response.status_code == 409:
                    logger.warning(f"API 409: Превышен лимит страниц (start={start})")
                    return {"result": {"items": [], "count": 0}}

                elif response.status_code == 400:
                    error_body = self._parse_error_body(response)
                    logger.error(f"API 400: Bad Request — {error_body}")
                    return {"result": {"items": [], "count": 0, "error": "bad_request"}}

                else:
                    error_body = self._parse_error_body(response)
                    logger.warning(f"API ошибка: статус={response.status_code}")
                    if response.status_code >= 500:
                        logger.error(f"5XX ошибка: {error_body}")

            except requests.Timeout:
                logger.error(f"Таймаут запроса (попытка {attempt+1}/{Config.RETRY_ATTEMPTS})")
            except requests.ConnectionError as e:
                logger.error(f"Ошибка соединения (попытка {attempt+1}): {e}")
            except Exception as e:
                logger.error(f"Неожиданная ошибка (попытка {attempt+1}): {e}")

            if attempt < Config.RETRY_ATTEMPTS - 1:
                delay = Config.RETRY_DELAY * (2 ** attempt)
                time.sleep(delay)

        return {"result": {"items": [], "count": 0, "error": "max_retries"}}

    def _parse_error_body(self, response) -> dict:
        """Парсинг тела ошибки API"""
        try:
            data = response.json()
            return {
                'status': data.get('status', ''),
                'message': data.get('message', ''),
                'time': data.get('time', ''),
                'requestId': data.get('requestId', '')
            }
        except Exception:
            return {'raw': response.text[:500]}

    def get_vri_details(self, vri_id: str) -> Dict:
        """
        Получение детальной информации о поверке
        GET /vri/{vri_id}
        """
        url = f"{Config.API_VRI}/{vri_id}"
        logger.debug(f"Детальный запрос: {url}")
        
        try:
            response = self.session.get(url, timeout=Config.REQUEST_TIMEOUT)
            
            if response.status_code == 200:
                return response.json()
            elif response.status_code == 404:
                logger.warning(f"Запись не найдена: {vri_id}")
                return {"error": "not_found"}
            else:
                logger.error(f"Ошибка детального запроса: {response.status_code}")
                return {"error": f"http_{response.status_code}"}
                
        except Exception as e:
            logger.error(f"Ошибка получения деталей {vri_id}: {e}")
            return {"error": str(e)}

    def search_mit(self, search_term: str = "*", start: int = 0, 
                   rows: int = Config.DEFAULT_ROWS) -> Dict:
        """
        Поиск в реестре типов СИ (MIT - Measurement Instruments Types)
        GET /mit
        """
        rows = min(rows, Config.MAX_ROWS_PER_REQUEST)
        encoded_search = quote(search_term, safe='*')
        
        params = {
            'search': encoded_search,
            'start': start,
            'rows': rows
        }
        
        logger.debug(f"MIT запрос: {Config.API_MIT} с параметрами {params}")
        
        try:
            response = self.session.get(
                Config.API_MIT,
                params=params,
                timeout=Config.REQUEST_TIMEOUT
            )
            
            if response.status_code == 200:
                return response.json()
            else:
                logger.warning(f"MIT API ошибка: {response.status_code}")
                return {"result": {"items": []}}
                
        except Exception as e:
            logger.error(f"Ошибка MIT запроса: {e}")
            return {"result": {"items": [], "error": str(e)}}

    def close(self):
        """Закрытие сессии"""
        self.session.close()
        logger.debug("API сессия закрыта")


# ============== АСИНХРОННЫЙ СБОРЩИК ==============
class AsyncCollector:
    """Промышленный асинхронный сборщик данных с расширенным логированием"""

    def __init__(self, db: Database = None, rate_limiter: 'DynamicRateLimiter' = None):
        self.db = db or Database()
        self.base_url = Config.API_VRI
        self.session = None
        self.total_collected = 0
        self.is_running = False
        self.stats = {'requests': 0, 'errors': 0, 'found': 0}
        self.rate_limiter = rate_limiter
        logger.info("AsyncCollector инициализирован")

    async def __aenter__(self):
        """Асинхронный вход"""
        self.session = aiohttp.ClientSession(headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json',
            'Accept-Language': 'ru-RU,ru;q=0.9',
        })
        logger.debug("aiohttp сессия создана")
        return self

    async def __aexit__(self, *args):
        """Асинхронный выход"""
        if self.session:
            await self.session.close()
            logger.debug("aiohttp сессия закрыта")

    async def fetch_page(self, year: int, page: int, search: str = "*") -> List[Dict]:
        """Асинхронный запрос одной страницы"""
        params = {
            'year': year,
            'search': quote(search, safe='*'),
            'rows': Config.MAX_ROWS_PER_REQUEST,
            'start': page * Config.MAX_ROWS_PER_REQUEST
        }
        
        self.stats['requests'] += 1
        
        try:
            async with self.session.get(self.base_url, params=params, timeout=15) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    items = data.get('result', {}).get('items', [])
                    self.stats['found'] += len(items)
                    logger.debug(f"Год {year}, стр.{page}: получено {len(items)} записей")
                    return items
                elif resp.status == 409:
                    logger.debug(f"Год {year}: превышен лимит страниц")
                    return []
                else:
                    self.stats['errors'] += 1
                    logger.debug(f"Год {year}, стр.{page}: статус {resp.status}")
                    return []
        except asyncio.TimeoutError:
            self.stats['errors'] += 1
            logger.debug(f"Год {year}, стр.{page}: таймаут")
            return []
        except Exception as e:
            self.stats['errors'] += 1
            logger.debug(f"Год {year}, стр.{page}: ошибка {e}")
            return []

    async def collect_year(self, year: int, progress_callback=None, search_term: str = "*") -> int:
        """Сбор данных за указанный год"""
        logger.info(f"Начало сбора за {year} год (поиск: '{search_term}')")
        
        existing_ids = self.db.get_existing_ids()
        new_count = 0
        page = 0
        max_pages = 50  # Ограничение API
        
        pbar = None
        if TQDM_AVAILABLE and progress_callback is None:
            pbar = tqdm(total=max_pages, desc=f"  {year}", unit="стр", leave=False)

        while page < max_pages and self.is_running:
            items = await self.fetch_page(year, page, search_term)

            if not items:
                logger.debug(f"Год {year}: данных больше нет на странице {page}")
                break

            # Фильтрация электрических счетчиков
            records = []
            for item in items:
                vri_id = item.get('vri_id')
                if vri_id and vri_id not in existing_ids and self._is_electric_meter(item):
                    record = VerificationRecord.from_api_response(
                        item, 
                        search_query=search_term,
                        search_year=year
                    )
                    records.append(record)
                    existing_ids.add(vri_id)

            if records:
                saved = self.db.save_records(records)
                new_count += saved
                logger.info(f"Год {year}, стр.{page}: сохранено {saved} новых записей")

                if progress_callback:
                    progress_callback(year, page, saved)

            if pbar:
                pbar.update(1)
                pbar.set_postfix({"найдено": new_count})

            # Пауза между запросами для соблюдения лимитов API
            await asyncio.sleep(0.3)

            # Если вернулось меньше максимального количества - последняя страница
            if len(items) < Config.MAX_ROWS_PER_REQUEST:
                break

            page += 1

        if pbar:
            pbar.close()

        logger.info(f"Год {year} завершен: собрано {new_count} новых записей")
        return new_count

    def _is_electric_meter(self, item: Dict) -> bool:
        """
        Проверка, является ли прибор электрическим счетчиком
        Использует точные запросы из конфигурации
        """
        mit_title = item.get('mit_title', '').lower()
        mit_notation = item.get('mit_notation', '').lower()
        mi_modification = item.get('mi_modification', '').lower()
        
        combined = f"{mit_title} {mit_notation} {mi_modification}"

        # Быстрое исключение неэлектрических приборов
        exclude_words = ['воды', 'газа', 'весы', 'термометр', 'манометр', 
                         'датчик', 'линейка', 'гиря', 'тепло', 'нефти',
                         'давления', 'температур', 'расходомер']
        
        for word in exclude_words:
            if word in combined:
                return False

        # Должно быть слово "счетчик"
        if 'счетчик' not in mit_title and 'счетчик' not in mit_notation:
            return False

        # Признаки электрического счетчика
        electric_indicators = [
            'электрический', 'электроэнергии', 'электросчетчик',
            'активной энергии', 'реактивной энергии', 'ватт', 'киловатт',
            'меркурий', 'нева', 'энергомера', 'инкотекс', 'матрица',
            'псч', 'сэт', 'це', 'квт', 'kwh', 'kw'
        ]

        for word in electric_indicators:
            if word in combined:
                return True

        return False

    async def collect_all_years(self, years: List[int], callback=None, 
                                 use_exact_queries: bool = True) -> Dict:
        """
        Сбор данных за все указанные годы с использованием гибридного метода
        """
        self.is_running = True
        results = {}
        
        logger.info(f"Запуск сбора за годы: {min(years)}-{max(years)}")
        
        if use_exact_queries and EXACT_QUERIES:
            logger.info(f"Используется гибридный режим с {len(EXACT_QUERIES)} точными запросами")
            
            # Этап 1: Точные запросы (быстрый сбор)
            for query in EXACT_QUERIES:
                if not self.is_running:
                    break
                
                logger.info(f"Точный запрос: '{query}'")
                for year in years:
                    count = await self.collect_year(year, callback, query)
                    results[f"{year}_{query[:30]}"] = count
        else:
            # Этап 2: Общий обход
            for year in years:
                if not self.is_running:
                    break
                count = await self.collect_year(year, callback)
                results[year] = count
                
                # Промежуточная статистика
                stats = self.db.get_stats()
                logger.info(f"Промежуточный итог: всего {stats['total']} записей")

        self.is_running = False
        logger.info(f"Сбор завершен. Статистика: {self.stats}")
        return results

    def stop(self):
        """Остановка сбора"""
        self.is_running = False
        logger.info("Получена команда остановки сбора")


# ============== ЭКСПОРТЕР ==============
class Exporter:
    """Экспорт данных с полной служебной информацией"""

    @staticmethod
    def ensure_export_dir():
        """Создание директории для экспорта"""
        os.makedirs(Config.EXPORT_DIR, exist_ok=True)

    @staticmethod
    def to_csv(data: List[Dict], filename: str = None,
               include_service_fields: bool = True) -> str:
        """
        Экспорт в CSV с русскими заголовками

        Args:
            data: Список записей
            filename: Имя файла
            include_service_fields: Включать служебные поля
        """
        Exporter.ensure_export_dir()

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{Config.EXPORT_DIR}/export_{timestamp}.csv"

        if not data:
            logger.warning("Нет данных для экспорта")
            with open(filename, 'w', encoding='utf-8-sig') as f:
                f.write("Нет данных")
            return filename

        # Русские заголовки для всех полей
        field_mapping = {
            # Основные данные из API
            'vri_id': 'ID',
            'mit_number': '№ реестра',
            'mit_title': 'Наименование',
            'mit_notation': 'Обозначение типа',
            'mi_modification': 'Модификация',
            'mi_number': 'Зав. номер',
            'verification_date': 'Дата поверки',
            'valid_date': 'Действ. до',
            'applicability': 'Пригоден',
            'org_title': 'Поверитель',
            'result_docnum': '№ документа',
            'manufacturer': 'Производитель',
            
            # Служебные поля
            'search_query': 'Поисковый запрос',
            'row_index': '№ строки',
            'id_pu': 'Id_ПУ',
            
            # Данные из исходного Excel файла
            'contract_number': 'Номер договора',
            'edo_code': 'Код ЭДО',
            'balance_owner': 'Балансовая принадлежность',
            'operation_responsibility': 'Эксплуатационная ответственность',
            'mpi': 'МПИ',
            
            # Техническое
            'collected_at': 'Время сохранения'
        }

        # Порядок полей для экспорта
        base_fields = [
            'vri_id', 'mit_number', 'mit_title', 'mit_notation',
            'mi_modification', 'mi_number', 'verification_date',
            'valid_date', 'applicability', 'org_title', 'result_docnum',
            'manufacturer'
        ]

        if include_service_fields:
            base_fields.extend([
                'id_pu', 'contract_number', 'edo_code',
                'balance_owner', 'operation_responsibility', 'mpi',
                'search_query', 'row_index'
            ])

        base_fields.append('collected_at')

        # Фильтруем только существующие поля и создаём mapping
        export_fields = [f for f in base_fields if f in field_mapping]
        russian_headers = [field_mapping[f] for f in export_fields]

        with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=export_fields, delimiter=';', extrasaction='ignore')
            # Пишем русские заголовки
            writer.writerow(dict(zip(export_fields, russian_headers)))
            
            for row in data:
                writer.writerow(row)

        logger.info(f"Экспортировано {len(data)} записей в {filename}")
        logger.debug(f"Поля экспорта: {export_fields}")
        return filename

    @staticmethod
    def to_excel(data: List[Dict], filename: str = None,
                 include_service_fields: bool = True) -> str:
        """Экспорт в Excel с русскими заголовками и форматированием"""
        if not PANDAS_AVAILABLE:
            logger.warning("pandas не установлен, используется CSV")
            return Exporter.to_csv(data, filename, include_service_fields)

        Exporter.ensure_export_dir()

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{Config.EXPORT_DIR}/export_{timestamp}.xlsx"

        # Русские заголовки для всех полей
        field_mapping = {
            # Основные данные из API
            'vri_id': 'ID',
            'mit_number': '№ реестра',
            'mit_title': 'Наименование',
            'mit_notation': 'Обозначение типа',
            'mi_modification': 'Модификация',
            'mi_number': 'Зав. номер',
            'verification_date': 'Дата поверки',
            'valid_date': 'Действ. до',
            'applicability': 'Пригоден',
            'org_title': 'Поверитель',
            'result_docnum': '№ документа',
            'manufacturer': 'Производитель',
            
            # Служебные поля
            'search_query': 'Поисковый запрос',
            'row_index': '№ строки',
            'id_pu': 'Id_ПУ',
            
            # Данные из исходного Excel файла
            'contract_number': 'Номер договора',
            'edo_code': 'Код ЭДО',
            'balance_owner': 'Балансовая принадлежность',
            'operation_responsibility': 'Эксплуатационная ответственность',
            'mpi': 'МПИ',
            
            # Техническое
            'collected_at': 'Время сохранения'
        }

        # Порядок полей для экспорта
        base_fields = [
            'vri_id', 'mit_number', 'mit_title', 'mit_notation',
            'mi_modification', 'mi_number', 'verification_date',
            'valid_date', 'applicability', 'org_title', 'result_docnum',
            'manufacturer'
        ]

        if include_service_fields:
            base_fields.extend([
                'id_pu', 'contract_number', 'edo_code',
                'balance_owner', 'operation_responsibility', 'mpi',
                'search_query', 'row_index'
            ])

        base_fields.append('collected_at')

        # Фильтруем только существующие поля
        if data:
            export_fields = [f for f in base_fields if f in data[0].keys()]
        else:
            export_fields = base_fields

        # Создаём DataFrame с русскими заголовками
        filtered_data = []
        for row in data:
            filtered_row = {field_mapping.get(k, k): v for k, v in row.items() if k in export_fields}
            filtered_data.append(filtered_row)

        df = pd.DataFrame(filtered_data)

        # Создание Excel с форматированием
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Результаты')

            workbook = writer.book
            worksheet = writer.sheets['Результаты']

            # Форматирование заголовков
            header_font = ExcelFont(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')

            for col_num, value in enumerate(df.columns.values):
                cell = worksheet.cell(row=1, column=col_num+1)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Автоподбор ширины колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        logger.info(f"Экспортировано {len(data)} записей в {filename}")
        return filename

    @staticmethod
    def export_stats(stats: Dict) -> str:
        """Экспорт статистики в JSON"""
        Exporter.ensure_export_dir()

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{Config.EXPORT_DIR}/stats_{timestamp}.json"

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(stats, f, ensure_ascii=False, indent=2)

        logger.info(f"Статистика экспортирована в {filename}")
        return filename
    
    @staticmethod
    def export_with_links(data: List[Dict], filename: str = None) -> str:
        """
        Специальный экспорт с сохранением ссылок на записи
        Включает все поля + record_url
        """
        if not data:
            logger.warning("Нет данных для экспорта")
            return ""
        
        # Гарантируем наличие URL
        for row in data:
            if 'record_url' not in row or not row['record_url']:
                vri_id = row.get('vri_id', '')
                if vri_id:
                    row['record_url'] = f"https://fgis.gost.ru/fundmetrology/cm/erts/?id={vri_id}"
        
        return Exporter.to_csv(data, filename, include_service_fields=True)


# ============== ОБРАБОТЧИК EXCEL ==============
class ExcelHandler:
    """Работа с Excel файлами для пакетного поиска с сохранением связей"""

    @staticmethod
    def detect_columns(df: pd.DataFrame) -> Dict[str, Tuple[str, int]]:
        """
        Автоматическое определение колонок по заголовкам
        Структура таблицы определяется автоматически - важен только набор нужных колонок,
        а не их порядок или положение в файле

        Returns:
            Dict[field_name: Tuple[column_name, column_index]]
        """
        column_mapping = {}

        # Расширенные возможные названия колонок для различных форматов файлов
        # Ключевые слова упорядочены по приоритету (более специфичные первыми)
        possible_columns = {
            # Серийный/заводской номер прибора - ищем точные совпадения
            'mi_number': [
                'серийный номер', 'заводской номер', 'номер пу', 'серийный номер пу',
                'зав. номер', 'зав №', 'serial number', 'device_number',
                'серийный', 'заводской', 'serial', 'pu_number', 'id_пу',
                'номер пу', 'прибор', 'счетчик', 'измеритель', 'id'
            ],
            # Номер в реестре типов СИ
            'mit_number': [
                'номер в реестре', 'реестровый номер', 'рег. номер', 'регистрационный',
                'тип си номер', 'номер типа', 'mit_number', 'vri_id',
                'реестр', 'тип си', 'тип'
            ],
            # Наименование типа СИ - приоритет коротким названиям
            'mit_title': [
                'модель', 'model', 'наименование типа', 'тип прибора', 'тип пу',
                'наименование', 'название', 'name', 'тип си', 'устройство',
                'описание', 'прибор'
            ],
            # Организация-поверитель - только специфичные слова
            'org_title': [
                'организация поверитель', 'поверитель', 'организация', 'org_title',
                'org', 'кто поверял', 'поверка', 'company', 'предприятие',
                'эксплуатационная ответственность'
            ],
            # Идентификатор записи VRI
            'vri_id': [
                'vri_id', 'vri', 'идентификатор', 'уникальный', 'uuid',
                'ключ', 'key', 'id записи'
            ],
            # Поисковый запрос (универсальное поле)
            'search_term': [
                'поисковый запрос', 'ключевое слово', 'поиск', 'query', 'запрос',
                'search', 'термин', 'ключ'
            ],
            # Год поверки
            'year': [
                'год поверки', 'поверка год', 'год поверки си', 'verification year',
                'год', 'year', 'дата поверки', 'mpi', 'мпИ'
            ],
            # Год выпуска прибора
            'manufacture_year': [
                'год выпуска', 'год производства', 'год изготовления', 'дата выпуска',
                'выпуск', 'manufacture', 'production year'
            ]
        }

        # Преобразуем названия колонок в нижний регистр для сравнения
        df_lower = [col.lower().strip() for col in df.columns]

        logger.debug(f"Определение колонок: {list(df.columns)}")

        # Сбор всех кандидатов (поле, колонка, оценка)
        candidates = []
        for field, keywords in possible_columns.items():
            for idx, col_lower in enumerate(df_lower):
                score = 0
                original_col = df.columns[idx]

                if col_lower == field:
                    score = 100
                else:
                    for priority, keyword in enumerate(keywords):
                        if keyword == col_lower:
                            score = max(score, 95 - priority)
                        elif col_lower.startswith(keyword):
                            score = max(score, 80 - priority)
                        elif keyword in col_lower:
                            score = max(score, 60 - priority)
                        elif len(keyword) > 4 and keyword[:4] in col_lower:
                            score = max(score, 40 - priority)

                if score > 50:
                    candidates.append((score, field, original_col, idx))

        # Сортировка по убыванию оценки — лучшие совпадения первыми
        candidates.sort(key=lambda x: x[0], reverse=True)

        # Назначение: каждая колонка и каждое поле — только один раз
        used_fields = set()
        used_cols = set()
        for score, field, col_name, idx in candidates:
            if field in used_fields or col_name in used_cols:
                continue
            column_mapping[field] = (col_name, idx)
            used_fields.add(field)
            used_cols.add(col_name)
            logger.debug(f"✅ Найдено: {field} -> {col_name} (score={score})")

        # Если не найдено поле search_term, но есть mi_number - используем его как поисковое
        if 'search_term' not in column_mapping and 'mi_number' in column_mapping:
            logger.debug("search_term не найден, будет использоваться mi_number для поиска")

        logger.info(f"Автоматически определено колонок: {len(column_mapping)} из {len(possible_columns)}")
        return column_mapping

    @staticmethod
    def read_queries(filename: str) -> List[Dict]:
        """
        Чтение запросов из Excel с сохранением исходных связей
        
        Returns:
            Список словарей с полями + row_index для связи
        """
        if not PANDAS_AVAILABLE:
            logger.error("pandas не установлен")
            raise ImportError("pandas не установлен")

        logger.info(f"Чтение Excel файла: {filename}")
        df = pd.read_excel(filename, dtype=str)
        logger.debug(f"Прочитано {len(df)} строк, колонки: {list(df.columns)}")

        # Определение колонок
        column_mapping = ExcelHandler.detect_columns(df)

        # Если не удалось определить, используем первую колонку как search_term
        if not column_mapping:
            logger.warning("Колонки не определены, используется первая колонка")
            column_mapping = {'search_term': (df.columns[0], 0)}

        queries = []
        for idx, row in df.iterrows():
            query = {
                'row_index': idx + 1,  # Индекс строки (1-based)
                'original_data': {}    # Исходные данные строки
            }
            
            # Сохранение всех исходных данных
            for col in df.columns:
                value = row[col]
                if value and str(value).lower() != 'nan':
                    query['original_data'][col] = str(value)
            
            # Маппинг полей
            for field, (col_name, _) in column_mapping.items():
                value = row[col_name]
                if value and str(value).lower() != 'nan':
                    query[field] = str(value)
            
            if 'search_term' in query or 'mi_number' in query:
                queries.append(query)

        logger.info(f"Загружено {len(queries)} запросов из Excel")
        return queries

    @staticmethod
    def create_template(filename: str = None) -> str:
        """Создание шаблона Excel файла"""
        if not PANDAS_AVAILABLE:
            return ""
        
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{Config.EXPORT_DIR}/batch_search_template_{timestamp}.xlsx"
        
        Exporter.ensure_export_dir()
        
        # Создание шаблона с примерами
        data = {
            'Заводской номер': ['12345678', '87654321', ''],
            'Наименование': ['Счетчик электрической энергии', '', ''],
            'Номер в реестре': ['77310-20', '', ''],
            'Год поверки': [2023, 2022, ''],
            'Примечание': ['Первый запрос', 'Второй запрос', '']
        }
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Шаблон')
            
            # Добавление комментариев
            worksheet = writer.sheets['Шаблон']
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 12
            worksheet.column_dimensions['E'].width = 25
        
        logger.info(f"Шаблон создан: {filename}")
        return filename


# ============== ГРАФИЧЕСКИЙ ИНТЕРФЕЙС (оптимизирован для 1200x700) ==============
class ArshinApp:
    """Главное окно приложения с оптимизированным интерфейсом"""

    def __init__(self, root):
        self.root = root
        self.root.title("ФГИС Аршин - Поиск и сбор данных v3.0")
        self.root.geometry(f"{Config.WINDOW_WIDTH}x{Config.WINDOW_HEIGHT}")
        
        # Минимальные размеры для корректного отображения
        self.root.minsize(1024, 600)

        # Инициализация компонентов
        self.rate_limiter = DynamicRateLimiter()
        self.db = Database()
        self.api_client = APIClient(rate_limiter=self.rate_limiter)
        self.collector = AsyncCollector(self.db, rate_limiter=self.rate_limiter)

        # Переменные для пакетного поиска
        self.batch_queries = []
        self.batch_results = []
        self.batch_all_columns = []  # Все колонки из Excel файла
        self.current_batch_index = 0
        self.batch_running = False
        self.batch_stop_requested = False

        # Переменные для одиночного поиска
        self.single_search_results = []  # Результаты последнего поиска

        # Переменные для асинхронного сбора
        self.async_running = False
        self.async_thread = None

        # Логирование
        logger.info(f"Приложение запущено, окно: {Config.WINDOW_WIDTH}x{Config.WINDOW_HEIGHT}")
        logger.info(f"Загружено {len(EXACT_QUERIES)} точных запросов, {len(MANUFACTURERS_RULES)} производителей")

        # Создание интерфейса
        self.setup_ui()

        # Загрузка статистики
        self.update_stats()

        # Обработка закрытия окна
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_ui(self):
        """Создание компактного интерфейса"""
        # Главное меню (компактное)
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # Меню "Файл"
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Экспорт в CSV", command=self.export_csv, accelerator="Ctrl+S")
        file_menu.add_command(label="Экспорт в Excel", command=self.export_excel)
        file_menu.add_command(label="Экспорт статистики", command=self.export_stats_json)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.on_close, accelerator="Ctrl+Q")

        # Меню "Инструменты"
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Инструменты", menu=tools_menu)
        tools_menu.add_command(label="Обновить статистику", command=self.update_stats)
        tools_menu.add_command(label="Очистить дубликаты", command=self.clean_duplicates)
        tools_menu.add_command(label="История поиска", command=self.show_search_history)
        tools_menu.add_separator()
        tools_menu.add_command(label="Создать шаблон Excel", command=lambda: ExcelHandler.create_template())

        # Меню "Помощь"
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Помощь", menu=help_menu)
        help_menu.add_command(label="О программе", command=self.show_about)
        help_menu.add_command(label="Лог файл", command=self.show_log_file)

        # Основной контейнер с вкладками (компактный)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)

        # Вкладки (оптимизированные)
        self.setup_single_search_tab()          # Поиск
        self.setup_batch_operations_tab()       # Пакетная операция (только файл)
        self.setup_results_tab()                # Результаты (с экспортом и статистикой)
        self.setup_filters_tab()                # Фильтры (config файлы)
        self.setup_stats_tab()                  # Статистика

        # Привязка горячих клавиш
        self.root.bind('<Control-s>', lambda e: self.export_csv())
        self.root.bind('<Control-q>', lambda e: self.on_close())

    def setup_single_search_tab(self):
        """Вкладка одиночного поиска - с управлением результатами"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔍 Поиск")

        # Верхняя панель с параметрами
        params_frame = ttk.LabelFrame(tab, text="Параметры поиска", padding=3)
        params_frame.pack(fill=tk.X, padx=5, pady=2)

        # Row 0: Заводской номер и кнопки
        search_frame = ttk.Frame(params_frame)
        search_frame.pack(fill=tk.X, pady=1)

        ttk.Label(search_frame, text="Зав. номер:", width=10).pack(side=tk.LEFT)
        self.search_entry = ttk.Entry(search_frame, width=30)
        self.search_entry.pack(side=tk.LEFT, padx=3)
        
        # Разрешить вставку мышкой
        self.search_entry.bind('<Button-3>', lambda e: None)  # Разрешить контекстное меню

        # Кнопки управления
        self.search_start_btn = ttk.Button(search_frame, text="▶ Найти", command=self.single_search, width=12, style='Green.TButton')
        self.search_start_btn.pack(side=tk.LEFT, padx=3)
        
        self.search_stop_btn = ttk.Button(search_frame, text="⏹ Стоп", command=self.stop_single_search, width=10, state=tk.DISABLED, style='Red.TButton')
        self.search_stop_btn.pack(side=tk.LEFT, padx=3)
        
        clear_btn = ttk.Button(search_frame, text="✖ Очистить", command=self.clear_single_search, width=12)
        clear_btn.pack(side=tk.LEFT, padx=3)

        # Row 1: Информация о поиске
        info_frame = ttk.Frame(params_frame)
        info_frame.pack(fill=tk.X, pady=1)
        
        self.search_info_var = tk.StringVar(value="Готов к поиску")
        ttk.Label(info_frame, textvariable=self.search_info_var, foreground='blue').pack(side=tk.LEFT, padx=5)

        # Row 2: Годы (заголовок)
        years_label_frame = ttk.Frame(params_frame)
        years_label_frame.pack(fill=tk.X, pady=1)
        ttk.Label(years_label_frame, text="Годы:").pack(side=tk.LEFT, padx=3)

        # Row 3: Годы (чекбоксы в одну строку)
        years_frame = ttk.Frame(params_frame)
        years_frame.pack(fill=tk.X, pady=1)

        # Тип поиска (только точное по умолчанию - в коде)
        self.search_type = tk.StringVar(value="exact")
        self.search_running = False
        self.search_stop_requested = False

        self.search_year_vars = {}
        for year in range(2010, 2027):
            var = tk.BooleanVar(value=True)
            self.search_year_vars[year] = var
            ttk.Checkbutton(years_frame, text=str(year), variable=var, width=5).pack(side=tk.LEFT, padx=0)

        # Панель действий с результатами
        actions_frame = ttk.Frame(tab)
        actions_frame.pack(fill=tk.X, padx=5, pady=2)

        ttk.Label(actions_frame, text="Действия:").pack(side=tk.LEFT, padx=3)
        
        save_csv_btn = ttk.Button(actions_frame, text="💾 Сохранить CSV", command=self.save_single_search_csv_from_tree, width=15)
        save_csv_btn.pack(side=tk.LEFT, padx=2)
        
        add_to_db_btn = ttk.Button(actions_frame, text="➕ В БД", command=self.add_single_search_to_db_from_tree, width=12)
        add_to_db_btn.pack(side=tk.LEFT, padx=2)
        
        delete_btn = ttk.Button(actions_frame, text="🗑 Удалить выбранные", command=self.delete_selected_rows, width=18)
        delete_btn.pack(side=tk.LEFT, padx=2)

        # Результаты поиска - полная таблица (10 колонок)
        results_frame = ttk.LabelFrame(tab, text="Результаты поиска", padding=0)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        # Контейнер для скроллов
        tree_container = ttk.Frame(results_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)

        # Все поля из ответа API
        columns = (
            'vri_id', 'mit_number', 'mit_title', 'mit_notation', 'mi_modification',
            'mi_number', 'verification_date', 'valid_date', 'org_title', 'result_docnum', 'applicability'
        )
        self.single_results_tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=12, selectmode='extended')

        # Заголовки на русском языке
        self.single_results_tree.heading('vri_id', text='ID записи')
        self.single_results_tree.heading('mit_number', text='№ в реестре')
        self.single_results_tree.heading('mit_title', text='Наименование СИ')
        self.single_results_tree.heading('mit_notation', text='Обозначение типа')
        self.single_results_tree.heading('mi_modification', text='Модификация')
        self.single_results_tree.heading('mi_number', text='Зав. номер')
        self.single_results_tree.heading('verification_date', text='Дата поверки')
        self.single_results_tree.heading('valid_date', text='Действ. до')
        self.single_results_tree.heading('org_title', text='Поверитель')
        self.single_results_tree.heading('result_docnum', text='№ документа')
        self.single_results_tree.heading('applicability', text='Пригоден')

        # Ширина колонок
        self.single_results_tree.column('vri_id', width=90)
        self.single_results_tree.column('mit_number', width=80)
        self.single_results_tree.column('mit_title', width=250)
        self.single_results_tree.column('mit_notation', width=100)
        self.single_results_tree.column('mi_modification', width=150)
        self.single_results_tree.column('mi_number', width=100)
        self.single_results_tree.column('verification_date', width=90)
        self.single_results_tree.column('valid_date', width=90)
        self.single_results_tree.column('org_title', width=180)
        self.single_results_tree.column('result_docnum', width=130)
        self.single_results_tree.column('applicability', width=70)

        # Скроллбары
        v_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.single_results_tree.yview)
        h_scroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.single_results_tree.xview)
        self.single_results_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.single_results_tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        self.setup_tree_context_menu(self.single_results_tree)
        
        # Статус бар вкладки
        self.single_status_var = tk.StringVar(value="Готов")
        status_bar = ttk.Label(tab, textvariable=self.single_status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, padx=3, pady=1)

    def setup_batch_operations_tab(self):
        """Вкладка пакетных операций - оптимизированная"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📦 Пакетная")

        # Файл (компактно)
        file_frame = ttk.Frame(tab)
        file_frame.pack(fill=tk.X, padx=5, pady=2)

        ttk.Label(file_frame, text="Excel:").pack(side=tk.LEFT)
        self.batch_file_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.batch_file_var, width=45).pack(side=tk.LEFT, padx=3, fill=tk.X, expand=True)
        ttk.Button(file_frame, text="📂", command=self.select_batch_file, width=3).pack(side=tk.LEFT, padx=2)
        ttk.Button(file_frame, text="📄", command=lambda: ExcelHandler.create_template(), width=3).pack(side=tk.LEFT, padx=2)

        # Параметры
        params_frame = ttk.LabelFrame(tab, text="Параметры", padding=2)
        params_frame.pack(fill=tk.X, padx=5, pady=2)

        # Верхняя строка: Годы (слева) + Кнопки (справа сверху)
        top_row = ttk.Frame(params_frame)
        top_row.pack(fill=tk.X, pady=1)

        # ===== ЛЕВАЯ ЧАСТЬ: Годы =====
        years_group = ttk.LabelFrame(top_row, text="Годы", padding=2)
        years_group.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.year_vars = {}
        
        # Строка 1: 2010-2016
        years_row1 = ttk.Frame(years_group)
        years_row1.pack(fill=tk.X, pady=0)
        for year in range(2010, 2017):
            var = tk.BooleanVar(value=True)
            self.year_vars[year] = var
            ttk.Checkbutton(years_row1, text=str(year), variable=var, width=5).pack(side=tk.LEFT, padx=0)

        # Строка 2: 2020-2026 + кнопка
        years_row2 = ttk.Frame(years_group)
        years_row2.pack(fill=tk.X, pady=0)
        for year in range(2020, 2027):
            var = tk.BooleanVar(value=True)
            self.year_vars[year] = var
            ttk.Checkbutton(years_row2, text=str(year), variable=var, width=5).pack(side=tk.LEFT, padx=0)
        
        ttk.Button(years_row2, text="Все...", command=self.select_all_years_dialog, width=6).pack(side=tk.LEFT, padx=2)

        # ===== ПРАВАЯ ЧАСТЬ: Кнопки =====
        buttons_frame = ttk.Frame(top_row)
        buttons_frame.pack(side=tk.RIGHT, padx=5)

        # Кнопка Запуск (светло-зелёная)
        self.batch_start_btn = ttk.Button(buttons_frame, text="▶ Запуск", command=self.start_batch_search, width=12)
        self.batch_start_btn.pack(side=tk.LEFT, padx=2)
        self.batch_start_btn.configure(style='Green.TButton')

        # Кнопка Стоп (светло-красная)
        self.batch_stop_btn = ttk.Button(buttons_frame, text="⏹ Стоп", command=self.stop_batch_search, width=10, state=tk.DISABLED)
        self.batch_stop_btn.pack(side=tk.LEFT, padx=2)
        self.batch_stop_btn.configure(style='Red.TButton')

        # Прогресс и статистика (в одну строку)
        progress_frame = ttk.Frame(params_frame)
        progress_frame.pack(fill=tk.X, pady=1)

        # Прогресс бар на всю ширину
        self.batch_progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.batch_progress.pack(fill=tk.X, pady=1)

        # Статистика в одну строку
        stats_row = ttk.Frame(progress_frame)
        stats_row.pack(fill=tk.X, pady=0)

        self.batch_processed_var = tk.StringVar(value="Обработано: 0 / 0")
        ttk.Label(stats_row, textvariable=self.batch_processed_var, width=25).pack(side=tk.LEFT, padx=5)

        self.batch_found_var = tk.StringVar(value="Найдено: 0")
        ttk.Label(stats_row, textvariable=self.batch_found_var, width=15).pack(side=tk.LEFT, padx=5)

        self.batch_filtered_var = tk.StringVar(value="Отфильтровано: 0")
        ttk.Label(stats_row, textvariable=self.batch_filtered_var, width=20).pack(side=tk.LEFT, padx=5)

        # Таблица запросов для предпросмотра - динамические колонки
        preview_frame = ttk.LabelFrame(tab, text="Запросы из файла (предпросмотр)", padding=0)
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

        # Контейнер для скроллов
        tree_container = ttk.Frame(preview_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)

        # Базовые колонки (будут расширены при загрузке файла)
        columns = ('row_num', 'id_pu', 'mi_number', 'mit_title', 'mit_number', 'year',
                   'contract_number', 'edo_code', 'balance_owner', 'operation_responsibility', 'mpi')
        self.batch_preview_tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=14)

        # Заголовки
        self.batch_preview_tree.heading('row_num', text='№')
        self.batch_preview_tree.heading('id_pu', text='Id_ПУ')
        self.batch_preview_tree.heading('mi_number', text='Серийный номер')
        self.batch_preview_tree.heading('mit_title', text='Модель')
        self.batch_preview_tree.heading('mit_number', text='Тип ПУ')
        self.batch_preview_tree.heading('year', text='Год')
        self.batch_preview_tree.heading('contract_number', text='Номер договора')
        self.batch_preview_tree.heading('edo_code', text='Код ЭДО')
        self.batch_preview_tree.heading('balance_owner', text='Балансовая принадлежность')
        self.batch_preview_tree.heading('operation_responsibility', text='Эксплуатационная ответственность')
        self.batch_preview_tree.heading('mpi', text='МПИ')

        # Ширина колонок
        self.batch_preview_tree.column('row_num', width=35)
        self.batch_preview_tree.column('id_pu', width=90)
        self.batch_preview_tree.column('mi_number', width=110)
        self.batch_preview_tree.column('mit_title', width=180)
        self.batch_preview_tree.column('mit_number', width=90)
        self.batch_preview_tree.column('year', width=50)
        self.batch_preview_tree.column('contract_number', width=100)
        self.batch_preview_tree.column('edo_code', width=80)
        self.batch_preview_tree.column('balance_owner', width=150)
        self.batch_preview_tree.column('operation_responsibility', width=180)
        self.batch_preview_tree.column('mpi', width=50)

        # Скроллы
        v_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.batch_preview_tree.yview)
        h_scroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.batch_preview_tree.xview)
        self.batch_preview_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.batch_preview_tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

    def setup_filters_tab(self):
        """Вкладка управления фильтрами и конфигурацией"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="🔧 Фильтры")

        # Точные запросы
        exact_frame = ttk.LabelFrame(tab, text="Точные запросы", padding=10)
        exact_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Список запросов
        self.exact_queries_list = scrolledtext.ScrolledText(exact_frame, height=10, font=('Consolas', 9))
        self.exact_queries_list.pack(fill=tk.BOTH, expand=True, pady=5)

        # Кнопки управления
        exact_btn_frame = ttk.Frame(exact_frame)
        exact_btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(exact_btn_frame, text="📂 Загрузить", command=self.load_exact_queries_file, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(exact_btn_frame, text="💾 Сохранить в файл", command=self.save_exact_queries_file, width=15).pack(side=tk.LEFT, padx=2)
        ttk.Button(exact_btn_frame, text="🗑 Очистить", command=self.clear_exact_queries, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(exact_btn_frame, text="➕ Добавить", command=self.add_exact_query, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(exact_btn_frame, text="🔄 Из файла", command=self.load_exact_queries_display, width=12).pack(side=tk.LEFT, padx=2)

        # Производители
        mfg_frame = ttk.LabelFrame(tab, text="Производители", padding=10)
        mfg_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Список производителей
        self.manufacturers_list = scrolledtext.ScrolledText(mfg_frame, height=10, font=('Consolas', 9))
        self.manufacturers_list.pack(fill=tk.BOTH, expand=True, pady=5)

        # Кнопки управления
        mfg_btn_frame = ttk.Frame(mfg_frame)
        mfg_btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(mfg_btn_frame, text="📂 Загрузить", command=self.load_manufacturers_file, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(mfg_btn_frame, text="💾 Сохранить в файл", command=self.save_manufacturers_file, width=15).pack(side=tk.LEFT, padx=2)
        ttk.Button(mfg_btn_frame, text="🗑 Очистить", command=self.clear_manufacturers, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(mfg_btn_frame, text="➕ Добавить", command=self.add_manufacturer, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(mfg_btn_frame, text="🔄 Из файла", command=self.load_manufacturers_display, width=12).pack(side=tk.LEFT, padx=2)

        # Статус
        self.filters_status_var = tk.StringVar(value="Готов")
        status_bar = ttk.Label(tab, textvariable=self.filters_status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, padx=3, pady=1)

        # Загрузка данных при старте
        self.load_exact_queries_display()
        self.load_manufacturers_display()

    def setup_stats_tab(self):
        """Вкладка статистики"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📊 Статистика")

        # Статистика
        stats_frame = ttk.LabelFrame(tab, text="База данных", padding=5)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        self.stats_text = scrolledtext.ScrolledText(stats_frame, height=20, font=('Consolas', 10))
        self.stats_text.pack(fill=tk.BOTH, expand=True)

        # Кнопки
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(pady=3)

        ttk.Button(btn_frame, text="Обновить", command=self.update_stats, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="Экспорт JSON", command=self.export_stats_json, width=12).pack(side=tk.LEFT, padx=3)

    def setup_results_tab(self):
        """Вкладка просмотра результатов - оптимизированная для 1200x700"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📋 Результаты")

        # Фильтры (компактно в 2 строки)
        filter_frame = ttk.Frame(tab)
        filter_frame.pack(fill=tk.X, padx=5, pady=2)

        # Row 0 - Поиск и кнопки
        row0 = ttk.Frame(filter_frame)
        row0.pack(fill=tk.X, pady=1)

        ttk.Label(row0, text="Поиск:").pack(side=tk.LEFT)
        self.db_search_var = tk.StringVar()
        ttk.Entry(row0, textvariable=self.db_search_var, width=25).pack(side=tk.LEFT, padx=2)
        ttk.Button(row0, text="🔍", command=self.load_db_data, width=3).pack(side=tk.LEFT)
        ttk.Button(row0, text="✖", command=self.reset_db_filters, width=3).pack(side=tk.LEFT, padx=5)

        # Кнопки справа
        btn_row = ttk.Frame(row0)
        btn_row.pack(side=tk.RIGHT)
        ttk.Button(btn_row, text="📄 CSV", command=self.export_results_csv, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="📊 Excel", command=self.export_results_excel, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="🗑 БД", command=self.clear_database, width=7).pack(side=tk.LEFT, padx=2)

        # Row 1 - Год и Производитель
        row1 = ttk.Frame(filter_frame)
        row1.pack(fill=tk.X, pady=1)

        ttk.Label(row1, text="Год:").pack(side=tk.LEFT)
        self.db_year_var = tk.StringVar(value="")
        year_combo = ttk.Combobox(row1, textvariable=self.db_year_var, width=5)
        year_combo['values'] = [''] + Config.YEARS
        year_combo.pack(side=tk.LEFT, padx=2)

        ttk.Label(row1, text="| Производитель:").pack(side=tk.LEFT, padx=(10, 0))
        self.db_manufacturer_var = tk.StringVar(value="")
        self.db_manufacturer_combo = ttk.Combobox(row1, textvariable=self.db_manufacturer_var, width=12)
        self.db_manufacturer_combo.pack(side=tk.LEFT, padx=2)

        # Таблица с результатами - расширенные поля
        table_frame = ttk.LabelFrame(tab, text="Результаты", padding=0)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

        # Контейнер для скроллов
        tree_container = ttk.Frame(table_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)

        # Столбцы с служебной информацией (20 колонок)
        columns = (
            'vri_id', 'id_pu', 'mit_number', 'mit_title', 'mit_notation', 'mi_modification',
            'mi_number', 'verification_date', 'valid_date', 'applicability',
            'org_title', 'result_docnum', 'manufacturer',
            'contract_number', 'edo_code', 'balance_owner', 'operation_responsibility', 'mpi',
            'search_query', 'row_index'
        )
        self.db_tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=12)

        # Заголовки
        self.db_tree.heading('vri_id', text='ID')
        self.db_tree.heading('id_pu', text='Id_ПУ')
        self.db_tree.heading('mit_number', text='№ реестра')
        self.db_tree.heading('mit_title', text='Наименование')
        self.db_tree.heading('mit_notation', text='Обозначение')
        self.db_tree.heading('mi_modification', text='Модиф.')
        self.db_tree.heading('mi_number', text='Зав. №')
        self.db_tree.heading('verification_date', text='Дата поверки')
        self.db_tree.heading('valid_date', text='Действ. до')
        self.db_tree.heading('applicability', text='Пригоден')
        self.db_tree.heading('org_title', text='Поверитель')
        self.db_tree.heading('result_docnum', text='№ док.')
        self.db_tree.heading('manufacturer', text='Производитель')
        self.db_tree.heading('contract_number', text='Номер договора')
        self.db_tree.heading('edo_code', text='Код ЭДО')
        self.db_tree.heading('balance_owner', text='Балансовая принадлежность')
        self.db_tree.heading('operation_responsibility', text='Эксплуатационная ответственность')
        self.db_tree.heading('mpi', text='МПИ')
        self.db_tree.heading('search_query', text='Запрос')
        self.db_tree.heading('row_index', text='№ строки')

        # Ширина колонок
        self.db_tree.column('vri_id', width=75)
        self.db_tree.column('id_pu', width=85)
        self.db_tree.column('mit_number', width=70)
        self.db_tree.column('mit_title', width=180)
        self.db_tree.column('mit_notation', width=85)
        self.db_tree.column('mi_modification', width=85)
        self.db_tree.column('mi_number', width=85)
        self.db_tree.column('verification_date', width=85)
        self.db_tree.column('valid_date', width=85)
        self.db_tree.column('applicability', width=60)
        self.db_tree.column('org_title', width=120)
        self.db_tree.column('result_docnum', width=95)
        self.db_tree.column('manufacturer', width=95)
        self.db_tree.column('contract_number', width=100)
        self.db_tree.column('edo_code', width=75)
        self.db_tree.column('balance_owner', width=130)
        self.db_tree.column('operation_responsibility', width=160)
        self.db_tree.column('mpi', width=45)
        self.db_tree.column('search_query', width=110)
        self.db_tree.column('row_index', width=50)

        # Скроллы
        v_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.db_tree.yview)
        h_scroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.db_tree.xview)
        self.db_tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.db_tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        self.setup_tree_context_menu(self.db_tree)

    def _show_tooltip(self, widget, text):
        """Показ всплывающей подсказки"""
        self.tooltip = tk.Toplevel(widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry("+%d+%d" % (widget.winfo_rootx(), widget.winfo_rooty() + widget.winfo_height()))
        
        label = ttk.Label(
            self.tooltip, 
            text=text, 
            background="#FFFFE0", 
            relief=tk.SOLID, 
            borderwidth=1,
            wraplength=300,
            justify=tk.LEFT
        )
        label.pack(ipadx=5, ipady=3)
    
    def _hide_tooltip(self):
        """Скрытие всплывающей подсказки"""
        if hasattr(self, 'tooltip') and self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None

    def setup_tree_context_menu(self, tree):
        """Контекстное меню с дополнительными опциями"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Копировать строку", command=lambda: self.copy_selected(tree))
        menu.add_command(label="Копировать все", command=lambda: self.copy_all(tree))
        menu.add_separator()
        menu.add_command(label="Открыть в браузере", command=lambda: self.open_in_browser(tree))
        menu.add_separator()
        menu.add_command(label="Экспорт выбранного", command=lambda: self.export_selected(tree))

        def show_menu(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                menu.post(event.x_root, event.y_root)

        tree.bind("<Button-3>", show_menu)
        tree.bind("<Double-1>", lambda e: self.open_in_browser(tree))

    # ========== Методы работы с контекстным меню ==========
    def copy_selected(self, tree):
        """Копирование выбранной строки"""
        item = tree.selection()[0]
        values = tree.item(item, 'values')
        self.root.clipboard_clear()
        self.root.clipboard_append('\t'.join(map(str, values)))
        logger.debug("Строка скопирована в буфер")

    def copy_all(self, tree):
        """Копирование всех строк"""
        lines = []
        for item in tree.get_children():
            lines.append('\t'.join(map(str, tree.item(item, 'values'))))
        self.root.clipboard_clear()
        self.root.clipboard_append('\n'.join(lines))
        logger.debug(f"Скопировано {len(lines)} строк")

    def open_in_browser(self, tree):
        """Открытие записи в браузере"""
        item = tree.selection()[0]
        values = tree.item(item, 'values')
        if values:
            vri_id = values[0]
            url = f"https://fgis.gost.ru/fundmetrology/cm/erts/?id={vri_id}"
            logger.info(f"Открытие в браузере: {url}")
            webbrowser.open(url)

    def export_selected(self, tree):
        """Экспорт выбранных записей"""
        items = tree.selection()
        if not items:
            messagebox.showwarning("Предупреждение", "Выберите записи для экспорта")
            return
        
        data = []
        for item in items:
            values = tree.item(item, 'values')
            # Получаем полные данные из БД
            vri_id = values[0]
            results = self.db.search(query=vri_id, query_type='exact', limit=1)
            if results:
                data.extend(results)
        
        if data:
            filename = filedialog.asksaveasfilename(
                title="Сохранить выбранные записи",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx")]
            )
            if filename:
                if filename.endswith('.xlsx') and PANDAS_AVAILABLE:
                    Exporter.to_excel(data, filename)
                else:
                    Exporter.to_csv(data, filename)
                messagebox.showinfo("Экспорт", f"Экспортировано {len(data)} записей")

    def export_results_csv(self):
        """Экспорт результатов из вкладки Результаты в CSV"""
        data = self.get_db_data()
        if not data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return

        filename = filedialog.asksaveasfilename(
            title="Экспорт в CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")]
        )
        if filename:
            Exporter.to_csv(data, filename)
            messagebox.showinfo("Экспорт", f"Экспортировано {len(data)} записей в {filename}")

    def export_results_excel(self):
        """Экспорт результатов из вкладки Результаты в Excel"""
        if not PANDAS_AVAILABLE:
            messagebox.showwarning("Предупреждение", "pandas не установлен")
            return

        data = self.get_db_data()
        if not data:
            messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
            return

        filename = filedialog.asksaveasfilename(
            title="Экспорт в Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filename:
            Exporter.to_excel(data, filename)
            messagebox.showinfo("Экспорт", f"Экспортировано {len(data)} записей в {filename}")

    def clear_database(self):
        """Очистка базы данных"""
        if not messagebox.askyesno("Подтверждение", 
            "Вы уверены, что хотите очистить всю базу данных?\n\nЭто действие нельзя отменить!"):
            return

        try:
            with sqlite3.connect(self.db.db_path) as conn:
                conn.execute('DELETE FROM verification_records')
                conn.commit()
            messagebox.showinfo("Очистка", "База данных очищена")
            self.load_db_data()
            self.update_stats()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось очистить БД:\n{e}")

    def get_db_data(self):
        """Получение данных из таблицы Результаты (20 колонок)"""
        data = []
        for item in self.db_tree.get_children():
            values = self.db_tree.item(item, 'values')
            if values and len(values) >= 20:
                data.append({
                    'vri_id': values[0],                     # ID
                    'id_pu': values[1],                      # Id_ПУ
                    'mit_number': values[2],                 # № реестра
                    'mit_title': values[3],                  # Наименование
                    'mit_notation': values[4],               # Обозначение
                    'mi_modification': values[5],            # Модиф.
                    'mi_number': values[6],                  # Зав. №
                    'verification_date': values[7],          # Дата поверки
                    'valid_date': values[8],                 # Действ. до
                    'applicability': values[9],              # Пригоден
                    'org_title': values[10],                 # Поверитель
                    'result_docnum': values[11],             # № док.
                    'manufacturer': values[12],              # Производитель
                    'contract_number': values[13],           # Номер договора
                    'edo_code': values[14],                  # Код ЭДО
                    'balance_owner': values[15],             # Балансовая принадлежность
                    'operation_responsibility': values[16],  # Эксплуатационная ответственность
                    'mpi': values[17],                       # МПИ
                    'search_query': values[18],              # Запрос
                    'row_index': values[19]                  # № строки
                })
        return data

    # ========== Методы пакетного поиска ==========
    def select_batch_file(self):
        """Выбор Excel файла с отображением данных"""
        filename = filedialog.askopenfilename(
            title="Выберите файл Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.batch_file_var.set(filename)
            try:
                if PANDAS_AVAILABLE:
                    # Чтение файла для предпросмотра
                    df = pd.read_excel(filename, dtype=str)
                    logger.info(f"Прочитано {len(df)} строк, колонки: {list(df.columns)}")
                    
                    # Сохраняем все колонки для динамического отображения
                    self.batch_all_columns = list(df.columns)

                    # Чтение запросов
                    queries = ExcelHandler.read_queries(filename)
                    self.batch_queries = queries

                    # Отображение статистики
                    self.batch_processed_var.set(f"Загружено: {len(queries)} запросов")
                    self.batch_found_var.set(f"Строк в файле: {len(df)}")
                    self.batch_filtered_var.set("Готов к поиску")
                    logger.info(f"Загружено {len(queries)} запросов из {filename}")

                    # Очистка и заполнение таблицы предпросмотра
                    for item in self.batch_preview_tree.get_children():
                        self.batch_preview_tree.delete(item)

                    # Показываем первые 100 записей для предпросмотра
                    for i, query in enumerate(queries[:100]):
                        row_index = query.get('row_index', i + 1)
                        original = query.get('original_data', {})

                        # Формируем значения для всех колонок
                        values = [row_index]
                        
                        # Основные поля
                        id_pu = original.get('Id_ПУ', '') or original.get('id_пу', '') or '-'
                        values.append(id_pu[:20] if id_pu else '-')
                        
                        search_term = query.get('mi_number', '') or query.get('search_term', '')
                        values.append(search_term[:20] if search_term else '-')
                        
                        mit_title = original.get('Модель', '') or original.get('Тип ПУ', '') or ''
                        values.append(mit_title[:40] if mit_title else '-')
                        
                        mit_number = original.get('Тип ПУ', '') or ''
                        values.append(mit_number[:20] if mit_number else '-')
                        
                        year = original.get('Год выпуска', '') or original.get('Год', '') or '-'
                        values.append(year)
                        
                        # Дополнительные поля из Excel
                        values.append(original.get('Номер договора', original.get('contract_number', '-')))
                        values.append(original.get('Код ЭДО', original.get('edo_code', '-')))
                        values.append(original.get('Балансовая принадлежность', original.get('balance_owner', '-')))
                        values.append(original.get('Эксплуатационная ответственность', original.get('operation_responsibility', '-')))
                        values.append(original.get('МПИ', original.get('mpi', '-')))

                        self.batch_preview_tree.insert('', tk.END, values=values)

                    if len(queries) > 100:
                        placeholder = [f'... еще {len(queries) - 100}'] + ['-'] * 10
                        self.batch_preview_tree.insert('', tk.END, values=placeholder)

                    # Сообщение если пусто
                    if not queries:
                        messagebox.showwarning(
                            "Предупреждение",
                            "Файл загружен, но не найдено поисковых запросов.\n\n"
                            "Убедитесь, что в файле есть колонка 'Серийный номер'."
                        )

                else:
                    messagebox.showwarning("Предупреждение", "pandas не установлен")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось прочитать файл:\n{e}")
                logger.error(f"Ошибка чтения файла: {e}", exc_info=True)

    def select_all_years_dialog(self):
        """Диалог выбора всех лет"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Выбор лет")
        dialog.geometry("400x300")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Выберите годы:").pack(pady=5)

        # Фрейм с чекбоксами
        years_frame = ttk.Frame(dialog)
        years_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # Сетка чекбоксов
        for i, year in enumerate(Config.YEARS):
            if year not in self.year_vars:
                self.year_vars[year] = tk.BooleanVar(value=True)
            cb = ttk.Checkbutton(years_frame, text=str(year), variable=self.year_vars[year])
            cb.grid(row=i//5, column=i%5, sticky=tk.W, padx=5, pady=2)

        # Кнопки
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)

        def select_all():
            for var in self.year_vars.values():
                var.set(True)
            dialog.destroy()

        def deselect_all():
            for var in self.year_vars.values():
                var.set(False)
            dialog.destroy()

        ttk.Button(btn_frame, text="Выбрать все", command=select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Снять все", command=deselect_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Отмена", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def start_batch_search(self):
        """Запуск пакетного поиска"""
        if not self.batch_queries:
            messagebox.showwarning("Предупреждение", "Загрузите файл с запросами")
            return

        selected_years = [year for year, var in self.year_vars.items() if var.get()]
        if not selected_years:
            messagebox.showwarning("Предупреждение", "Выберите хотя бы один год")
            return

        self.batch_running = True
        self.batch_stop_requested = False
        self.batch_start_btn.config(state=tk.DISABLED)
        self.batch_stop_btn.config(state=tk.NORMAL)

        self.batch_results = []

        # Запуск в потоке
        thread = threading.Thread(target=self._batch_search_thread, args=(selected_years,))
        thread.daemon = True
        thread.start()

    def _batch_search_thread(self, years):
        """Поток пакетного поиска с сохранением связей и фильтрацией по электросчетчикам"""
        try:
            total_queries = len(self.batch_queries)
            logger.info(f"Запуск пакетного поиска: {total_queries} запросов, годы={years}")
            found_count = 0
            filtered_count = 0  # Счетчик отфильтрованных записей

            for i, query_data in enumerate(self.batch_queries):
                if self.batch_stop_requested:
                    break

                # Получение поискового запроса с сохранением связи
                search_term = query_data.get('mi_number', '') or query_data.get('search_term', '')
                row_index = query_data.get('row_index', i + 1)

                if not search_term:
                    continue

                logger.debug(f"Запрос {i+1}/{total_queries}: '{search_term}' (строка {row_index})")

                # Поиск по каждому году
                for year in years:
                    if self.batch_stop_requested:
                        break

                    response = self.api_client.search_vri(search_term=search_term, year=year)
                    items = response.get('result', {}).get('items', [])

                    for item in items:
                        # ФИЛЬТРАЦИЯ: проверяем, относится ли к электросчетчикам
                        mit_title = item.get('mit_title', '')
                        if not VerificationRecord.is_electric_meter(mit_title):
                            filtered_count += 1
                            logger.debug(f"Отфильтровано (не электросчетчик): {mit_title}")
                            continue

                        # Получение Id_ПУ из исходных данных
                        original_data = query_data.get('original_data', {})
                        id_pu = original_data.get('Id_ПУ', '') or original_data.get('id_пу', '')

                        # Создание записи с сохранением всех связей
                        record = VerificationRecord.from_api_response(
                            item,
                            search_query=search_term,
                            row_index=row_index,
                            id_pu=id_pu,
                            original_data=original_data
                        )
                        self.batch_results.append(record)
                        found_count += 1

                        # Добавление в таблицу на вкладке Результаты
                        self.root.after(0, lambda r=record: self._add_result_to_db_tree(r))

                    # Прогресс и статистика
                    progress = int((i + 1) / total_queries * 100)
                    self.root.after(0, lambda p=progress: self.batch_progress.configure(value=p))
                    self.root.after(0, lambda c=i+1, t=total_queries: self.batch_processed_var.set(
                        f"Обработано: {c} / {t}"
                    ))
                    self.root.after(0, lambda f=found_count: self.batch_found_var.set(
                        f"Найдено: {f}"
                    ))
                    self.root.after(0, lambda fl=filtered_count: self.batch_filtered_var.set(
                        f"Отфильтровано: {fl}"
                    ))

                    time.sleep(0.1)  # Пауза между запросами

            # Сохранение в БД
            if self.batch_results:
                saved = self.db.save_records(self.batch_results)
                self.root.after(0, lambda: self.batch_processed_var.set(f"✅ Завершено"))
                self.root.after(0, lambda: self.batch_found_var.set(f"Найдено: {len(self.batch_results)}"))
                self.root.after(0, lambda: self.batch_filtered_var.set(f"Сохранено: {saved}"))
                logger.info(f"✅ Завершено. Найдено {len(self.batch_results)}, сохранено {saved} новых, отфильтровано {filtered_count}")
                # Переключаемся на вкладку Результаты
                self.root.after(0, lambda: self.notebook.select(2))
            else:
                self.root.after(0, lambda: self.batch_processed_var.set("⚠ Завершено"))
                self.root.after(0, lambda: self.batch_found_var.set("Результатов нет"))

        except Exception as e:
            error_msg = f"Ошибка: {e}"
            self.root.after(0, lambda m=error_msg: self.batch_status_var.set(m))
            logger.error(f"Ошибка в пакетном поиске: {e}", exc_info=True)

        finally:
            self.root.after(0, self._finish_batch_search)

    def _add_result_to_db_tree(self, record: VerificationRecord):
        """Добавление результата в таблицу на вкладке Результаты (20 полей)"""
        title = record.mit_title[:60] + '...' if len(record.mit_title) > 60 else record.mit_title
        ver_date = record.verification_date[:10] if record.verification_date and len(record.verification_date) >= 10 else ''
        valid_date = record.valid_date[:10] if record.valid_date and len(record.valid_date) >= 10 else ''

        self.db_tree.insert('', tk.END, values=(
            record.vri_id,                              # 0: ID
            record.id_pu,                               # 1: Id_ПУ
            record.mit_number,                          # 2: № реестра
            title,                                      # 3: Наименование
            record.mit_notation,                        # 4: Обозначение
            record.mi_modification,                     # 5: Модиф.
            record.mi_number,                           # 6: Зав. №
            ver_date,                                   # 7: Дата поверки
            valid_date,                                 # 8: Действ. до
            'Да' if record.applicability else 'Нет',   # 9: Пригоден
            record.org_title,                           # 10: Поверитель
            record.result_docnum,                       # 11: № док.
            record.manufacturer,                        # 12: Производитель
            record.contract_number,                     # 13: Номер договора
            record.edo_code,                            # 14: Код ЭДО
            record.balance_owner,                       # 15: Балансовая принадлежность
            record.operation_responsibility,            # 16: Эксплуатационная ответственность
            record.mpi,                                 # 17: МПИ
            record.search_query,                        # 18: Запрос
            record.row_index                            # 19: № строки
        ))

    def _finish_batch_search(self):
        """Завершение пакетного поиска"""
        self.batch_running = False
        self.batch_start_btn.config(state=tk.NORMAL)
        self.batch_stop_btn.config(state=tk.DISABLED)
        self.update_stats()

    def stop_batch_search(self):
        """Остановка пакетного поиска"""
        self.batch_stop_requested = True
        self.batch_status_var.set("Остановка...")
        logger.info("Остановка пакетного поиска по запросу пользователя")

    # ========== Методы одиночного поиска ==========
    def single_search(self):
        """Одиночный поиск с сохранением результатов и проверкой на дубли"""
        search_term = self.search_entry.get().strip()
        if not search_term:
            messagebox.showwarning("Предупреждение", "Введите заводской номер")
            return

        # Получение выбранных годов из чекбоксов
        selected_years = [year for year, var in self.search_year_vars.items() if var.get()]
        if not selected_years:
            messagebox.showwarning("Предупреждение", "Выберите хотя бы один год")
            return

        # Блокировка кнопок
        self.search_running = True
        self.search_stop_requested = False
        self.search_start_btn.config(state=tk.DISABLED)
        self.search_stop_btn.config(state=tk.NORMAL)

        start_time = time.time()

        # Очистка таблицы
        for item in self.single_results_tree.get_children():
            self.single_results_tree.delete(item)

        # Поиск через API по всем выбранным годам (получаем все результаты, до 500)
        logger.info(f"Поиск: '{search_term}', годы={selected_years}")
        all_items = []
        self.single_search_results_all = []  # Сохраняем все оригинальные записи

        for year in selected_years:
            if self.search_stop_requested:
                self.search_info_var.set("⏹ Поиск остановлен пользователем")
                break
                
            self.search_info_var.set(f"🔍 Поиск... год={year}, найдено={len(all_items)}")
            self.root.update_idletasks()
            
            start = 0
            rows = 100

            while True:
                if self.search_stop_requested:
                    break
                    
                response = self.api_client.search_vri(
                    search_term=search_term,
                    year=year,
                    start=start,
                    rows=rows
                )
                items = response.get('result', {}).get('items', [])
                count = response.get('result', {}).get('count', 0)

                if not items:
                    break

                all_items.extend(items)
                self.single_search_results_all.extend(items)
                start += rows

                if start >= count or len(all_items) >= 500:
                    break

        elapsed_ms = int((time.time() - start_time) * 1000)

        # Разблокировка кнопок
        self.search_running = False
        self.search_start_btn.config(state=tk.NORMAL)
        self.search_stop_btn.config(state=tk.DISABLED)

        # Сохраняем результаты для последующего экспорта/добавления в БД
        self.single_search_results = all_items

        # Отображение всех результатов (10 колонок)
        for item in all_items:
            self.single_results_tree.insert('', tk.END, values=(
                item.get('vri_id', ''),                              # 0: ID записи
                item.get('mit_number', ''),                          # 1: № в реестре
                item.get('mit_title', '')[:100],                     # 2: Наименование СИ
                item.get('mit_notation', ''),                        # 3: Обозначение типа
                item.get('mi_modification', ''),                     # 4: Модификация
                item.get('mi_number', ''),                           # 5: Зав. номер
                item.get('verification_date', '')[:10] if item.get('verification_date') else '',  # 6: Дата поверки
                item.get('valid_date', '')[:10] if item.get('valid_date') else '',  # 7: Действ. до
                item.get('org_title', '')[:50],                      # 8: Поверитель
                item.get('result_docnum', ''),                       # 9: № документа
                'Да' if item.get('applicability') else 'Нет'        # 10: Пригоден
            ))

        # Логирование в историю (первый год из выбранных)
        self.db.log_search(search_term, self.search_type.get(), selected_years[0] if selected_years else None, len(all_items), elapsed_ms)

        self.single_status_var.set(f"Найдено {len(all_items)} за {elapsed_ms}мс")
        
        if self.search_stop_requested:
            self.search_info_var.set(f"⏹ Остановлено. Найдено: {len(all_items)}")
        else:
            self.search_info_var.set(f"✅ Найдено: {len(all_items)} за {elapsed_ms}мс")

        # Сообщение с опциями
        msg = f"Найдено {len(all_items)} записей\nВремя: {elapsed_ms}мс"
        
        # Создаем диалоговое окно с кнопками
        dialog = tk.Toplevel(self.root)
        dialog.title("Поиск завершён")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text=msg, justify=tk.CENTER).pack(pady=20)
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        def save_csv():
            self.save_single_search_csv(all_items, search_term)
            dialog.destroy()
        
        def add_to_db():
            self.add_single_search_to_db(all_items, search_term)
            dialog.destroy()
        
        def close():
            dialog.destroy()
        
        ttk.Button(btn_frame, text="💾 Сохранить в CSV", command=save_csv, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="➕ Добавить в БД", command=add_to_db, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="✖ Закрыть", command=close, width=10).pack(side=tk.LEFT, padx=5)

    def save_single_search_csv(self, items: List[Dict], search_term: str):
        """Сохранение результатов одиночного поиска в CSV"""
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для сохранения")
            return
        
        filename = filedialog.asksaveasfilename(
            title="Сохранить результаты поиска",
            defaultextension=".csv",
            initialfile=f"search_{search_term}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            filetypes=[("CSV files", "*.csv")]
        )
        
        if filename:
            try:
                # Конвертируем в формат VerificationRecord для экспорта
                records = [VerificationRecord.from_api_response(item, search_term) for item in items]
                data = [r.to_dict() for r in records]
                
                Exporter.to_csv(data, filename, include_service_fields=True)
                messagebox.showinfo("Экспорт", f"Экспортировано {len(items)} записей в {filename}")
                logger.info(f"Экспорт одиночного поиска: {filename}, {len(items)} записей")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось экспортировать:\n{e}")
                logger.error(f"Ошибка экспорта: {e}")

    def add_single_search_to_db(self, items: List[Dict], search_term: str):
        """Добавление результатов одиночного поиска в БД с проверкой на дубли"""
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для добавления")
            return
        
        # Создаём записи (служебные поля пустые для одиночного поиска)
        records = []
        for i, item in enumerate(items):
            record = VerificationRecord.from_api_response(
                item,
                search_query=search_term,
                row_index=i + 1,
                id_pu="",
                original_data=None
            )
            records.append(record)
        
        # Сохраняем с проверкой на дубли
        saved = self.db.save_records(records)
        duplicates = len(records) - saved
        
        msg = f"Добавлено: {saved} новых записей"
        if duplicates > 0:
            msg += f"\nПропущено дублей: {duplicates}"
        
        messagebox.showinfo("Добавление в БД", msg)
        logger.info(f"Одиночный поиск в БД: добавлено {saved}, дубликатов {duplicates}")

        # Обновляем статистику
        self.update_stats()

    # ========== Методы управления поиском ==========
    def stop_single_search(self):
        """Остановка одиночного поиска"""
        self.search_stop_requested = True
        self.search_info_var.set("⏹ Остановка поиска...")
        logger.info("Остановка одиночного поиска по запросу пользователя")

    def clear_single_search(self):
        """Очистка результатов поиска для нового запроса"""
        self.search_entry.delete(0, tk.END)
        for item in self.single_results_tree.get_children():
            self.single_results_tree.delete(item)
        self.single_search_results = []
        self.search_info_var.set("Готов к поиску")
        self.single_status_var.set("Готов")
        logger.info("Очистка результатов поиска")

    def delete_selected_rows(self):
        """Удаление выбранных строк из таблицы результатов"""
        selected = self.single_results_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите строки для удаления")
            return

        count = len(selected)
        if messagebox.askyesno("Подтверждение", f"Удалить {count} выбранных строк?"):
            for item in selected:
                self.single_results_tree.delete(item)
            
            # Обновляем internal список результатов
            self.single_search_results = []
            for item in self.single_results_tree.get_children():
                values = self.single_results_tree.item(item, 'values')
                if values:
                    # Находим соответствующую запись в original_items
                    for orig_item in self.single_search_results_all:
                        if orig_item.get('vri_id') == values[0]:
                            self.single_search_results.append(orig_item)
                            break
            
            self.search_info_var.set(f"Удалено: {count} строк")
            logger.info(f"Удалено {count} строк из результатов поиска")

    def save_single_search_csv_from_tree(self):
        """Сохранение текущих результатов из таблицы в CSV"""
        items = self.single_search_results
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для сохранения")
            return
        
        search_term = self.search_entry.get().strip() or "search"
        self.save_single_search_csv(items, search_term)

    def add_single_search_to_db_from_tree(self):
        """Добавление текущих результатов из таблицы в БД"""
        items = self.single_search_results
        if not items:
            messagebox.showwarning("Предупреждение", "Нет данных для добавления")
            return
        
        search_term = self.search_entry.get().strip() or "search"
        self.add_single_search_to_db(items, search_term)

    # ========== Методы работы с БД ==========
    def load_db_data(self):
        """Загрузка данных из БД с фильтрами (20 колонок)"""
        search = self.db_search_var.get().strip()
        year = self.db_year_var.get()
        year = int(year) if year.isdigit() else None
        manufacturer = self.db_manufacturer_var.get().strip()

        # Очистка таблицы
        for item in self.db_tree.get_children():
            self.db_tree.delete(item)

        # Загрузка
        results = self.db.search(search, 'substring' if search else 'exact', year, manufacturer, limit=500)

        for row in results:
            self.db_tree.insert('', tk.END, values=(
                row.get('vri_id', ''),                     # 0: ID
                row.get('id_pu', ''),                      # 1: Id_ПУ
                row.get('mit_number', ''),                 # 2: № реестра
                row.get('mit_title', '')[:80],             # 3: Наименование
                row.get('mit_notation', ''),               # 4: Обозначение
                row.get('mi_modification', ''),            # 5: Модиф.
                row.get('mi_number', ''),                  # 6: Зав. №
                row.get('verification_date', '')[:10] if row.get('verification_date') else '',  # 7: Дата поверки
                row.get('valid_date', '')[:10] if row.get('valid_date') else '',  # 8: Действ. до
                'Да' if row.get('applicability', 1) else 'Нет',  # 9: Пригоден
                row.get('org_title', ''),                  # 10: Поверитель
                row.get('result_docnum', ''),              # 11: № док.
                row.get('manufacturer', ''),               # 12: Производитель
                row.get('contract_number', ''),            # 13: Номер договора
                row.get('edo_code', ''),                   # 14: Код ЭДО
                row.get('balance_owner', ''),              # 15: Балансовая принадлежность
                row.get('operation_responsibility', ''),   # 16: Эксплуатационная ответственность
                row.get('mpi', ''),                        # 17: МПИ
                row.get('search_query', ''),               # 18: Запрос
                row.get('row_index', '')                   # 19: № строки
            ))

        # Обновление списка производителей
        with sqlite3.connect(self.db.db_path) as conn:
            cursor = conn.execute('SELECT DISTINCT manufacturer FROM verification_records ORDER BY manufacturer')
            manufacturers = [row[0] for row in cursor.fetchall() if row[0]]

        self.db_manufacturer_combo['values'] = [''] + manufacturers
        logger.debug(f"Загружено {len(results)} записей из БД")

    def reset_db_filters(self):
        """Сброс фильтров"""
        self.db_search_var.set('')
        self.db_year_var.set('')
        self.db_manufacturer_var.set('')
        self.load_db_data()

    # ========== Методы статистики и экспорта ==========
    def update_stats(self):
        """Обновление статистики"""
        stats = self.db.get_stats()

        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete(1.0, tk.END)

        self.stats_text.insert(tk.END, "📊 СТАТИСТИКА БАЗЫ ДАННЫХ\n")
        self.stats_text.insert(tk.END, "="*50 + "\n")
        self.stats_text.insert(tk.END, f"Всего записей: {stats['total']}\n")
        self.stats_text.insert(tk.END, f"Пригодных: {stats['applicable']} ({stats['applicable']/stats['total']*100:.1f}%)\n" if stats['total'] else "Пригодных: 0\n")
        self.stats_text.insert(tk.END, f"Непригодных: {stats['inapplicable']}\n\n")

        if stats['by_year']:
            self.stats_text.insert(tk.END, "📅 ПО ГОДАМ:\n")
            for year, count in sorted(stats['by_year'].items()):
                pct = (count/stats['total']*100) if stats['total'] else 0
                self.stats_text.insert(tk.END, f"  {year}: {count} ({pct:.1f}%)\n")
            self.stats_text.insert(tk.END, "\n")

        if stats['by_manufacturer']:
            self.stats_text.insert(tk.END, "🏭 ПО ПРОИЗВОДИТЕЛЯМ (ТОП-15):\n")
            for man, count in list(stats['by_manufacturer'].items())[:15]:
                pct = (count/stats['total']*100) if stats['total'] else 0
                self.stats_text.insert(tk.END, f"  {man}: {count} ({pct:.1f}%)\n")

        self.stats_text.config(state=tk.DISABLED)
        logger.debug("Статистика обновлена")

    def clean_duplicates(self):
        """Очистка дубликатов"""
        deleted = self.db.clean_duplicates()
        messagebox.showinfo("Дедупликация", f"Удалено дубликатов: {deleted}")
        self.update_stats()

    def export_csv(self):
        """Экспорт в CSV с полной служебной информацией"""
        filename = filedialog.asksaveasfilename(
            title="Сохранить как CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            try:
                with sqlite3.connect(self.db.db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.execute('SELECT * FROM verification_records ORDER BY verification_date DESC')
                    data = [dict(row) for row in cursor.fetchall()]

                Exporter.to_csv(data, filename, include_service_fields=True)
                messagebox.showinfo("Экспорт", f"Данные экспортированы в {filename}\nВключая служебные поля (Id_ПУ, № строки, Поисковый запрос)")
                logger.info(f"Экспорт в CSV: {filename}, {len(data)} записей")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось экспортировать: {e}")
                logger.error(f"Ошибка экспорта: {e}")

    def export_excel(self):
        """Экспорт в Excel"""
        if not PANDAS_AVAILABLE:
            messagebox.showwarning("Предупреждение", "pandas не установлен. Используйте CSV.")
            return

        filename = filedialog.asksaveasfilename(
            title="Сохранить как Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            try:
                with sqlite3.connect(self.db.db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.execute('SELECT * FROM verification_records ORDER BY verification_date DESC')
                    data = [dict(row) for row in cursor.fetchall()]

                Exporter.to_excel(data, filename, include_service_fields=True)
                messagebox.showinfo("Экспорт", f"Данные экспортированы в {filename}\nURL генерируется формулой из vri_id")
                logger.info(f"Экспорт в Excel: {filename}, {len(data)} записей")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось экспортировать: {e}")
                logger.error(f"Ошибка экспорта: {e}")

    def export_stats_json(self):
        """Экспорт статистики"""
        stats = self.db.get_stats()
        filename = Exporter.export_stats(stats)
        messagebox.showinfo("Экспорт", f"Статистика экспортирована в {filename}")

    # ========== Вспомогательные методы ==========
    def show_search_history(self):
        """Показ истории поиска"""
        history = self.db.get_search_history(50)
        
        dialog = tk.Toplevel(self.root)
        dialog.title("История поиска")
        dialog.geometry("700x400")
        dialog.transient(self.root)

        text = scrolledtext.ScrolledText(dialog, height=20, font=('Consolas', 10))
        text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        for record in history:
            line = f"{record['search_date'][:19]} | {record['query'][:30]:30} | {record['results_count']:4} рез. | {record['duration_ms']:4}мс\n"
            text.insert(tk.END, line)

        text.config(state=tk.DISABLED)

    def show_log_file(self):
        """Показ файла логов"""
        if os.path.exists(LOG_FILE):
            os.startfile(LOG_FILE) if sys.platform == 'win32' else os.system(f'xdg-open "{LOG_FILE}"')
        else:
            messagebox.showinfo("Лог файл", f"Файл не найден: {LOG_FILE}")

    # ========== Методы управления фильтрами ==========
    def load_exact_queries_display(self):
        """Загрузка и отображение точных запросов из стандартного файла"""
        try:
            with open(Config.EXACT_QUERIES_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
            self.exact_queries_list.config(state=tk.NORMAL)
            self.exact_queries_list.delete(1.0, tk.END)
            self.exact_queries_list.insert(tk.END, content)
            self.exact_queries_list.config(state=tk.DISABLED)
            self.filters_status_var.set("✅ Загружено из exact_queries.csv")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл:\n{e}")
            self.filters_status_var.set("❌ Ошибка загрузки")

    def load_exact_queries_file(self):
        """Загрузка точных запросов из выбранного файла"""
        filename = filedialog.askopenfilename(
            title="Загрузить точные запросы",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    content = f.read()
                self.exact_queries_list.config(state=tk.NORMAL)
                self.exact_queries_list.delete(1.0, tk.END)
                self.exact_queries_list.insert(tk.END, content)
                self.exact_queries_list.config(state=tk.DISABLED)
                self.filters_status_var.set(f"✅ Загружено из {os.path.basename(filename)}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось загрузить файл:\n{e}")

    def save_exact_queries_file(self):
        """Сохранение точных запросов в выбранный файл"""
        filename = filedialog.asksaveasfilename(
            title="Сохранить точные запросы",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            try:
                content = self.exact_queries_list.get(1.0, tk.END)
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(content.strip() + '\n')
                messagebox.showinfo("Сохранено", f"Файл сохранен:\n{filename}")
                self.filters_status_var.set(f"✅ Сохранено в {os.path.basename(filename)}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{e}")

    def clear_exact_queries(self):
        """Очистка списка точных запросов"""
        if messagebox.askyesno("Подтверждение", "Очистить список точных запросов?"):
            self.exact_queries_list.config(state=tk.NORMAL)
            self.exact_queries_list.delete(1.0, tk.END)
            self.exact_queries_list.config(state=tk.DISABLED)
            self.filters_status_var.set("🗑 Очищено")

    def add_exact_query(self):
        """Добавление нового запроса"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Добавить запрос")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Запрос:").pack(pady=5)
        query_entry = ttk.Entry(dialog, width=50)
        query_entry.pack(pady=5)

        ttk.Label(dialog, text="Категория (electric/gas/water/other):").pack(pady=5)
        cat_entry = ttk.Entry(dialog, width=20)
        cat_entry.insert(0, "electric")
        cat_entry.pack(pady=5)

        ttk.Label(dialog, text="Приоритет (high/medium/low):").pack(pady=5)
        prio_entry = ttk.Entry(dialog, width=20)
        prio_entry.insert(0, "medium")
        prio_entry.pack(pady=5)

        def save():
            query = query_entry.get().strip()
            cat = cat_entry.get().strip()
            prio = prio_entry.get().strip()
            if query:
                new_line = f"\n{query},{cat},{prio}"
                self.exact_queries_list.config(state=tk.NORMAL)
                self.exact_queries_list.insert(tk.END, new_line)
                self.exact_queries_list.config(state=tk.DISABLED)
                dialog.destroy()
                self.filters_status_var.set("➕ Добавлено")

        ttk.Button(dialog, text="Добавить", command=save).pack(pady=10)

    def load_manufacturers_display(self):
        """Загрузка и отображение производителей из стандартного файла"""
        try:
            with open(Config.MANUFACTURERS_FILE, 'r', encoding='utf-8') as f:
                content = f.read()
            self.manufacturers_list.config(state=tk.NORMAL)
            self.manufacturers_list.delete(1.0, tk.END)
            self.manufacturers_list.insert(tk.END, content)
            self.manufacturers_list.config(state=tk.DISABLED)
            self.filters_status_var.set("✅ Загружено из manufacturers.csv")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить файл:\n{e}")
            self.filters_status_var.set("❌ Ошибка загрузки")

    def load_manufacturers_file(self):
        """Загрузка производителей из выбранного файла"""
        filename = filedialog.askopenfilename(
            title="Загрузить производителей",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    content = f.read()
                self.manufacturers_list.config(state=tk.NORMAL)
                self.manufacturers_list.delete(1.0, tk.END)
                self.manufacturers_list.insert(tk.END, content)
                self.manufacturers_list.config(state=tk.DISABLED)
                self.filters_status_var.set(f"✅ Загружено из {os.path.basename(filename)}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось загрузить файл:\n{e}")

    def save_manufacturers_file(self):
        """Сохранение производителей в выбранный файл"""
        filename = filedialog.asksaveasfilename(
            title="Сохранить производителей",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            try:
                content = self.manufacturers_list.get(1.0, tk.END)
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(content.strip() + '\n')
                messagebox.showinfo("Сохранено", f"Файл сохранен:\n{filename}")
                self.filters_status_var.set(f"✅ Сохранено в {os.path.basename(filename)}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{e}")

    def clear_manufacturers(self):
        """Очистка списка производителей"""
        if messagebox.askyesno("Подтверждение", "Очистить список производителей?"):
            self.manufacturers_list.config(state=tk.NORMAL)
            self.manufacturers_list.delete(1.0, tk.END)
            self.manufacturers_list.config(state=tk.DISABLED)
            self.filters_status_var.set("🗑 Очищено")

    def add_manufacturer(self):
        """Добавление нового производителя"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Добавить производителя")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Ключевое слово (в нижнем регистре):").pack(pady=5)
        kw_entry = ttk.Entry(dialog, width=50)
        kw_entry.pack(pady=5)

        ttk.Label(dialog, text="Название производителя:").pack(pady=5)
        mfg_entry = ttk.Entry(dialog, width=50)
        mfg_entry.pack(pady=5)

        ttk.Label(dialog, text="Категория (electric/other):").pack(pady=5)
        cat_entry = ttk.Entry(dialog, width=20)
        cat_entry.insert(0, "electric")
        cat_entry.pack(pady=5)

        ttk.Label(dialog, text="Приоритет (high/medium/low):").pack(pady=5)
        prio_entry = ttk.Entry(dialog, width=20)
        prio_entry.insert(0, "medium")
        prio_entry.pack(pady=5)

        def save():
            kw = kw_entry.get().strip().lower()
            mfg = mfg_entry.get().strip()
            cat = cat_entry.get().strip()
            prio = prio_entry.get().strip()
            if kw and mfg:
                new_line = f"\n{kw},{mfg},{cat},{prio}"
                self.manufacturers_list.config(state=tk.NORMAL)
                self.manufacturers_list.insert(tk.END, new_line)
                self.manufacturers_list.config(state=tk.DISABLED)
                dialog.destroy()
                self.filters_status_var.set("➕ Добавлено")

        ttk.Button(dialog, text="Добавить", command=save).pack(pady=10)

    def delete_manufacturer(self):
        """Удаление выбранного производителя"""
        try:
            sel = self.manufacturers_list.tag_ranges(tk.SEL)
            if not sel:
                messagebox.showwarning("Предупреждение", "Выделите строку для удаления")
                return
            self.manufacturers_list.config(state=tk.NORMAL)
            self.manufacturers_list.delete(tk.SEL_FIRST, tk.SEL_LAST)
            self.manufacturers_list.config(state=tk.DISABLED)
            self.filters_status_var.set("🗑 Удалено")
        except:
            pass

    def show_about(self):
        """О программе"""
        about_text = """
        ФГИС Аршин - Поиск и сбор данных
        Версия 2.1 (оптимизированная)

        Возможности:
        • Поиск с использованием файлов конфигурации
        • Пакетный поиск из Excel с сохранением связей
        • Промышленный сбор данных (асинхронный)
        • Экспорт с полной служебной информацией
        • Расширенное логирование для отладки
        • Оптимизированный интерфейс для 1200x700

        API: https://fgis.gost.ru/fundmetrology/eapi/
        Лог файл: {}
        """.format(LOG_FILE)
        
        messagebox.showinfo("О программе", about_text)

    def on_close(self):
        """Обработка закрытия"""
        if self.batch_running:
            if not messagebox.askyesno("Подтверждение", "Пакетный поиск выполняется. Завершить?"):
                return
            self.batch_stop_requested = True

        if self.async_running:
            if not messagebox.askyesno("Подтверждение", "Сбор данных выполняется. Завершить?"):
                return
            self.collector.stop()

        self.api_client.close()
        logger.info("Приложение закрыто")
        self.root.destroy()


# ============== ЗАПУСК ==============
def main():
    """Точка входа"""
    # Создание директорий
    os.makedirs(Config.EXPORT_DIR, exist_ok=True)
    os.makedirs(Config.CONFIG_DIR, exist_ok=True)
    os.makedirs("logs", exist_ok=True)

    logger.info("="*60)
    logger.info("ЗАПУСК ПРИЛОЖЕНИЯ ФГИС АРШИН v2.1")
    logger.info(f"Python: {sys.version}")
    logger.info(f"Платформа: {sys.platform}")
    logger.info(f"Лог файл: {LOG_FILE}")
    logger.info("="*60)

    # Запуск GUI
    root = tk.Tk()

    # Применение темы и стилей
    try:
        style = ttk.Style()
        if 'clam' in style.theme_names():
            style.theme_use('clam')
        
        # Настройка цветов кнопок
        style.configure('Green.TButton', 
            background='#90EE90',  # светло-зелёный
            foreground='black',
            font=('TkDefaultFont', 10))
        
        style.map('Green.TButton',
            background=[('active', '#98FB98'), ('pressed', '#32CD32')])
        
        style.configure('Red.TButton',
            background='#FFB6C1',  # светло-красный
            foreground='black',
            font=('TkDefaultFont', 10))
        
        style.map('Red.TButton',
            background=[('active', '#FFC0CB'), ('pressed', '#DC143C')])
    except:
        pass

    app = ArshinApp(root)
    app.load_db_data()
    root.mainloop()


if __name__ == "__main__":
    main()
